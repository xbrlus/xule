"""XuleFunctions

Xule is a rule processor for XBRL (X)brl r(ULE). 

DOCSKIP
See https://xbrl.us/dqc-license for license information.  
See https://xbrl.us/dqc-patent for patent infringement notice.
Copyright (c) 2017 - 2023 XBRL US, Inc.

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

    http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.

$Change: 23454 $
DOCSKIP
"""

from aniso8601 import parse_duration
from arelle.ModelValue import qname, QName
from arelle import FileSource, PackageManager, FunctionIxt
import collections
from collections.abc import Iterable
from contextlib import contextmanager
import datetime
import decimal
import json
import openpyxl
import random
from lxml import etree as et
from .XuleRunTime import XuleProcessingError
from . import XuleValue as xv
from . import XuleRollForward as xrf
from . import XuleInstanceFunctions as xif
from . import XuleUtility as xu


def func_exists(xule_context, *args):   
    #return xv.xv.XuleValue(xule_context, args[0].type not in ('unbound', 'none'), 'bool')
    return xv.XuleValue(xule_context, args[0].type != 'unbound', 'bool')

def func_missing(xule_context, *args):
    #return xv.XuleValue(xule_context, args[0].type in ('unbound', 'none'), 'bool')
    return xv.XuleValue(xule_context, args[0].type == 'unbound', 'bool')

def func_date(xule_context, *args):
    arg = args[0]

    if arg.type == 'instant':
        return arg
    elif arg.type == 'string':
        return xv.XuleValue(xule_context, xv.iso_to_date(xule_context, arg.value), 'instant')
    else:
        raise XuleProcessingError(_("function 'date' requires a string argument, found '%s'" % arg.type), xule_context)

def func_duration(xule_context, *args):
    start = args[0]
    end = args[1]
    
    start_instant = func_date(xule_context, start)
    end_instant = func_date(xule_context, end)
    
    if end_instant.value < start_instant.value:
        return xv.XuleValue(xule_context, None, 'unbound')
    else:
        return xv.XuleValue(xule_context, (start_instant.value, end_instant.value), 'duration', from_model=start.from_model or end.from_model)

def func_forever(xule_context, *args):
    return xv.XuleValue(xule_context, (datetime.datetime.min, datetime.datetime.max), 'duration')

def func_unit(xule_context, *args):
    
    if len(args) == 0 or len(args) > 2:
        raise XuleProcessingError(_("The unit() function takes 1 or 2 arguments, found {}".format(len(args))), xule_context)
    
    return xv.XuleValue(xule_context, xv.XuleUnit(*args), 'unit')

def func_entity(xule_context, *args):
    scheme = args[0]
    identifier = args[1]
    
    if scheme.type != 'string' or identifier.type != 'string':
        raise XuleProcessingError(_("The entity scheme and identifier must be strings. Found '%s' and '%s'" % (scheme.type, identifier.type)), xule_context)
    
    return xv.XuleValue(xule_context, (scheme.value, identifier.value), 'entity')

def func_qname(xule_context, *args):
    namespace_uri_arg = args[0]
    local_name_arg = args[1]
    
    if namespace_uri_arg.type not in ('string', 'uri', 'unbound', 'none'):
        raise XuleProcessingError(_("Function 'qname' requires the namespace_uri argument to be a string, uri or none, found '%s'" % namespace_uri_arg.type), xule_context)
    if local_name_arg.type != 'string':
        raise XuleProcessingError(_("Function 'qname' requires the local_part argument to be a string, found '%s'" % local_name_arg.type), xule_context)

    if namespace_uri_arg.type == 'unbound':
        return xv.XuleValue(xule_context, qname(local_name_arg.value, noPrefixIsNoNamespace=True), 'qname')
    else:
        # get the prefix from the rule file
        prefix = get_prefix(xule_context, namespace_uri_arg.value)
        return xv.XuleValue(xule_context, QName(prefix, namespace_uri_arg.value, local_name_arg.value), 'qname')

def get_prefix(xule_context, uri):
    for k, v in xule_context.global_context.catalog['namespaces'].items():
        if v['uri'] == uri:
            if k == '*':
                return None
            else:
                return k
    return None

def func_uri(xule_context, *args):
    arg = args[0]

    if arg.type == 'string':
        return xv.XuleValue(xule_context, arg.value, 'uri')
    elif arg.type == 'uri':
        return arg
    else:
        raise XuleProcessingError(_("The 'uri' function requires a string argument, found '%s'." % arg.type), xule_context) 

def func_time_span(xule_context, *args):
    arg = args[0]
    
    if arg.type != 'string':
        raise XuleProcessingError(_("Function 'time-span' expects a string, fount '%s'." % arg.type), xule_context)
    
    try:
        return xv.XuleValue(xule_context, parse_duration(arg.value.upper()), 'time-period')
    except:
        raise XuleProcessingError(_("Could not convert '%s' into a time-period." % arg.value), xule_context)

def func_schema_type(xule_context, *args):
    arg = args[0]
    
    if arg.type == 'qname':
        if arg.value in xule_context.model.qnameTypes:
            return xv.XuleValue(xule_context, xule_context.model.qnameTypes[arg.value], 'type')
        else:
            return xv.XuleValue(xule_context, None, 'none')
    else:
        raise XuleProcessingError(_("Function 'schema' expects a qname argument, found '%s'" % arg.type), xule_context)

def func_num_to_string(xule_context, *args):
    arg = args[0]
    
    if arg.type in ('int', 'float', 'decimal'):
        return xv.XuleValue(xule_context, format(arg.value, ","), 'string')
    else:
        raise XuleProcessingError(_("function 'num_to_string' requires a numeric argument, found '%s'" % arg.type), xule_context)
        
def func_mod(xule_context, *args):
    numerator = args[0]
    denominator = args[1]
    
    if numerator.type not in ('int', 'float', 'decimal'):
        raise XuleProcessingError(_("The numerator for the 'mod' function must be numeric, found '%s'" % numerator.type), xule_context) 
    if denominator.type not in ('int', 'float', 'decimal'):
        raise XuleProcessingError(_("The denominator for the 'mod' function must be numeric, found '%s'" % denominator.type), xule_context)
    
    combined_type, numerator_compute_value, denominator_compute_value = xv.combine_xule_types(numerator, denominator, xule_context)
    return xv.XuleValue(xule_context, numerator_compute_value % denominator_compute_value, combined_type)

def func_random(xule_context, *args):
    if len(args) == 0:
        return xv.XuleValue(xule_context, decimal.Decimal(random.random()), 'decimal')
    
    if len(args) >= 1:
        if args[0].type not in ('int', 'decimal', 'float', 'none'):
            raise XuleProcessingError(_("First argument of the random() function must be numeric, found '{}'".format(args[0].type)), xule_context)
        start = args[0].value or 0.0
    else:
        start = 0.0
    
    if len(args) >= 2:
        if args[1].type not in ('int', 'decimal', 'float', 'none'):
            raise XuleProcessingError(_("Second argument of the random() function must be numeric, found '{}'".format(args[1].type)), xule_context)
        end = args[1].value or 1.0
    else:
        end = 1.0

    if len(args) == 3:
        if args[2].type != 'string':
            raise XuleProcessingError(_("Third argument of the random() function must be a string of 'int' or 'decimal', found non string of type '{}'".format(args[2].type)), xule_context)
        if args[2].value not in ('int', 'decimal'):
            raise XuleProcessingError(_("Third argument of the random() function must be a string of 'int' or 'decimal', found '{}''".format(args[2].value)), xule_context)
        result_type = args[2].value
    else:
        result_type = 'decimal'

    if result_type == 'int':
        random_val = round(random.uniform(start, end))
    else:
        random_val = decimal.Decimal(random.uniform(start, end))

    return xv.XuleValue(xule_context, random_val, result_type)

def func_extension_concept(xule_context, *args):   
    extension_ns_value_set = xule_context._const_extension_ns()
    if len(extension_ns_value_set.values) > 0:
        extension_ns = extension_ns_value_set.values[None][0].value
    else:
        raise XuleProcessingError(_("Cannot determine extension namespace."), xule_context)
    
    concepts = set(xv.XuleValue(xule_context, x, 'concept') for x in xule_context.model.qnameConcepts.values() if (x.isItem or x.isTuple) and x.qname.namespaceURI == extension_ns)
    
    return xv.XuleValue(xule_context, frozenset(concepts), 'set')

def agg_count(xule_context, values):
    alignment = values[0].alignment if len(values) > 0 else None
    return_value = xv.XuleValue(xule_context, len(values), 'int', alignment=alignment)
    tags = {}
    facts = collections.OrderedDict()
    
    for current_value in values:
        if current_value.tags is not None:
            tags.update(current_value.tags)
        if current_value.facts is not None:
            facts.update(current_value.facts)
    if len(tags) > 0:
        return_value.tags = tags
    if len(facts) > 0:
        return_value.facts = facts
    return return_value #xv.XuleValue(xule_context, len(values), 'int', alignment=alignment)

def agg_sum(xule_context, values):
    agg_value = values[0].clone()
    tags = {} if agg_value.tags is None else agg_value.tags
    facts = collections.OrderedDict() if agg_value.facts is None else agg_value.facts
    
    for current_value in values[1:]:
        combined_types = xv.combine_xule_types(agg_value, current_value, xule_context)
        if combined_types[0] == 'set':
            agg_value = xv.XuleValue(xule_context, combined_types[1] | combined_types[2], combined_types[0], alignment=agg_value.alignment)
        else:
            agg_value = xv.XuleValue(xule_context, combined_types[1] + combined_types[2], combined_types[0], alignment=agg_value.alignment)
        if current_value.tags is not None:
            tags.update(current_value.tags)
        if current_value.facts is not None:
            facts.update(current_value.facts)
            
    if len(tags) > 0:
        agg_value.tags = tags
    if len(facts) > 0:
        agg_value.facts = facts
    
    return agg_value 

def agg_all(xule_context, values):
    all_value = True
    tags = {}
    facts = collections.OrderedDict()
    
    for current_value in values:
        if current_value.type != 'bool':
            raise XuleProcessingError(_("Function all can only operator on booleans, but found '%s'." % current_value.type), xule_context)
        if current_value.value and current_value.tags is not None:
            tags.update(current_value.tags)
        if current_value.value and current_value.facts is not None:
            facts.update(current_value.facts)      
        all_value = all_value and current_value.value
        if not all_value:
            break
    
    return_value = xv.XuleValue(xule_context, all_value, 'bool')
    if len(tags) > 0:
        return_value.tags = tags
    if len(facts) > 0:
        return_value.facts = facts
        
    return return_value #xv.XuleValue(xule_context, all_value, 'bool')

def agg_any(xule_context, values):
    any_value = False
    tags = {}
    facts = collections.OrderedDict()
    
    for current_value in values:
        if current_value.type != 'bool':
            raise XuleProcessingError(_("Function all can only operator on booleans, but found '%s'." % current_value.type), xule_context)
        if current_value.value and current_value.tags is not None:
            tags.update(current_value.tags)
        if current_value.value and current_value.facts is not None:
            facts.update(current_value.facts)
                    
        any_value = any_value or current_value.value
        if any_value:
            break
    return_value = xv.XuleValue(xule_context, any_value, 'bool')
    if len(tags) > 0:
        return_value.tags = tags
    if len(facts) > 0 :
        return_value.facts = facts
    return return_value #xv.XuleValue(xule_context, any_value, 'bool')

def agg_first(xule_context, values):
    return values[0].clone()

def agg_max(xule_context, values):
    agg_value = values[0].clone()
    
    for current_value in values[1:]:
        if agg_value.value < current_value.value:
            agg_value = current_value.clone()
            
    return agg_value

def agg_min(xule_context, values):
    agg_value = values[0].clone()
    
    for current_value in values[1:]:
        if agg_value.value > current_value.value:
            agg_value = current_value.clone()
            
    return agg_value    
    
def agg_list(xule_context, values):
#Commented out for elimination of the shadow_collection
#     list_values = []
#     shadow = []
#     
#     for current_value in values:
#         list_values.append(current_value)
#         shadow.append(current_value.shadow_collection if current_value.type in ('list','set') else current_value.value)
#         
#     return xv.XuleValue(xule_context, tuple(list_values), 'list', shadow_collection=tuple(shadow))

    list_values = []
    shadow = []
    tags = {}
    facts = collections.OrderedDict()
    
    for current_value in values:
        list_values.append(current_value)
        shadow.append(current_value.shadow_collection if current_value.type in ('list','set', 'dictionary') else current_value.value)
        if current_value.tags is not None:
            tags.update(current_value.tags)
        if current_value.facts is not None:
            facts.update(current_value.facts)
    
    return_value = xv.XuleValue(xule_context, tuple(list_values), 'list', shadow_collection=tuple(shadow))
    if len(tags) > 0:
        return_value.tags = tags
    if len(facts) > 0:
        return_value.facts = facts
    return return_value #xv.XuleValue(xule_context, tuple(list_values), 'list')

def agg_set(xule_context, values):
    set_values = []
    shadow = []
    tags = {}
    facts = collections.OrderedDict()
    
    for current_value in values:
        if current_value.type in ('set', 'list', 'dictionary'):
            if current_value.shadow_collection not in shadow:
                set_values.append(current_value)
                shadow.append(current_value.shadow_collection)
                if current_value.tags is not None:
                    tags.update(current_value.tags)
                if current_value.facts is not None:
                    facts.update(current_value.facts)
        else:
            if current_value.value not in shadow:
                set_values.append(current_value)
                shadow.append(current_value.value)
                if current_value.tags is not None:
                    tags.update(current_value.tags)
                if current_value.facts is not None:
                    facts.update(current_value.facts)
    
    return_value = xv.XuleValue(xule_context, frozenset(set_values), 'set', shadow_collection=frozenset(shadow))
    if len(tags) > 0:
        return_value.tags = tags
    if len(facts) > 0:
        return_value.facts = facts
    return return_value #xv.XuleValue(xule_context, frozenset(set_values), 'set')

def agg_dict(xule_context, values):
    shadow = []
    tags = {}
    facts = collections.OrderedDict()
    
    dict_values = dict()
    shadow = dict()
    
    for current_value in values:
        if current_value.type != 'list':
            raise XuleProcessingError(_("Arguments for the dict() function must be lists of key/value pairs, found %s" % current_value.type),
                                      xule_context)
        if len(current_value.value) != 2:
            raise XuleProcessingError(_("Arguments for the dict() function must be lists of length 2 (key/value pair). Found list of length %i" % len(current_value.value)))
    
        key = current_value.value[0]
        if key.type == 'dictionary':
            raise XuleProcessingError(_("Key to a dictionary cannot be a dictionary."), xule_context)
        
        value = current_value.value[1]
        
        # A dictionary can only have one value for a key. If the key is already in the dicitionary then the current set of values can be skipped.
        # This is determined by the shadow (the underlying value of the XuleValue) as it is done for agg_set.
        if key.type in ('set', 'list'):
            if key.shadow_collection in shadow:
                continue
        else:
            if key.value in shadow:
                continue
        
        if key.tags is not None:
            tags.update(key.tags)
        if value.tags is not None:
            tags.update(value.tags)
        if key.facts is not None:
            facts.update(key.facts)     
        if value.facts is not None:
            facts.update(value.facts)                       
        
        dict_values[key] = value
  
        shadow[key.shadow_collection if key.type in ('list', 'set') else key.value] = value.shadow_collection if value.type in ('list', 'set', 'dictionary') else value.value

    return_value = xv.XuleValue(xule_context, frozenset(dict_values.items()), 'dictionary', shadow_collection=frozenset(shadow.items()))
    if len(tags) > 0:
        return_value.tags = tags
    if len(facts) > 0:
        return_value.facts = facts
    return return_value  

def func_taxonomy(xule_context, *args):
    if len(args) == 0:
        if xule_context.model is None:
            # There is no instance doucment
            return xv.XuleValue(xule_context, None, 'none')
        else:
            setattr(xule_context.model, 'taxonomy_name', 'instance')
            return xv.XuleValue(xule_context, xule_context.model, 'taxonomy')
    elif len(args) == 1:
        taxonomy_url = args[0]
        if taxonomy_url.type not in ('string', 'uri'):
            raise XuleProcessingError(_("The taxonomy() function takes a string or uri, found {}.".format(taxonomy_url.type)), xule_context)
        
        other_taxonomy = xule_context.get_other_taxonomies(taxonomy_url.value)
        setattr(other_taxonomy, 'taxonomy_name', taxonomy_url.value)
        return xv.XuleValue(xule_context, other_taxonomy , 'taxonomy')
    else:
        raise XuleProcessingError(_("The taxonomy() function takes at most 1 argument, found {}".format(len(args))))

def func_excel_data(xule_context, *args):
    """Read an excel file/url
    
    Arguments:
        file_url (stirng or url)
        range (string) - A sheet name, named range or cell range
        has_header (boolean) - determines if the first line of the cs file has headers
        type_list (list) - list of xule types/transformatons in the order of the columns of the columns of the specified range in the excel sheet. If omitted
                           the columns will be strings
        as_dictionary (boolean) - return the row as a dictionary instead of a list. Optional
    """
    # Validate the arguments
    if len(args) < 1:
        raise XuleProcessingError(_("The excel-datat() function requires at least 1 argument (file url, found {} arguments".format(len(args))), xule_context)
    if len(args) > 5:
        raise XuleProcessingError(_("The excel-data() function takes no more than 5 arguments (file url, range, has_headers, columne types, as dictionary), found {}".format(len(args))), xule_context)

    file_url = args[0]
    if file_url.type not in ('string', 'uri'):
        raise XuleProcessingError(_("The file url argument (1st) of the excel-data() function must be a string or uri, found {}".format(file_url.type)), xule_context)

    if len(args) > 1:
        range_descriptor = args[1]
        if range_descriptor.type != 'string':
            raise XuleProcessingError(_("The cell range argument (2nd) of the excel-data() function must be a string, found {}".format(range_descriptor.type)), xule_context)
    else:
        range_descriptor = None

    if len(args) > 2:
        if args[2].type != 'bool':
            raise XuleProcessingError(_("The has headers argument (3rd) of the excel-data() function muset be a boolean, found '{}'.".format(args[2].type)), xule_context)
        has_headers = args[2].value
    else: # default is false
        has_headers = False

    if len(args) >= 4: 
        ordered_cols = validate_data_field_types(args[3], 'excel-data', xule_context)   
    else:
        ordered_cols = None
    
    if len(args) == 5:
        if args[4].type != 'bool':
            raise XuleProcessingError(_("The as dictionary argument (5th) of the excel-data() function must be a boolean, found '{}'.".format(args[4].type)), xule_context)
        if args[2].value:
            return_row_type = 'dictionary'
        else:
            return_row_type = 'list'
    else:
        return_row_type = 'list'
        
    if return_row_type == 'dictionary' and not has_headers:
        raise XuleProcessingError(_("When the excel-data() function is returning the rows as dictionaries (5th argument), the has headers argument (3rd argument) must be true."), xule_context)

    # Open the workbook      
    # Using the FileSource object in arelle. This will open the file and handle taxonomy package mappings.
    file_source = FileSource.openFileSource(file_url.value, xule_context.global_context.cntlr)
    file = file_source.file(file_url.value, binary=True)
    # file is  tuple of one item as a BytesIO stream. Since this is in bytes, it needs to be converted to text via a decoder.
    # Assuming the file is in utf-8. 
    with open_excel(file[0]) as wb:

        # get cell range - for named ranges there may be multiple, so the cell range will be a tuple of ranges
        # cell range may be a sheet name, named range or a cell reference (i.e. sheet1!a1:c3) or a single cell
        sheet_name = None
        cell_range_text = None
        named_range_text = None

        if range_descriptor is not None:
            if ':' in range_descriptor.value or '!' in range_descriptor.value:
                # this is a cell range (i.e sheet11:a1:c3)
                if '!' in range_descriptor.value:
                    sheet_name, cell_range_text = range_descriptor.value.split('!')
                else:
                    cell_range_text = range_descriptor.value
            else:
                # This is either a sheet name, named range or a single cell
                if range_descriptor.value in wb.sheetnames:
                    sheet_name = range_descriptor.value
                elif range_descriptor.value in wb.defined_names:
                    named_range_text = range_descriptor.value
                else: # this may be a single cell
                    cell_range_text = range_descriptor.value
            
            if named_range_text is not None:
                ranges = wb.defined_names[named_range_text].destinations # This returns a list of tuples. The tuple contest 0 - sheet anem 1 - cell range
            else:
                if sheet_name is None:
                    sheet_name = wb.active.title
                ranges = [(sheet_name, cell_range_text)] # This mimics what wb.defined_names returns
        
        result = []
        result_shadow = []
        first_line = True
        row_num = 0
        for range_sheet, range_cells_text in ranges:
            try:
                ws = wb[range_sheet]
            except KeyError:
                raise XuleProcessingError(_("Sheet '{}' does not exist in workbook '{}'".format(range_sheet, file_url.value)), xule_context)
            if range_cells_text is None:
                # it will be the whole sheet
                rows = ws.rows
            else:
                try:
                    rows = ws[range_cells_text]
                except ValueError:
                    raise XuleProcessingError(_("Cell range '{}' is not a valid cell range for worksheet '{}' in workbook '{}'".format(range_cells_text, range_sheet, file_url.value)), xule_context)

            if not isinstance(rows, Iterable):
                # This is a single cell refernce. Row will be the cell
                rows = ((rows,),)
            for row in rows:
                row_num += 1
                if first_line and has_headers:
                    first_line = False
                    #skip the headers line
                    if return_row_type == 'dictionary':
                        # Need to get the names from the first row
                        column_names = [x.value for x in row]
                        if len(column_names) != len(set(column_names)):
                            raise XuleProcessingError(_("There are duplicate column names in the excel file. This is not allowed when return rows as dictionaries. File: {}".format(file_url.value)), xule_context)
                    continue # now skip to the next line
                if return_row_type == 'list':
                    result_line = list()
                    result_line_shadow = list()
                else: #dictionary
                    result_line = dict()
                    result_line_shadow = dict()
                
                for col_num, item in enumerate(row):
                    if ordered_cols is not None and col_num >= len(ordered_cols):
                        raise XuleProcessingError(_("The nubmer of columns on row {} is greater than the number of column types provided in the 4th argument of the excel-data() function. File: {}".format(row_num, file_url.value)), xule_context)
                    
                    if type(item.value) == datetime.datetime:
                        v = item.value.isoformat()
                    else:
                        v = str(item.value)

                    item_value = convert_file_data_item(v, ordered_cols[col_num] if ordered_cols is not None else None, xule_context)

                    if return_row_type == 'list':
                        result_line.append(item_value)
                        result_line_shadow.append(item_value.value)
                    else: #dictonary
                        if col_num >= len(column_names):
                            raise xule_context(_("The number of columns on row {} is greater than the number of headers in the csv file. File: {}".format(row_num, 
                                                                                                                                                        mappedUrl if mapped_file_url == file_url.value else file_url.value + ' --> ' + mapped_file_url)), xule_context)

                        result_line[xv.XuleValue(xule_context, column_names[col_num], 'string')] = item_value
                        result_line_shadow[column_names[col_num]] = item_value.value
                        
                if return_row_type == 'list':
                    result.append(xv.XuleValue(xule_context, tuple(result_line), 'list', shadow_collection=tuple(result_line_shadow)))
                    result_shadow.append(tuple(result_line_shadow))
                else: #dictionary
                    result.append(xv.XuleValue(xule_context, frozenset(result_line.items()), 'dictionary', shadow_collection=frozenset(result_line_shadow.items())))
                    result_shadow.append(frozenset(result_line_shadow.items()))
          
    return xv.XuleValue(xule_context, tuple(result), 'list', shadow_collection=tuple(result_shadow))

@contextmanager
def open_excel(file_name):
    workbook = openpyxl.load_workbook(file_name)
    yield workbook
    workbook.close()

def func_csv_data(xule_context, *args):
    """Read a csv file/url.
    
    Arguments:
        file_url (string or url)
        has_header (boolean) - determines if the first line of the csv file has headers
        type list (list) - list of xule types in the order of the columns of the csv file. This is optional. If not provided, then all the data will be
                           treated as stirngs.
        as_dictionary (boolean) - return the row as a dictionary instead of a list. This is optional.
    """
    if len(args) == 0:
        raise XuleProcessingError(_("The csv-data() function requires at least 1 argument (file url), found no arguments."), xule_context)
    if len(args) > 4:
        raise XuleProcessingError(_("The csv-data() function takes no more than 4 arguments (file url, has headers, column types, as dictionary), found {} arguments.".format(len(args))), xule_context)

    file_url = args[0]
    

    if file_url.type not in ('string', 'uri'):
        raise XuleProcessingError(_("The file url argument (1st argument) of the csv-dta() function must be a string or uri, found '{}'.".format(file_url.value)), xule_context)
    
    if len(args) > 1:
        if args[1].type != 'bool':
            raise XuleProcessingError(_("The has headers argument (2nd argument) of the csv-data() function muset be a boolean, found '{}'.".format(args[1].type)), xule_context)
        has_headers = args[1].value
    else: # default is false
        has_headers = False

    if len(args) >= 3: 
        ordered_cols = validate_data_field_types(args[2], 'csv-data', xule_context)   
    else:
        ordered_cols = None
    
    if len(args) == 4:
        if args[3].type != 'bool':
            raise XuleProcessingError(_("The as dictionary argument (4th argument) of the csv-data() function must be a boolean, found '{}'.".format(args[3].type)), xule_context)
        if args[3].value:
            return_row_type = 'dictionary'
        else:
            return_row_type = 'list'
    else:
        return_row_type = 'list'
        
    if return_row_type == 'dictionary' and not has_headers:
        raise XuleProcessingError(_("When the csv-data() function is returning the rows as dictionaries (4th argument), the has headers argument (2nd argument) must be true."), xule_context)
        
    result = list()
    result_shadow = list()
    
    mapped_file_url = PackageManager.mappedUrl(file_url.value)

    # Using the FileSource object in arelle. This will open the file and handle taxonomy package mappings.
    file_source = FileSource.openFileSource(file_url.value, xule_context.global_context.cntlr)
    file = file_source.file(file_url.value, binary=True)
    # file is  tuple of one item as a BytesIO stream. Since this is in bytes, it needs to be converted to text via a decoder.
    # Assuming the file is in utf-8. 
    data_source = [x.decode('utf-8') for x in file[0].readlines()]
 
    import csv
    reader = csv.reader(data_source)
    first_line = True
    row_num = 0
    for line in reader:
        row_num += 1
        if first_line and has_headers:
            first_line = False
            #skip the headers line
            if return_row_type == 'dictionary':
                # Need to get the names from the first row
                column_names = [x for x in line]
                if len(column_names) != len(set(column_names)):
                    raise XuleProcessingError(_("There are duplicate column names in the csv file. This is not allowed when return rows as dictionaries. File: {}".format(file_url.value)), xule_context)
                
            continue
        
        if return_row_type == 'list':
            result_line = list()
            result_line_shadow = list()
        else: #dictionary
            result_line = dict()
            result_line_shadow = dict()
            
        for col_num, item in enumerate(line):
            if ordered_cols is not None and col_num >= len(ordered_cols):
                raise XuleProcessingError(_("The nubmer of columns on row {} is greater than the number of column types provided in the third argument of the csv-data() function. File: {}".format(row_num, file_url.value)), xule_context)
            
            item_value = convert_file_data_item(item, ordered_cols[col_num] if ordered_cols is not None else None, xule_context)

            if return_row_type == 'list':
                result_line.append(item_value)
                result_line_shadow.append(item_value.value)
            else: #dictonary
                if col_num >= len(column_names):
                    raise xule_context(_("The number of columns on row {} is greater than the number of headers in the csv file. File: {}".format(row_num, 
                                                                                                                                                  mappedUrl if mapped_file_url == file_url.value else file_url.value + ' --> ' + mapped_file_url)), xule_context)

                result_line[xv.XuleValue(xule_context, column_names[col_num], 'string')] = item_value
                result_line_shadow[column_names[col_num]] = item_value.value
                
        if return_row_type == 'list':
            result.append(xv.XuleValue(xule_context, tuple(result_line), 'list', shadow_collection=tuple(result_line_shadow)))
            result_shadow.append(tuple(result_line_shadow))
        else: #dictionary
            result.append(xv.XuleValue(xule_context, frozenset(result_line.items()), 'dictionary', shadow_collection=frozenset(result_line_shadow.items())))
            result_shadow.append(frozenset(result_line_shadow.items()))
          
    return xv.XuleValue(xule_context, tuple(result), 'list', shadow_collection=tuple(result_shadow))

def validate_data_field_types(column_types, func_name, xule_context):

    if column_types.type == 'none':
        ordered_cols = None
    elif column_types.type == 'list':
        ordered_cols = list()
        for col in column_types.value:
            if col.type not in  ('string', 'qname', 'list'): # qnames are used for transforms
                raise XuleProcessingError(_("The type list argument (3rd argument) of the {}() function must be a list of strings or qnames (for transforms) or a 2 item list of the transform and output type, found '{}'.".format(func_name, col.type)), xule_context)
            ordered_cols.append(col)
    else:
        raise XuleProcessingError(_("The type list argument (3rd argument) of the {}() fucntion must be list, found '{}'.".format(func_name, column_types.type)), xule_context)

    return ordered_cols

def convert_file_data_item(val, type, xule_context):
    
    if type is None:
        return xv.XuleValue(xule_context, val, 'string')

    if type.type in ('qname', 'list'):
        # this is a transform
        # Need to convert the stering version of the value to the canonical string version using the transform format.
        if type.type == 'qname': # This is just a transform, the type will be a string
            f = type.value
            output_type = 'string'
        else: # this is a list of 2 items. the first is the transform and the second is output type
            if len(type.value) != 2:
                raise XuleProcessingError(_("When the type value is a list, it must have 2 items, the first is the transform name and the second is output type"), xule_context)
            if type.value[0].type != 'qname':
                raise XuleProcessingError(_("The fist item in a type list must be a qname for a transform. Found '{}'".format(type.value[0].type)), xule_context)
            f = type.value[0].value
            if type.value[1].type != 'string':
                raise XuleProcessingError(_("The second item in a type list must be sting indicating the output type. Found '{}'".format(type.value[1].type)), xule_context)
            output_type = type.value[1].value

        if f.namespaceURI in FunctionIxt.ixtNamespaceFunctions:
            try:
                v = FunctionIxt.ixtNamespaceFunctions[f.namespaceURI][f.localName](val)
            except Exception as err:
                raise XuleProcessingError(_("Unable to convert '{}' using transform '{}'.".format(val, f.clarkNotation)))
        else:
            try:
                v = xule_context.model.modelManager.customTransforms[f](val)
            except KeyError as err:
                raise XuleProcessingError(_("Transform '{}' is unknown".format(f.clarkNotation)))
            except Exception as err:
                raise XuleProcessingError(_("Unable to convert '{}' using transform '{}'.".format(val, f.clarkNotation)))
    else: #This is a string indicating the output type
        output_type = type.value
        v = val

    if output_type == 'qname':
        if v.count(':') == 0:
            prefix = '*' # This indicates the default namespace
            local_name = v
        elif v.count(':') == 1:
            prefix, local_name = v.split(':')
        else:
            raise XuleProcessingError(_("While processing a data file, QName in a file can only have one ':', found {} ':'s".format(val.count(':'))), xule_context)
        
        namespace = xule_context.rule_set.getNamespaceUri(prefix)
        
        return xv.XuleValue(xule_context, QName(prefix if prefix != '*' else None, namespace, local_name), 'qname')
    elif output_type == 'int':
        try:
            return xv.XuleValue(xule_context, int(v), 'int')
        except ValueError:
            raise XuleProcessingError(_("While processing a data file, cannot convert '{}' to an {}.".format(val, type)), xule_context)
    elif output_type == 'float':
        try:
            return xv.XuleValue(xule_context, float(v), 'float')
        except ValueError:
            raise XuleProcessingError(_("While processing a data file, cannot convert '{}' to a {}.".format(val, type)), xule_context)
    elif output_type == 'decimal':
        try:
            return xv.XuleValue(xule_context, decimal.Decimal(v), 'decimal')
        except decimal.InvalidOperation:
            raise XuleProcessingError(_("While processing a data file, cannot convert '{}' to a {}.".format(val, type)), xule_context)
    elif output_type == 'string':
        return xv.XuleValue(xule_context, v, 'string')  
    elif output_type == 'date':
        return xv.XuleValue(xule_context, datetime.datetime.fromisoformat(v).date(), 'date')
    elif output_type == 'boolean':
        return xv.XuleValue(xule_context, bool(v), 'boolean')
    else:
        raise XuleProcessingError(_("While processing a data file, {} is not implemented.".format(output_type)), xule_context)


def func_json_data(xule_context, *args):
    """Read a json file/url.
    
    Arguments:
        file_url (string or url)

    Returns a dictionary/list of the json data.
    """
    
    file_url = args[0]

    if file_url.type not in ('string', 'uri'):
        raise XuleProcessingError(_("The file url argument of the json-dta() function must be a string or uri, found '{}'.".format(file_url.value)), xule_context)

    mapped_file_url = PackageManager.mappedUrl(file_url.value)

    # Using the FileSource object in arelle. This will open the file and handle taxonomy package mappings.
    file_source = FileSource.openFileSource(file_url.value, xule_context.global_context.cntlr)
    file = file_source.file(file_url.value, binary=True)
    # file is  tuple of one item as a BytesIO stream. Since this is in bytes, it needs to be converted to text via a decoder.
    # Assuming the file is in utf-8. 
    data_source = [x.decode('utf-8') for x in file[0].readlines()]
    try:
        json_source = json.loads(''.join(data_source))
    #except JSONDecodeError:
    except ValueError:
        raise XuleProcessingError(_("The file '{}' is not a valid JSON file.".format(file_url.value)), xule_context)
    
    x = xv.system_collection_to_xule(json_source, xule_context)
    return xv.system_collection_to_xule(json_source, xule_context)

def func_xml_data_flat(xule_context, *args):
    """Reads an XML file and returns a list of records
     
       This function take 5 arguments:
          1 - file name
          2 - xpath expression of elements to get
          3 - list of fields to return. Each field is an xpath expression based on the element retrieved.
          4 - column types
          5 - namespace map
    """
    result = []
    shadow_result = []

    if len(args) < 3:
        raise XuleProcessingError(_(f"The xml-data-flat() funciton requires at least 3 arguments (file, locator xpath, list of fields), found {len(args)}"), xule_context)

    file_url = args[0]
    locator_xpath = args[1]
    fields = args[2]

    if file_url.type not in ('string', 'uri'):
        raise XuleProcessingError(_("The file url argument of the xml-data-flat() function must be a string or uri, found '{}'.".format(file_url.value)), xule_context)

    if locator_xpath.type != 'string':
        raise XuleProcessingError(_("The xpath locator for function xml-data-flat() is not a string"), xule_context)
    
    if fields.type != 'list':
        raise XuleProcessingError(_("The fields list for function xml-data-flat() is not a string"), xule_context)

    # Column types
    if len(args) >= 4:  
        ordered_cols = validate_data_field_types(args[3], 'xml-data-flat', xule_context)
    else:
        ordered_cols = None

    # namespace dictionary
    nsmap = {k: v['uri'] for k, v in xule_context.global_context.catalog['namespaces'].items() if k != '*'} 
    if len(args) == 5:
        namespaces = args[4]
        if namespaces.type != 'dictionary':
            raise XuleProcessingError(_("The namespace map argument of the xml-data-flat() function (4th argument) must be a dictionary"), xule_context)
        nsmap.update(namespaces.shadow_dictionary)
    if None in nsmap.keys():
        del nsmap[None] # this means there is no namespace for the default. A default is not allowed for etree element.xpath()

    field_count = 0
    for field in fields.value:
        field_count += 1
        if field.type != 'string':
            raise XuleProcessingError(_(f"Field in the {field_count} position is not a string"), xule_context)
    
    mapped_file_url = PackageManager.mappedUrl(file_url.value)
    try:
        xml = et.parse(FileSource.openFileStream(xule_context.global_context.cntlr, file_url.value))
    except IOError:
        raise XuleProcessingError(_(f"Cannot find file {file_url.value} for function xml-data-flat()"), xule_context)
    except:
        raise XuleProcessingError(_(f"Cannot open file {file_url.value} as an xml file"), xule_context)
    
    try:
        nodes = xml.xpath(locator_xpath.value, namespaces=nsmap)
    except:
        raise XuleProcessingError(_(f"In function xml-data-flat(), xpath failed. File: {file_url.value}, XPath: {locator_xpath.value}"), xule_context)
    
    for node in nodes:
        record = []
        shadow_record = []
        field_count = 0
        for field in fields.value:
            try:
                field_result = node.xpath(field.value, namespaces=nsmap)
                if len(field_result) == 0:
                    field_val = xv.XuleValue(xule_context, None, 'none')
                else:
                    if hasattr(field_result[0], 'text'):
                        field_val = field_result[0].text
                    else:
                        field_val = field_result[0]
                    if ordered_cols is not None and field_count >= len(ordered_cols):
                        raise XuleProcessingError(_("The nubmer of columns on row {} is greater than the number of column types provided in the third argument of the csv-data() function. File: {}".format(row_num, file_url.value)), xule_context)
            
                    field_val = convert_file_data_item(field_val, ordered_cols[field_count] if ordered_cols is not None else None, xule_context)
            except:
                raise XuleProcessingError(_(f"In function xml-data-flat, field xpath of '{field.value + 1}' is not valid."), xule_context)

            record.append(field_val)
            shadow_record.append(field_val.value)

            field_count += 1

        record_xule_val = xv.XuleValue(xule_context, tuple(record), 'list', shadow_collection=shadow_record)
        result.append(record_xule_val)
        shadow_result.append(shadow_record)
        
    return xv.XuleValue(xule_context, tuple(result), 'list', shadow_collection=shadow_result)

def func_first_value(xule_context, *args):
    """Return the first non None value.

    This function can take any number of arguments. It will return the first argument that is not None
    """
    for arg in args:
        if arg.value is not None:
            return arg.clone()
    # If here, either there were no arguments, or they were all none
    return xv.XuleValue(xule_context, None, 'unbound')

def func_range(xule_context, *args):
    """Return a list of numbers.

    If there is one argument it is the stop value of the range with the start value of 1.
    If there are 2 arguments, the first is the start and the second is the stop.
    If there are 3 argument, the first is the start, the second the stop and the third is the step.
    All the arguments must be convertible to an integer

    """
    # Check all arguments are numbers.
    for position, arg in enumerate(args, 1):
        if arg.type not in ('int', 'float', 'decimal'):
            ordinal = "%d%s" % (position,"tsnrhtdd"[(position/10%10!=1)*(position%10<4)*position%10::4])
            raise XuleProcessingError(
                _("The {} argument of the 'range' function must be a number, found '{}'".format(ordinal, arg.type)), xule_context)
        if not xv.xule_castable(arg, 'int', xule_context):
            ordinal = "%d%s" % (position, "tsnrhtdd"[(position / 10 % 10 != 1) * (position % 10 < 4) * position % 10::4])
            raise XuleProcessingError(
                _("The {} argument of the 'range' function must be an integer, found '{}'".format(ordinal, arg.value)),
                xule_context)


    if len(args) == 0:
        raise XuleProcessingError(_("The 'range' function requires at least one argument."), xule_context)
    elif len(args) == 1:
        start_num = 1
        stop_num = int(args[0].value) + 1
        step = 1
    elif len(args) == 2:
        start_num = int(args[0].value)
        stop_num = int(args[1].value) + 1
        step = 1
    else:
        start_num = int(args[0].value)
        stop_num = int(args[1].value) + 1
        step = int(args[2].value)

    # Check that the
    number_list = list(range(start_num, stop_num, step))
    number_list_values = tuple(xv.XuleValue(xule_context, x, 'int') for x in number_list)
    return xv.XuleValue(xule_context, number_list_values, 'list', shadow_collection=tuple(number_list))

def func_difference(xule_context, *args):
    '''Difference between 2 sets'''

    if args[0].type != 'set':
        raise XuleProcessingError(
            _("The first argument to the difference() fucntion must be a set, found '{}'".format(args[0].type)),
            xule_context)
    if args[1].type != 'set':
        raise XuleProcessingError(
            _("The second argument to the difference() fucntion must be a set, found '{}'".format(args[1].type)),
            xule_context)

    return xu.subtract_sets(xule_context, args[0], args[1])

def func_symmetric_difference(xule_context, *args):
    '''Symmetric difference between 2 sets'''

    if args[0].type != 'set':
        raise XuleProcessingError(
            _("The first argument to the symmetric_difference() fucntion must be a set, found '{}'".format(args[0].type)),
            xule_context)
    if args[1].type != 'set':
        raise XuleProcessingError(
            _("The second argument to the symmetric_difference() fucntion must be a set, found '{}'".format(args[1].type)),
            xule_context)

    return xu.symetric_difference(xule_context, args[0], args[1])

def func_version(xule_context, *args):
    '''Get the version number of the rule set'''
    version = xule_context.global_context.catalog.get('version', None)
    if version is None:
        return xv.XuleValue(xule_context, None, 'none')
    else:
        return xv.XuleValue(xule_context, version, 'string')

def func_rule_name(xule_context, *args):
    '''Get the name of the current executing rule'''
    if xule_context.rule_name is None:
        return xv.XuleValue(xule_context, None, 'none')
    else:
        return xv.XuleValue(xule_context, xule_context.rule_name, 'string')

# def func_alignment(xule_context, *args):
    
#     result = dict()
#     if xule_context.iteration_table.current_alignment is None:
#         return xv.XuleValue(xule_context, 'No Alignment', 'string')
#     for name_info, value in xule_context.iteration_table.current_alignment or tuple():
#         # name_info is a tuple 0 =  'builtin' or 'explicit_dimension' (even for typed dimensions), 1 = the value of the aspect
#         if name_info[0] == 'builtin':
#             if name_info[1] == 'entity':
#                 result['entity'] = f'[\"{value[0]}\", \"{value[1]}\"]'
#                 #xule_value = xv.XuleValue(xule_context, (xv.XuleValue(xule_context, value[0], 'string'), xv.XuleValue(xule_context, value[1], 'string')), 'entity')
#             elif name_info[1] == 'unit':
#                 result['unit'] = repr(value)
#                 #xule_value = xv.XuleValue(xule_context, value, 'unit')
#             elif name_info[1] == 'concept':
#                 result['concept'] = value.clarkNotation
#             elif name_info[1] == 'period': 
#                 if isinstance(value, tuple):
#                     # This is a duration
#                     if value[0] == datetime.datetime.min and value[1] == datetime.datetime.max:
#                         result['period'] = 'forever'
#                     else:
#                         result['period'] = f'{value[0].isoformat()}/{value[1].isoformat()}'
#                 else:
#                     result['period'] = value.isoformat()
#                 #xule_type, xule_value = xv.model_to_xule_type(xule_context, value)
#                 #xule_value = xv.XuleValue(xule_context, xule_value, xule_type)
#                 #result[xv.XuleValue(xule_context, name_info[1], 'string')] = xule_value
#         else:
#             if isinstance(value, QName):
#                 result[name_info[1].clarkNotation] = value.clarkNotation
#             else:
#                 result[name_info[1].clarkNotation] = str(value)

#             #xule_type, xule_value = xv.model_to_xule_type(xule_context, value)
#             #xule_value = xv.XuleValue(xule_context, xule_value, xule_type)
#             #result[xv.XuleValue(xule_context, name_info[1].clarkNotation, 'string')] = xule_value
    
#     return xv.XuleValue(xule_context, json.dumps(result), 'string')
#     #return xv.XuleValue(xule_context, frozenset(result.items()), 'dictionary')


def func_alignment(xule_context, *args):

    # if the context as 'result_alignment', then it is already calculated and is being used in an output message.
    return getattr(xule_context, 'result_alignment', _calc_alignment(xule_context))

def _calc_alignment(xule_context):

    result = dict()
    if xule_context.iteration_table.current_alignment is None:
        return xv.XuleValue(xule_context, None, 'none')
    for name_info, value in xule_context.iteration_table.current_alignment or tuple():
        # name_info is a tuple 0 =  'builtin' or 'explicit_dimension' (even for typed dimensions), 1 = the value of the aspect
        if name_info[0] == 'builtin':
            if name_info[1] == 'entity':
                #result['entity'] = f'[\"{value[0]}\", \"{value[1]}\"]'
                #xule_value = xv.XuleValue(xule_context, (xv.XuleValue(xule_context, value[0], 'string'), xv.XuleValue(xule_context, value[1], 'string')), 'entity')
                result[xv.XuleValue(xule_context, 'entity', 'string')] = xv.XuleValue(xule_context, (value[0], value[1]), 'entity') # a entity value is a tuple of the scheme and identifier
            elif name_info[1] == 'unit':
                #result['unit'] = repr(value)
                #xule_value = xv.XuleValue(xule_context, value, 'unit')
                result[xv.XuleValue(xule_context, 'unit', 'string')] = xv.XuleValue(xule_context, value, 'unit')
            elif name_info[1] == 'concept':
                #result['concept'] = value.clarkNotation
                result[xv.XuleValue(xule_context, 'concept', 'string')] = xv.XuleValue(xule_context, value, 'qname')
            elif name_info[1] == 'period': 
                if isinstance(value, tuple):
                    # This is a duration
                    result[xv.XuleValue(xule_context, 'period', 'string')] = xv.XuleValue(xule_context, value, 'duration')

                    # if value[0] == datetime.datetime.min and value[1] == datetime.datetime.max:
                    #     result['period'] = 'forever'
                    # else:
                    #     result['period'] = f'{value[0].isoformat()}/{value[1].isoformat()}'
                else:
                    result[xv.XuleValue(xule_context, 'period', 'string')] = xv.XuleValue(xule_context, value, 'instant')
                    #result['period'] = value.isoformat()
                #xule_type, xule_value = xv.model_to_xule_type(xule_context, value)
                #xule_value = xv.XuleValue(xule_context, xule_value, xule_type)
                #result[xv.XuleValue(xule_context, name_info[1], 'string')] = xule_value
        else:
            # if isinstance(value, QName):
            #     result[name_info[1].clarkNotation] = value.clarkNotation
            # else:
            #     result[name_info[1].clarkNotation] = str(value)

            xule_type, xule_value = xv.model_to_xule_type(xule_context, value)
            xule_value = xv.XuleValue(xule_context, xule_value, xule_type)
            result[xv.XuleValue(xule_context, name_info[1], 'qname')] = xule_value
    
    #return xv.XuleValue(xule_context, json.dumps(result), 'string')
    return xv.XuleValue(xule_context, frozenset(result.items()), 'dictionary')


#the position of the function information
FUNCTION_TYPE = 0
FUNCTION_EVALUATOR = 1
FUNCTION_ARG_NUM = 2
#aggregate only 
FUNCTION_DEFAULT_VALUE = 3
FUNCTION_DEFAULT_TYPE = 4
#non aggregate only
FUNCTION_ALLOW_UNBOUND_ARGS = 3
FUNCTION_RESULT_NUMBER = 4

def built_in_functions():
    funcs = {
#              'all': ('aggregate', agg_all, 1, True, 'bool'),
#              'any': ('aggregate', agg_any, 1, False, 'bool'),
#              'first': ('aggregate', agg_first, 1, None, None),
#              'count': ('aggregate', agg_count, 1, 0, 'int'),
#              'sum': ('aggregate', agg_sum, 1, None, None),
#              'max': ('aggregate', agg_max, 1, None, None), 
#              'min': ('aggregate', agg_min, 1, None, None),
             'list': ('aggregate', agg_list, 1, tuple(), 'list'),
             #'list': ('aggregate', agg_list, 1, None, None),
             'set': ('aggregate', agg_set, 1, frozenset(), 'set'),
             #'set': ('aggregate', agg_set, 1, None, None),
             'dict': ('aggregate', agg_dict, 1, frozenset(), 'dictionary'),
             
             'exists': ('regular', func_exists, 1, True, 'single'),
             'missing': ('regular', func_missing, 1, True, 'single'),
             #'instant': ('regular', func_instant, 1, False, 'single'),
             'date': ('regular', func_date, 1, False, 'single'),
             'duration': ('regular', func_duration, 2, False, 'single'),
             'forever': ('regular', func_forever, 0, False, 'single'),
             'unit': ('regular', func_unit, -2, False, 'single'),
             'entity': ('regular', func_entity, 2, False, 'single'),
             'qname': ('regular', func_qname, 2, True, 'single'),
             'uri': ('regular', func_uri, 1, False, 'single'),
             'time-span': ('regular', func_time_span, 1, False, 'single'),
             'schema-type': ('regular', func_schema_type, 1, False, 'single'),
             'num-to-string': ('regular', func_num_to_string, 1, False, 'single'),
             'mod': ('regular', func_mod, 2, False, 'single'),
             'random': ('regular', func_random, -3, False, 'single'),
             'extension-concepts': ('regular', func_extension_concept, 0, False, 'single'),
             'taxonomy': ('regular', func_taxonomy, -1, False, 'single'),
             'csv-data': ('regular', func_csv_data, -4, False, 'single'),
             'json-data': ('regular', func_json_data, 1, False, 'single'),
             'xml-data-flat': ('regular', func_xml_data_flat, -5, False, 'single'),
             'excel-data': ('regular', func_excel_data, -5, False, 'single'),
             'first-value': ('regular', func_first_value, None, True, 'single'),
             'range': ('regular', func_range, -3, False, 'single'),
             'difference': ('regular', func_difference, 2, False, 'single'),
             'symmetric_difference': ('regular', func_symmetric_difference, 2, False, 'single'),
             'version': ('regular', func_version, 0, False, 'single'),
             'rule-name': ('regular', func_rule_name, 0, False, 'single'),
             'alignment': ('regular', func_alignment, 0, False, 'single')
             }    

    try:
        funcs.update(xrf.BUILTIN_FUNCTIONS)
        funcs.update(xif.BUILTIN_FUNCTIONS)
    except NameError:
        pass
    
    return funcs

BUILTIN_FUNCTIONS = built_in_functions()



#BUILTIN_FUNCTIONS = {}
