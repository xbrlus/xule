"""XuleProcessor

Xule is a rule processor for XBRL (X)brl r(ULE). 

The XuleProcessor module is the main module for processing a rule set against an instance.

DOCSKIP
See https://xbrl.us/dqc-license for license information.  
See https://xbrl.us/dqc-patent for patent infringement notice.
Copyright (c) 2017 - 2021 XBRL US, Inc.

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

    http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.

$Change$
DOCSKIP
"""
from .XuleContext import XuleGlobalContext, XuleRuleContext  # XuleContext
from .XuleFunctions import func_alignment
from .XuleRelattionshipSet import XuleRelationshipSet
from .XuleRunTime import XuleProcessingError, XuleIterationStop, XuleException, XuleBuildTableError, XuleReEvaluate
from .XuleValue import *
from . import XuleConstants as xc
from . import XuleModelIndexer as xmi
from . import XuleUtility
import itertools as it
from arelle.ModelValue import QName, dayTimeDuration, DateTime, gYear, gMonthDay, gYearMonth, qname
from arelle.ModelInstanceObject import ModelFact
from arelle.ModelRelationshipSet import ModelRelationshipSet
from arelle.ModelDtsObject import ModelConcept
from arelle.ModelObject import ModelObject
from arelle.XmlValidate import VALID
import decimal
import datetime
from aniso8601 import parse_duration, parse_datetime, parse_date
import collections
import copy
from threading import Thread
from . import XuleFunctions
from . import XuleProperties
import os
from openpyxl import load_workbook, Workbook
import json
import regex as re

def process_xule(rule_set, model_xbrl, cntlr, options, saved_taxonomies=None, is_validator=False):
    """Run xule rules against a filing.
    
    :param rule_set: An opened rule set
    :type rule_set: XuleRuleSet
    :param model_xbrl: An Arelle model of the instance document
    :type model_xbrl: ModelXbrl
    :param cntlr: An Arelle controller
    :type cntlr: Cntlr
    :param options: The command line options
    :type options: optparse
    
    This is the main function to process a Xule ruleset against a filing. This function just sets a few things up. 
    The most import item is the processor context. The context saves the state of the processor throughout the 
    processing of the rules.
    """

    # Save the controller in XuleValue
    init_cntlr(cntlr)

    global_context = XuleGlobalContext(rule_set, model_xbrl, cntlr, options)
    if saved_taxonomies is not None and len(saved_taxonomies) > 0:
        global_context.other_taxonomies = saved_taxonomies

        # Set up trace files
    if getattr(global_context.options, "xule_trace_count", False):
        try:
            os.remove(global_context.options.xule_trace_count + ".csv")
        except FileNotFoundError:
            pass
        try:
            os.remove(global_context.options.xule_trace_count + ".txt")
        except FileNotFoundError:
            pass

    if getattr(global_context.options, "xule_time", None) is not None:
        total_start = datetime.datetime.today()

    if getattr(global_context.options, "xule_multi", False):
        from .XuleMultiProcessing import output_message_queue
        t = Thread(target=output_message_queue, args=(global_context,))
        t.name = "Message Queue"
        t.start()

    if getattr(global_context.options, "xule_time", None) is not None:
        fact_index_start = datetime.datetime.today()

    # Create the processing context to build the index
    xule_context = XuleRuleContext(global_context, cache_size_bytes=getattr(global_context.options, 'xule_cache_size_bytes', getattr(global_context.options, 'xule_cache_size_bytes', 1_000_000)))
    # Build an index on the facts in the model.
    xmi.index_model(xule_context)
    # Clean up
    del xule_context

    if getattr(global_context.options, "xule_time", None) is not None:
        fact_index_end = datetime.datetime.today()
        global_context.message_queue.print("Index build time %s." % (fact_index_end - fact_index_start), "xule.buildIndexTime")

    # Determine if constants should be precalced. This is determined by the --xule-precalc-constants optoin on the command line. This is useful to simulate how the processor works in the server
    # environment.
    if getattr(global_context.options, "xule_precalc_constants", False):
        constant_start = datetime.datetime.today()
        process_precalc_constants(global_context)
        constant_end = datetime.datetime.today()
        constant_time = constant_end - constant_start
        global_context.message_queue.print("Time to calculated non instance constants: %s" % (constant_time), "xule.precalcConsts")

    # Load any reloadable saved constnats
    if getattr(global_context.options, "xule_args_file", False):
        constant_start = datetime.datetime.today()
        process_reloadable_constants(global_context)
        global_context.message_queue.print("Time to calculated non instance constants: %s" % (datetime.datetime.today() - constant_start), "xule.loadSavedConstants")

    # Determine if constants should be outputed
    if getattr(global_context.options, "xule_output_constants", None) is not None:
        output_constant(global_context, cntlr)

    if getattr(global_context.options, "xule_run", False) or is_validator:
    # global_context.message_queue.logging("Processing Filing...")
        evaluate_rule_set(global_context)

    if getattr(global_context.options, "xule_time", None) is not None:
        total_end = datetime.datetime.today()
        if getattr(global_context.options, "xule_precalc_constants", False):
            global_context.message_queue.print(
                "Time to process excluding non instance constant: %s." % (total_end - total_start - constant_time), "xule.ruleEvalTime")
        global_context.message_queue.print("Total time to process: %s." % (total_end - total_start), "xule.ruleEvalTime")
    # Shutdown Message Queue
    if getattr(global_context.options, "xule_multi", False):
        global_context.message_queue.stop()
        global_context.message_queue.clear()
        t.join()
        # Save any taxonomies that were opened
    saved_taxonomies = global_context.other_taxonomies
    # clean up
    del global_context

    return saved_taxonomies

def output_constant(global_context, cntlr):
    # Change the message format for the log handler so it only outputs the message
    save_log_handler = cntlr.logHandler
    new_log_handler = copy.copy(save_log_handler)
    import logging
    # create formatter and add it to the handlers
    formatter = logging.Formatter('%(message)s')
    new_log_handler.setFormatter(formatter)
    cntlr.logger.removeHandler(cntlr.logHandler)
    cntlr.logger.addHandler(new_log_handler)

    save_to_file = getattr(global_context.options, "xule_output_constants_file", None)
    const_objs = {} # for saving to a file

    for constant_name_raw in getattr(global_context.options, "xule_output_constants").split(','):
        constant_name = constant_name_raw.strip()
        if constant_name != '':
            const_info = precalc_constant(global_context, constant_name)
            if const_info is None: # The constant was not found
                cntlr.addToLog(f"Constant {constant_name} does not exist", 'warning')
            else:
                for alignment, values in const_info['value'].values.items():
                    for val in values:
                        if save_to_file:
                            const_objs[constant_name] = val.reloadable_value(constant_name)
                        else:
                            cntlr.addToLog(val.format_value(), constant_name, messageArgs={'alignment': alignment, 'name': constant_name})

    if save_to_file:
        with open(save_to_file, "w") as fh:
            json.dump(const_objs, fh)

    cntlr.logger.addHandler(save_log_handler)

def evaluate_rule_set(global_context):
    """Process the rule set.
    
    :param global_context: The global processing context
    :type global_context: XuleGlobalContext
    :param skip_rules: A list of rule names to skip
    :type skip_rules: list
    
    This function loops through all the rules in the rule set and evaluates each rule.
    
    During evaluation of a rule, the evaluator can produce a XuleIterationStop exception. This exception indicates
    that processing of the current iteration of the rule can stop and go to the next iteration.
    """
    if getattr(global_context.options, "xule_time", None) is not None:
        times = []

    # Create a list of rules to skip. These are determined by the --xule-skip option on the command line.
    skip_rules = getattr(global_context.options, "xule_skip", None).split(",") if getattr(global_context.options,
                                                                                          "xule_skip",
                                                                                          None) is not None else None

    # Create a list of run only rules. This is the opposite of skip_rules. If run_only is not NOne than only those rules will be processed.
    run_only_rules = getattr(global_context.options, "xule_run_only", None).split(",") if getattr(
        global_context.options, "xule_run_only", None) is not None else None
    run_only_rule_pattern = re.compile(getattr(global_context.options, "xule_run_only_pattern")) if getattr(
        global_context.options, "xule_run_only_pattern", None) is not None else None

    # use the term "cat" for catalog information. Read through the list of rules in the catalog.
    for file_num, cat_rules in global_context.catalog['rules_by_file'].items():
        for rule_name in sorted(cat_rules.keys()):
            cat_rule = cat_rules[rule_name]

            if skip_rules is not None and rule_name in skip_rules:
                global_context.message_queue.print("Skipping rule: %s" % rule_name, "xule.ruleSkipped")
                continue

            if not (run_only_rules is None or rule_name in run_only_rules):
                continue
            if not (run_only_rule_pattern is None or run_only_rule_pattern.match(rule_name)):
                continue

            # get the AST for the rule from the ruleset
            rule = global_context.rule_set.getItem(cat_rule)

            if getattr(global_context.options, "xule_debug", False):
                global_context.message_queue.print(
                    "Processing: %s - %s" % (rule_name, datetime.datetime.today().strftime("%H:%M:%S.%f")),
                    "xule.ruleDebug")
                if global_context.model is not None:
                    global_context.message_queue.print(global_context.model.modelDocument.uri, "xule.ruleDebug")

            try:
                if getattr(global_context.options, "xule_time", None) is not None or getattr(global_context.options,
                                                                                             "xule_trace_count", False):
                    rule_start = datetime.datetime.today()

                # Establish the rule context. A new context is created for each rule.
                xule_context = XuleRuleContext(global_context,
                                               rule_name,
                                               file_num)
                # add the main table
                xule_context.iteration_table.add_table(rule['node_id'], xule_context.get_processing_id(rule['node_id']))

                # Evaluate the rule. 
                if global_context.model is not None:
                    global_context.model.modelManager.showStatus("Processing rule {}".format(rule_name))
                evaluate(rule, xule_context)

            except (XuleProcessingError, XuleBuildTableError) as e:
                if getattr(global_context.options, "xule_crash", False):
                    raise
                else:
                    xule_context.global_context.message_queue.error("xule:error", "rule %s: %s" % (rule_name, str(e)))

            except XuleIterationStop:
                pass

            except Exception as e:
                if getattr(global_context.options, "xule_crash", False):
                    # if global_context.crash_on_error:
                    raise
                else:
                    xule_context.global_context.message_queue.error("xule:error", "rule %s: %s" % (rule_name, str(e)))

            if getattr(global_context.options, "xule_time", None) is not None:
                rule_end = datetime.datetime.today()
                times.append((xule_context.rule_name, rule_end - rule_start))
                if getattr(global_context.options, "xule_debug", False):
                    global_context.message_queue.print("%s time took: %s - %s " % (
                    rule_name, (rule_end - rule_start).total_seconds(),
                    datetime.datetime.today().strftime("%H:%M:%S.%f")), "xule.ruleEvalTime")

            if getattr(global_context.options, "xule_trace_count", False):
                total_time = datetime.datetime.today() - rule_start
                global_context.message_queue.print(
                      "Total iterations:", xule_context.iter_count,
                      "Messages:", xule_context.iter_message_count,
                      "Pass:", xule_context.iter_pass_count,
                      "Non alignment:", xule_context.iter_misaligned_count,
                      "Exception:", xule_context.iter_except_count, "xule.traceCount")
                write_trace_count_csv(global_context.options.xule_trace_count, rule_name,
                                      global_context.expression_trace, rule, xule_context.iter_count, total_time)
                write_trace_count_string(global_context.options.xule_trace_count, rule_name,
                                         global_context.expression_trace, rule, xule_context.iter_count, total_time)

            # clean up
            del xule_context

    # output stats
    if getattr(global_context.options, "xule_rule_stats_file", None) is not None:
        stats_file_name = getattr(global_context.options, "xule_rule_stats_file")
        if not stats_file_name.endswith('.json'):
            stats_file_name = stats_file_name + '.json'
        with open(stats_file_name, 'w') as stats_file:
            json.dump(global_context.stats, stats_file)

    # Display timing information
    if getattr(global_context.options, "xule_time", None) is not None:
        global_context.message_queue.print("Total number of rules processed: %i" % len(times), "xule.ruleEvalTime")
        # slow_rules = [timing_info for timing_info in times if timing_info[1].total_seconds() > 0.001]
        slow_rules = sorted([timing_info for timing_info in times if
                             timing_info[1].total_seconds() > getattr(global_context.options, "xule_time", None)],
                            key=lambda tup: tup[1], reverse=True)
        # slow_rules = [timing_info for timing_info in times if timing_info[1].total_seconds() > global_context.show_timing]
        global_context.message_queue.print(
            "Number of rules over %ss: %i" % (getattr(global_context.options, "xule_time", None), len(slow_rules)), "xule.ruleEvalTime")
        for slow_rule in slow_rules:
            global_context.message_queue.print("Rule %s end. Took %s" % (slow_rule[0], slow_rule[1]), "xule.ruleEvalTime")
            # global_context.message_queue.print("Global expression cache size: %i" % len(global_context.expression_cache))

def index_property_start(model_fact):
    if model_fact.context.isStartEndPeriod:
        return model_fact.context.startDatetime
    elif model_fact.context.isInstantPeriod:
        return model_fact.context.endDatetime - datetime.timedelta(days=1)


def index_property_end(model_fact):
    if model_fact.context.isStartEndPeriod:
        return model_fact.context.endDatetime - datetime.timedelta(days=1)
    elif model_fact.context.isInstantPeriod:
        return model_fact.context.endDatetime - datetime.timedelta(days=1)


def index_property_days(model_fact):
    if model_fact.context.isStartEndPeriod:
        return (model_fact.context.endDatetime - model_fact.context.startDatetime).days
    elif model_fact.context.isInstantPeriod:
        return 0


def index_property_balance(model_fact):
    if model_fact.concept.balance is not None:
        return model_fact.concept.balance


# These are aspect properties that are used in a factset. I.e. @concept.is-monetary. The 'is-monetary' is an aspect property of the
# aspect 'concept'. This list identifies all the expected properties that will be in the fact index. This a dictionary with the key
# being the property identifier and the value being the function to calculate the value of the property with a fact. The property identifier
# is a 3 part tuple:
#    1 - 'property' - This part is always 'property'.
#    2 - the aspect name
#    3 - the property name
# _FACT_INDEX_PROPERTIES = {
#     ('property', 'concept', 'period-type'): lambda f: f.concept.periodType,
#     ('property', 'concept', 'balance'): index_property_balance,
#     ('property', 'concept', 'data-type'): lambda f: f.concept.typeQname,
#     ('property', 'concept', 'base-type'): lambda f: f.concept.baseXbrliTypeQname,
#     ('property', 'concept', 'is-monetary'): lambda f: f.concept.isMonetary,
#     ('property', 'concept', 'is-numeric'): lambda f: f.concept.isNumeric,
#     ('property', 'concept', 'substitution'): lambda f: f.concept.substitutionGroupQname,
#     ('property', 'concept', 'namespace-uri'): lambda f: f.concept.qname.namespaceURI,
#     ('property', 'concept', 'local-name'): lambda f: f.concept.qname.localName,
#     ('property', 'concept', 'is-abstract'): lambda f: f.concept.isAbstract,
#     ('property', 'concept', 'id'): lambda f: f.id,
#     ('property', 'period', 'start'): index_property_start,
#     ('property', 'period', 'end'): index_property_end,
#     ('property', 'period', 'days'): index_property_days,
#     ('property', 'entity', 'scheme'): lambda f: f.context.entityIdentifier[0],  # entityIdentifier[0] is the scheme
#     ('property', 'entity', 'id'): lambda f: f.context.entityIdentifier[1]
# # entityIdentifer[1] is the id
# }

# def index_table_properties(xule_context):
#     """"Add the table properites to the fact index

#     :type xule_context: XuleRuleContext
#     """
#     # Go through each table.
#     for cube_base in XuleDimensionCube.base_dimension_sets(xule_context.model):
#         cube = XuleDimensionCube(xule_context.model, *cube_base, include_facts=True)
#         xule_context.global_context.fact_index[('builtin', 'cube')][cube] |= cube.facts
#         xule_context.global_context.fact_index[('property', 'cube', 'name')][cube.hypercube.qname] |= cube.facts
#         xule_context.global_context.fact_index[('property', 'cube', 'drs-role')][cube.drs_role.roleURI] |= cube.facts

# def get_decimalized_value(fact_a, fact_b, xule_context):
#     """Adjust 2 fact values based on accuracy.
    
#     :param fact_a: First fact
#     :type fact_a: ModelFact
#     :param fact_b: Second fact
#     :type fact_b: ModelFact
#     :returns: A tuple of the rounded fact values and the new decimals value for each
#     :rtype: tuple
    
#     Round the fact values to the minimum accuracy defined by the decimals attribute of the facts. 
    
#     Example:
#         Fact value of 1,234,567 with decimals -3 is rounded to 1,235,000
        
#     Arguments:
#         fact_a (ModelFact): First fact
#         fact_b (ModelFact): Second fact
#         xule_context (XuleRuleContext): Processing context
#     """
#     fact_a_decimals = get_decimals(fact_a, xule_context)
#     fact_b_decimals = get_decimals(fact_b, xule_context)

#     min_decimals = min(fact_a_decimals, fact_b_decimals)

#     fact_a_value = fact_a.xValue if fact_a_decimals == float('inf') else round(fact_a.xValue, min_decimals)
#     fact_b_value = fact_b.xValue if fact_b_decimals == float('inf') else round(fact_b.xValue, min_decimals)

#     return fact_a_value, fact_a_decimals, fact_b_value, fact_b_decimals


# def get_decimals(fact, xule_context):
#     """Return the decimals of a fact as a number.
    
#     :param fact: The fact to get the accuracy from
#     :type fact: ModelFact
#     :param xule_context: Rule processing context
#     :type xule_context: XuleRuleContext
#     """
#     if fact.decimals is None:
#         return float('inf')

#     if fact.decimals.strip() == 'INF':
#         return float('inf')
#     else:
#         try:
#             return int(fact.decimals)
#         except ValueError:
#             raise XuleProcessingError(_("%s Fact contains invalid decimal value of %s" % (fact.qname, fact.decimals)),
#                                       xule_context)


def evaluate(rule_part, xule_context, trace_dependent=False, override_table_id=None):
    """General evaluator for an expression.
    
    :param rule_part: The expression being evaluated
    :type rule_part: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :param trace_dependent: Debugging indicator
    :type trace_dependent: bool
    :param override_table_id: A table id to use instead of the table id of the rule_part.
    :type override_table_id: int
    
    This is the main evaluator for evlatuating rule expressions. If the rule_part is an iterable expression it will be evaluated. The returned
    values will be put on the iteration table and a single value will be selected as the value of the expression for the current iteration. If the
    the rule_part is a singleton expression, then it will be evaluated and the value returned. 
    
    Each type of expression (i.e. assertion, if statement, for loop, literal interger, add operation) has its own evaluator. This evaluator will call
    the appriate evaluator for the expression.
    
    This evaluator handles caching expression evaluations for performance.
    
    This evaluator also includes capturing information about the evaluation for debugging purposes.
    """

    if xule_context.iter_count > xule_context.global_context.maximum_iterations:
        raise XuleProcessingError('Rule has run too many iterations, either remove catesian products in the rule or the instance or increase the iterations using --xule-max-rule-iterations option.')

    try:
        # Setup trace information.
        if getattr(xule_context.global_context.options, "xule_trace", False) or getattr(
                xule_context.global_context.options, "xule_trace_count", False):
            trace_is_dependent = "D" if trace_dependent else " "
            trace_source = "U"
            trace_written = False
        if getattr(xule_context.global_context.options, "xule_trace_count", False):
            if rule_part['node_id'] not in xule_context.expression_trace:
                xule_context.expression_trace[rule_part['node_id']] = {'iterations': 1,  # total
                                                                       'iterations-t': datetime.timedelta(0),
                                                                       'U': 0,
                                                                       # unknown iterations - should always be none
                                                                       'U-t': datetime.timedelta(0),
                                                                       'E': 0,  # evaluated iterables
                                                                       'E-t': datetime.timedelta(0),
                                                                       'c': 0,  # from cache
                                                                       'c-t': datetime.timedelta(0),
                                                                       'T': 0,  # from table
                                                                       'T-t': datetime.timedelta(0),
                                                                       'e': 0,  # evaluate non iterable
                                                                       'e-t': datetime.timedelta(0),
                                                                       'R': 0,  # re-evaluate
                                                                       'R-t': datetime.timedelta(0),
                                                                       'r': 0,  # re-evaluate non iterable
                                                                       'r-t': datetime.timedelta(0),
                                                                       'isE': 0,
                                                                       # iteration stop on evaluation of iterable
                                                                       'isE-t': datetime.timedelta(0),
                                                                       'ise': 0,
                                                                       # iteration stop on evaluate of non iterable
                                                                       'ise-t': datetime.timedelta(0),
                                                                       'isu': 0,
                                                                       # iteration stop on unbound during post value processing
                                                                       'isu-t': datetime.timedelta(0),
                                                                       'ex': 0,  # exception during iteration evaluation
                                                                       'ex-t': datetime.timedelta(0),
                                                                       'name': rule_part['exprName']
                                                                       }
            else:
                xule_context.expression_trace[rule_part['node_id']]['iterations'] += 1
            expression_trace_start = datetime.datetime.today()

        #processing_id = xule_context.get_processing_id(rule_part['node_id'])
        processing_id = xule_context.get_column_id(rule_part['node_id'])
        rule_part_name = rule_part['exprName']
        # trace
        if getattr(xule_context.global_context.options, "xule_trace", False):
            xule_context.trace_level += 1

            trace = "  " * xule_context.trace_level
            trace += rule_part_name + " " + str(processing_id)  # + " " + str(rule_part)
            xule_context.global_context.message_queue.print(f">{trace_is_dependent} {processing_id} " + trace.replace("\n", " "), "xule.ruleTrace")

        has_key = False



        if ('is_iterable' in rule_part):
            # is_iterable is always true if it is present, so don't need to check the actual value
            xule_context.used_expressions.add(processing_id)

            if 'is_dependent' in rule_part:
                is_dependent = rule_part['is_dependent']

                value = xule_context.iteration_table.current_value(processing_id, xule_context)
                if value is None:
                    # Will evaluate or get from cache.
                    values = None
                    if is_dependent:
                        xule_context.iteration_table.current_table.make_dependent()
                    try:
                        if getattr(xule_context.global_context.options, "xule_no_cache", False):
                            # if xule_context.global_context.no_cache:
                            values = None
                        else:
                            if (rule_part.get('table_id') != xule_context.iteration_table.main_table_id or
                                    is_dependent):
                                local_cache_key = get_local_cache_key(rule_part, xule_context)
                                if local_cache_key is not None:
                                    
                                    has_key = True
                                    
                                    values = xule_context.local_cache.get(local_cache_key)
                                else:
                                    values = None
                        if values is None:
                            # if has_key:
                            #     print("Not I", rule_part['node_id'],rule_part['exprName'])

                            values = EVALUATOR[rule_part_name](rule_part, xule_context)
                            trace_source = "E"
                            if not getattr(xule_context.global_context.options, "xule_no_cache", False):
                                # if not xule_context.global_context.no_cache:
                                if (rule_part.get('table_id') != xule_context.iteration_table.main_table_id or
                                        is_dependent):
                                    local_cache_key = get_local_cache_key(rule_part, xule_context)
                                    if local_cache_key is not None:
                                        # print("caching", rule_part['node_id'], [(x[0], x[1].format_value()[:10]) for x in local_cache_key[1]], len(values.values))
                                        xule_context.local_cache[local_cache_key] = values
                        else:
                            # if has_key:
                            #     print("Found I", rule_part['node_id'], rule_part['exprName'])

                            # print("using cache", rule_part['node_id'], [x[0] for x in local_cache_key[1]])
                            trace_source = "c"
                    except XuleIterationStop:
                        if getattr(xule_context.global_context.options, "xule_trace", False):
                            # if xule_context.show_trace:
                            xule_context.trace_level -= 1
                        if getattr(xule_context.global_context.options, "xule_trace_count", False):
                            # if xule_context.show_trace_count:
                            xule_context.expression_trace[rule_part['node_id']][
                                'iterations-t'] += datetime.datetime.today() - expression_trace_start
                            xule_context.expression_trace[rule_part['node_id']][
                                'isE-t'] += datetime.datetime.today() - expression_trace_start
                            xule_context.expression_trace[rule_part['node_id']]['isE'] += 1
                            trace_written = True
                        raise
                    except XuleReEvaluate:
                        trace_source = 'R'
                        raise
                    else:
                        # add - add values to expression cache
                        xule_context.iteration_table.add_column(rule_part, override_table_id or rule_part['table_id'],
                                                                processing_id, values, xule_context)
                        value = xule_context.iteration_table.current_value(processing_id, xule_context)
                        # The tags on the value may not apply to this iteration.  For exmaple, if the expression is not dependent, then it will
                        # be evaluated once and stored in the local cache with the tags from the first evaluation.
                        if value is not None and value.tags is not None:
                            new_tags = value.tags.copy()
                            new_tags.update(xule_context.tags)
                            value.tags = new_tags
                else:
                    trace_source = "T"
                    # The tags on the value may not apply to this iteration.  The value was pulled from the iteration table. It
                    # will have tags from the previously calculated value.
                    if value is not None and value.tags is not None:
                        new_tags = value.tags.copy()
                        new_tags.update(xule_context.tags)
                        value.tags = new_tags
            else:
                raise XuleProcessingError(
                    _("Internal error: Found iterable (%s) that does not have a dependency flag." % rule_part_name),
                    xule_context)
        else:  # is_iterable
            trace_source = "e"
            # Check the cache - only if the expression does have something in it that produces multiple results and its not a varRef.
            if getattr(xule_context.global_context.options, "xule_no_cache", False):
                local_cache_key = None
            else:
                if rule_part['number'] == 'single' and rule_part['exprName'] not in ('varRef', 'tagRef'):
                    local_cache_key = get_local_cache_key(rule_part, xule_context)
                else:
                    local_cache_key = None
            if local_cache_key is None:
                value = None
            else:

                has_key = True

                cache_value = xule_context.local_cache.get(local_cache_key)
                value = cache_value.clone() if cache_value is not None else None
                # The tags on the value may not apply to this iteration.  For exmaple, if the expression is not dependent, then it will
                # be evaluated once and stored in the local cache with the tags from the first evaluation.
                if value is not None and value.tags is not None:
                    new_tags = value.tags.copy()
                    new_tags.update(xule_context.tags)
                    value.tags = new_tags

            # if value is not None and has_key:
            #     print("Found N", rule_part['node_id'], rule_part['exprName'])

            if value is None:

                # if has_key:
                #     print("Not N", rule_part['node_id'], rule_part['exprName'])

                try:
                    value = EVALUATOR[rule_part_name](rule_part, xule_context)
                except XuleIterationStop:
                    if getattr(xule_context.global_context.options, "xule_trace", False):
                        # if xule_context.show_trace:
                        xule_context.trace_level -= 1
                    if getattr(xule_context.global_context.options, "xule_trace_count", False):
                        # if xule_context.show_trace_count:
                        xule_context.expression_trace[rule_part['node_id']][
                            'iterations-t'] += datetime.datetime.today() - expression_trace_start
                        xule_context.expression_trace[rule_part['node_id']][
                            'ise-t'] += datetime.datetime.today() - expression_trace_start
                        xule_context.expression_trace[rule_part['node_id']]['ise'] += 1
                        trace_written = True
                    raise
                except XuleReEvaluate as e:
                    trace_source = 'r'
                    raise

                if not getattr(xule_context.global_context.options, "xule_no_cache", False):
                    if value is not None:
                        local_cache_key = get_local_cache_key(rule_part, xule_context)
                        if local_cache_key is not None:
                            # The cache value is cloned so it is not corrupted by further processing after this point.
                            # Copy the value and copy the exisiting tags
                            cache_value = value.clone()
                            cache_value.tags = copy.copy(xule_context.tags)
                            cache_value.facts = copy.copy(xule_context.facts)
                            #cache_value.aligned_result_only = xule_context.aligned_result_only
                            #cache_value.used_expressions = copy.copy(xule_context.used_expressions)
                            xule_context.local_cache[local_cache_key] = cache_value

        # If the look_for_alignment flag is set, check if there is now alignment after adding the column. This is used in 'where' clause processing.
        if (xule_context.look_for_alignment and
                # rule_part.has_alignment and
                value.aligned_result_only and
                rule_part.get('table_id') in xule_context.where_table_ids and
                rule_part['node_id'] in xule_context.where_dependent_iterables):
            raise XuleReEvaluate(xule_context.iteration_table.any_alignment)

        if getattr(xule_context.global_context.options, "xule_trace", False):
            # if xule_context.show_trace:
            sugar = sugar_trace(value, rule_part, xule_context)
            trace_info = (xule_context.trace_level, rule_part_name, sugar, value)
            xule_context.trace.appendleft(trace_info)

            post_trace = "  " * xule_context.trace_level
            post_trace += ("NONE" if value is None else value.format_value()) + format_trace_info(trace_info[1],
                                                                                                  trace_info[2], {},
                                                                                                  xule_context)
            xule_context.global_context.message_queue.print(f"<{trace_is_dependent}{trace_source}{processing_id}" + post_trace.replace("\n", " "), "xule.ruleTrace")

            xule_context.trace_level -= 1
        try:
            value = post_evaluate_value(rule_part, value, xule_context)
        finally:
            if getattr(xule_context.global_context.options, "xule_trace_count", False):
                # if xule_context.show_trace_count:
                xule_context.expression_trace[rule_part['node_id']][trace_source] += 1
                xule_context.expression_trace[rule_part['node_id']]['iterations-t'] += (
                            datetime.datetime.today() - expression_trace_start)
                xule_context.expression_trace[rule_part['node_id']][trace_source + '-t'] += (
                            datetime.datetime.today() - expression_trace_start)
                trace_written = True
    finally:
        if getattr(xule_context.global_context.options, "xule_trace_count", False) and not trace_written:
            xule_context.expression_trace[rule_part['node_id']][
                'iterations-t'] += datetime.datetime.today() - expression_trace_start
            xule_context.expression_trace[rule_part['node_id']][trace_source + '-t'] += (
                        datetime.datetime.today() - expression_trace_start)
            xule_context.expression_trace[rule_part['node_id']][trace_source] += 1
            trace_written = True
    return value


def post_evaluate_value(rule_part, value, xule_context):
    """Track tags and facts for the evaluated value.
    
    :param rule_part: The expression being evaluated
    :type rule_part: dict
    :param value: The evaluated value
    :type value: XuleValue
    :param xule_context: The rule processing context
    :type xule_context: XuleRuleContext
    """
    if value is None:
        raise XuleIterationStop(XuleValue(xule_context, None, 'unbound'))
        # value = XuleValue(xule_context, None, 'unbound')

    if value.fact is not None:
        # xule_context.facts.append(value.fact)
        xule_context.facts[value.fact] = None
    if value.facts is not None:
        #         print("before", len(xule_context.facts), len(set(xule_context.facts)))
        #         xule_context.facts.extend(value.facts)
        #         print("after", len(xule_context.facts), len(set(xule_context.facts)))
        xule_context.facts.update(value.facts)
    if value.tags is not None:
        #         # Need to make sure that the current version of the tags stay and the update only adds new tags from the value.
        #         new_tags = value.tags.copy()
        #         new_tags.update(xule_context.tags)
        #         xule_context.tags = new_tags
        xule_context.tags.update(value.tags)
    if value.aligned_result_only == True:
        xule_context.aligned_result_only = True

    #     if value.used_vars is not None:
    #         new_keys = value.used_vars.keys() - xule_context.vars.keys()
    #         for new_key in new_keys:
    #             xule_context.vars[new_key] = value.used_vars[new_key]

    if value.used_expressions is not None:
        # print("add",rule_part['exprName'], rule_part['node_id'], len(xule_context.used_expressions), len(value.used_expressions))
        xule_context.used_expressions.update(value.used_expressions)

    if value.type == 'unbound':
        raise XuleIterationStop(value)

    return value


def get_local_cache_key(rule_part, xule_context):
    """Get a cache key for storing a value in the cache
    
    :param rule_part: xule expression
    :type rule_part: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    """
    # Don't cache for function refs that are not cacheable
    if (rule_part['exprName'] in ('functionReference', 'macroRef') and rule_part.get('cacheable') != True) or rule_part[
        'exprName'] == 'forBodyExpr':
        return None

    dep_var_index = set()
    for var_ref in rule_part['var_refs']:
        # var_ref tuple: 0 = var declaration id, 1 = var name, 2 = var ref ast, 3 = var type (1=block variable or 'for' variable, 2=constant, 3=function argument, factset variable)
        var_info = xule_context.find_var(var_ref[1], var_ref[0])
        if var_info['type'] == xule_context._VAR_TYPE_CONSTANT:

            if var_info['calculated']:
                const_values = var_info['value']
                # if the constant only have one alignment which is None and there is only one value, then thats the value, there is no
                # other option
                if len(const_values.values) == 1 and None in const_values.values and len(const_values.values[None]) == 1:
                    const_value = const_values.values[None][0]                      
                else:
                    # Here there is more than one value for the constant, so find it on the iteration table.
                    const_value = xule_context.iteration_table.current_value(var_ref[0], xule_context)
                if const_value is not None:
                    dep_var_index.add((var_info['name'], const_value.hashable_system_value))  


            # if val is None: # we are checking the cache
            #     if var_info['calculated']:
            #         #const_processing_id = xule_context.get_column_id(var_ref[2]['var_declaration'])
            #         #const_value = xule_context.iteration_table.current_value(const_processing_id, xule_context)
            #         const_value = evaluate(var_info['expr'],
            #                            xule_context,
            #                            override_table_id=var_ref[2]['table_id'])
            #         if const_value is not None:
            #         # const_value = evaluate(var_info['expr'],
            #         #                    xule_context,
            #         #                    override_table_id=var_ref[2]['table_id'])
            #             dep_var_index.add((var_info['name'], const_value.hashable_system_value))
            #         else:
            #             # In this case I have a constant that the rule part is dependent on, but there is no value for, so
            #             # we cannot rely on the cache key because it is incomplete, so return None which will force
            #             # doing the evaluation.
            #             return None
            # else: # get a cache key to store into the cache
            #     if var_info['calculated']:
            #         const_values = var_info['value']
            #         if len(const_values.values) == 1 and None in const_values.values and len(const_values.values[None]) == 1:
            #             const_value = const_values.values[None][0]                      
            #         else:
            #             const_value = xule_context.iteration_table.current_value(var_ref[0], xule_context)
            #         if const_value is not None:
            #             dep_var_index.add((var_info['name'], const_value.hashable_system_value))  



                # if isinstance(val, XuleValueSet):
                #     # collect all the used expressions from the values in the value set
                #     used_expressions = set()
                #     for values in val.values.values():
                #         for value in values:
                #             used_expressions.update(value.used_expressions or set())
                #     if var_info['expr']['node_id'] in used_expressions | xule_context.used_expressions:
                #         const_value = evaluate(var_info['expr'],
                #                        xule_context,
                #                        override_table_id=var_ref[2]['table_id'])
                #         dep_var_index.add((var_info['name'], const_value.hashable_system_value))
                # else: # val is a single XuleValue
                #     if var_info['calculated']:
                #     #if var_info['expr']['node_id'] in (val.used_expressions or set())| xule_context.used_expressions:
                #         const_value = var_info['value']
                #         # const_value = evaluate(var_info['expr'],
                #         #                xule_context,
                #         #                override_table_id=var_ref[2]['table_id'])
                #         dep_var_index.add((var_info['name'], const_value.hashable_system_value))

            # if xule_context.get_processing_id(var_info['expr']['node_id']) in xule_context.used_expressions:
            #     const_value = evaluate(var_info['expr'],
            #                            xule_context,
            #                            override_table_id=var_ref[2]['table_id'])
            #     dep_var_index.add((var_info['name'], const_value))

        else:
            if var_ref[0] in xule_context.vars:
                if var_info['calculated']:
                    dep_var_index.add((var_info['name'], var_info['value']))

    # specail handling for for loops. This ensures that the for variable is included in the cache key. If the body of the for
    # loop did not use the for variable it would not in the list dependent vars list for the expression and so it would not
    # be included in the cache key.
    for for_var in xule_context.find_for_vars():
        dep_var_index.add((for_var['name'], for_var['value']))

    alignment = xule_context.iteration_table.current_alignment if rule_part['has_alignment'] else None

    cache_key = (rule_part['node_id'], frozenset(dep_var_index), alignment)
    return cache_key


def evaluate_assertion(assert_rule, xule_context):
    """Evaluator for an assertion rule.
    
    :param assert_rule: Rule expression for an assertion
    :type assert_rule: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    
    This evaluator evaluates an assertion rule. This evalutor will keep evaluting the assertion rule until the iteration table is empty. This 
    is how it can produce multiple messages for a single rule. For example, a rule @Assets < 0 will produce a message for each individual
    Asset value that is less than zero. If there are 3 values for Assets and 2 of them are less than zero, it will produce 2 messages. 
    
    An assertion will evaluate to a boolean value. If the assertion is marked as 'satisfied' and the
    evaluated value is true, the assertion will produce a message. If the assertion is marked as unstatisfied and the evaluated value is false, 
    the assertion will produce a message.
    """

    # Keep evaluating the rule while there are iterations. This is done in a While True loop so there is always at least one iteration. This is for rules that 
    # do not have iterable expressions in them (i.e. 1 + 2).
    rule_start = datetime.datetime.today()
    while True:
        xule_context.iter_count += 1
        try:
            xule_value = evaluate(assert_rule['body'], xule_context)
        except XuleIterationStop:
            xule_context.iter_skip_count += 1
            pass
        except:
            xule_context.iter_except_count += 1
            raise
        else:
            # There were no exceptions. Check the results of the rule and create the message.

            # Check if the rule expects only aligned values. When a rule has aligned values then the none aligned results are ignored.
            # This prevents a rule like 1 + @Assets from producing a result when there are no Assets in a filing. When this happens, the @Assetss will have 
            # a none aligned unbound value. The plus operation will treat this as 0 and produce a value of 1. However, this value should not result in a message.
            # When a factset is evluated, it trips the aligned_result_only flag. 
            if xule_value.type != 'unbound' and not (
                    xule_context.iteration_table.current_alignment is None and xule_context.aligned_result_only):

                if xule_value.type != 'bool':
                    raise XuleProcessingError(_("Raise %s did not evaluate to a boolean, found '%s'." % (
                    xule_context.rule_name, xule_value.type)), xule_context)

                # Determine if a message should be sent
                send_message = ((assert_rule['satisfactionType'] == 'satisfied' and xule_value.value == True) or
                                (assert_rule['satisfactionType'] == 'unsatisfied' and xule_value.value == False))

                if send_message:
                    xule_context.iter_message_count += 1
                    messages = dict()
                    # Process each of the results in the rule. The Results are the messages that are produced.
                    for rule_result in assert_rule.get('results', list()):
                        messages[rule_result['resultName']] = result_message(assert_rule, rule_result, xule_value,
                                                                             xule_context)

                    # get severity
                    if 'severity' not in messages:
                        # default severity
                        messages['severity'] = 'error'
                    severity = messages['severity']

                    # message - this is the main message
                    main_message = messages['message'].value if 'message' in messages else XuleString(
                        'No message supplied')
                    messages.pop('message', None)

                    full_rule_name = xule_context.rule_name
                    # Handle rule suffix
                    if 'rule-suffix' in messages:
                        full_rule_name += '.' + messages['rule-suffix']

                    filing_url = xule_context.model.modelDocument.uri if xule_context.model is not None else ''
                    # The rule_focus is the model object that is the focus fo the rule. This can be a modelFact, modelConcept or modelDocument.
                    # It is used by the logger to provide additional location information about the thing (i.e. fact) that is the focus of the 
                    # message fom the rule.

                    rule_focus = messages.pop('rule-focus', None)
                    if rule_focus is None:
                        rule_focus = next(iter(xule_context.facts.keys()), None)

                    # Prep the main_message for the logger. The logger wants a %-style format string and the substitutions passed as named arguments.
                    if isinstance(main_message, XuleString):
                        format_string_message = main_message.format_string
                        substitutions = main_message.substitutions
                    else:
                        format_string_message = main_message
                        substitutions = dict()

                    # combine the substitutions and the messages dictionary
                    messages.update(substitutions)

                    xule_context.global_context.message_queue.log(severity.upper(),
                                                                  full_rule_name,
                                                                  _(format_string_message),
                                                                  # sourceFileLine=source_location,
                                                                  filing_url=filing_url,
                                                                  modelObject=rule_focus,
                                                                  **messages)

                else:
                    xule_context.iter_pass_count += 1
            else:
                xule_context.iter_misaligned_count += 1

        # xule_context.iteration_table.del_current()
        # if xule_context.iteration_table.is_empty:
        xule_context.iteration_table.next(assert_rule['node_id'])
        if xule_context.iteration_table.is_table_empty(assert_rule['node_id']):
            break
        else:
            xule_context.reset_iteration()

    # Rule Statistics
    handle_stats(xule_context, rule_start, datetime.datetime.today(), 'assert')  

def evaluate_output_rule(output_rule, xule_context):
    """Evaluator for an output rule.
    
    :param output_rule: Rule expression for an assertion
    :type output_rule: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    
    This evaluator evaluates an output rule. This evalutor will keep evaluting the output rule until the iteration table is empty. This 
    is how it can produce multiple messages for a single rule. 
    
    An output rule will produce a value and then create a message based on the evaluated value.
    """
    # Keep evaluating the rule while there are iterations. This is done in a While True loop so there is always at least one iteration. This is for rules that 
    # do not have iterable expressions in them (i.e. 1 + 2).
    rule_start = datetime.datetime.today()
    while True:
        xule_context.iter_count += 1
        try:
            xule_value = evaluate(output_rule['body'], xule_context)
        except XuleIterationStop:
            xule_context.iter_skip_count += 1
            pass
        except:
            xule_context.iter_except_count += 1
            raise
        else:
            # There were no exceptions. Check the results of the rule and create the message.

            # Check if the rule expects only aligned values. When a rule has aligned values then the none aligned results are ignored.
            # This prevents a rule like 1 + @Assets from producing a result when there are no Assets in a filing. When this happens, the @Assetss will have 
            # a none aligned unbound value. The plus operation will treat this as 0 and produce a value of 1. However, this value should not result in a message.
            # When a factset is evluated, it trips the aligned_result_only flag. 
            if xule_value.type != 'unbound' and not (
                    xule_context.iteration_table.current_alignment is None and xule_context.aligned_result_only):
                # Determine if a message should be sent
                xule_context.iter_message_count += 1
                # produce file output
                result_file(output_rule, xule_value, xule_context)

                messages = dict()
                # Process each of the results in the rule. The Results are the messages that are produced.
                for rule_result in output_rule.get('results', list()):
                    messages[rule_result['resultName']] = result_message(output_rule, rule_result, xule_value,
                                                                         xule_context)

                # get severity
                if 'severity' not in messages:
                    # default severity
                    messages['severity'] = 'info'
                severity = messages['severity']
                # message - this is the main message

                # Check if there is an instane output. If so, then don't produce a message unless there is an explicit "message" output.
                #has_instance_output = process_instance_output(xule_context, messages)
                main_message = messages.get('message', xule_value) # This will default the message to the value of the rule.
                if main_message.type == 'string':
                    main_message = main_message.value
                else:
                    main_message = main_message.format_value()

                messages.pop('message', None)
            

                full_rule_name = xule_context.rule_name
                # Handle rule suffix
                if 'rule-suffix' in messages:
                    full_rule_name += '.' + messages['rule-suffix']

                filing_url = xule_context.model.modelDocument.uri if xule_context.model is not None else ''
                # The rule_focus is the model object that is the focus fo the rule. This can be a modelFact, modelConcept or modelDocument.
                # It is used by the logger to provide additional location information about the thing (i.e. fact) that is the focus of the 
                # message fom the rule.
                rule_focus = messages.pop('rule-focus', None)
                if rule_focus is None:
                    rule_focus = next(iter(xule_context.facts.keys()), None)

                # Prep the main_message for the logger. The logger wants a %-style format string and the substitutions passed as named arguments.
                if isinstance(main_message, XuleString):
                    format_string_message = main_message.format_string
                    substitutions = main_message.substitutions
                else:
                    format_string_message = main_message
                    substitutions = dict()

                # combine the substitutions and the messages dictionary
                messages.update(substitutions)

                xule_context.global_context.message_queue.log(severity.upper(),
                                                              full_rule_name,
                                                              _(format_string_message),
                                                              # sourceFileLine=source_location,
                                                              filing_url=filing_url,
                                                              modelObject=rule_focus,
                                                              **messages)
            else:
                xule_context.iter_misaligned_count += 1

        # xule_context.iteration_table.del_current()
        # if xule_context.iteration_table.is_empty:
        xule_context.iteration_table.next(output_rule['node_id'])
        if xule_context.iteration_table.is_table_empty(output_rule['node_id']):
            break
        else:
            xule_context.reset_iteration()

    # Rule Statistics
    handle_stats(xule_context, rule_start, datetime.datetime.today(), 'output')    

def handle_stats(xule_context, rule_start, rule_end, rule_type):

    if (getattr(xule_context.global_context.options, "xule_rule_stats_log", False) or 
        getattr(xule_context.global_context.options, "xule_rule_stats_file", None) is not None):

        stats = {'rule_name': xule_context.rule_name,
                 'rule_type': rule_type,
                'start_time': rule_start.isoformat(),
                'end_time': rule_end.isoformat(),
                'time': str(rule_end - rule_start),
                'total': xule_context.iter_count,
                'pass': xule_context.iter_pass_count,
                'skip': xule_context.iter_skip_count,
                'message': xule_context.iter_message_count,
                'misaligned': xule_context.iter_misaligned_count,
                'except': xule_context.iter_except_count
        }

        if getattr(xule_context.global_context.options, "xule_rule_stats_log", False):
            xule_context.global_context.message_queue.log('INFO',
                                                        f"stats-{xule_context.rule_name}",
                                                        json.dumps(stats),
                                                        filing_url=xule_context.model.modelDocument.uri if xule_context.model is not None else '',
                                                        **stats)
        
        if getattr(xule_context.global_context.options, "xule_rule_stats_file", None) is not None:
            xule_context.global_context.stats.append(stats)

def evaluate_bool_literal(literal, xule_context):
    """Evaluator for literal boolean expressions
    
    :param literal: Rule expression
    :type literal: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue
    
    A boolean literal is either 'true' or 'false'.
    """
    if literal['value'] == "true":
        return XuleValue(xule_context, True, 'bool')
    elif literal['value'] == "false":
        return XuleValue(xule_context, False, 'bool')
    else:
        raise XuleProcessingError(_("Invalid boolean literal found: %s" % literal.value), xule_context)

def evaluate_period_literal(literal, xule_context):
    """Evaluate a period literal

    :param literal: Rule expression
    :type literal: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue

    Currently, the only type of period literal is 'forever'. Other periods (instance, duration) are created using
    the period() function.
    """

    if literal.get('forever', False):
        return XuleValue(xule_context, (datetime.datetime.min, datetime.datetime.max), 'duration')
    else:
        return XuleValue(xule_context, None, 'none')

def evaluate_string_literal(literal, xule_context):
    """Evaluate a string literal

    :param literal: Rule expression
    :type literal: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue
    
    A string can consist of a string of characters, escaped characters or an expression to evaluate. The literal will contain a list
    of these components that make up the string. For example: 
    
            "The value of the rule is {$rule-value}.\nThis is based on the fact value {$fact}.".
    
    In this example the literal would be a list of:
        * string of characters: "The value of the rule is "
        * an expression: $rule-value
        * string of characters: "."
        * escape character: "\n"
        * string of characters: "This is based on the fact value "
        * an expression: $fact
        * string of characters: "."
    
    This evaluator will evaluate all the components of the string literal and concatenate them to a string.
    """
    # result_string = ''
    # The underlying value for the XuleValue that is created from this evaluator is a XuleString. A XuleString is subclassed
    # from a python str. A XuleString stores a format string along with the subsitutions. It will create the formatted string and set that
    # as the value of the XuleSting. This way it will act and feel like a python string but will contain the original format string and
    # subsitutiions. Having the format string and substitutions separate is usefuly when logging messages to arelle.
    format_string = ''
    substitutions = []
    sub_num = 0

    for string_item in literal['stringList']:
        if string_item['exprName'] == 'baseString':
            format_string += string_item['value']
        elif string_item['exprName'] == 'escape':
            if string_item['value'] == 'n':
                format_string += '\n'
            elif string_item['value'] == 't':
                format_string += '\t'
            else:
                format_string += string_item['value']
        else:
            # This is an expression.
            expr_value = evaluate(string_item, xule_context)
            # The result of the expression is not directly put in the format string. Instead a substitution is used
            sub_name = 'sub{}'.format(sub_num)
            sub_num += 1
            # Substitutions is a list of a 3 part tuple 0=location in format string, 1=substitution name, 2=substitution value
            substitutions.append((len(format_string), sub_name, expr_value.format_value()))

    return XuleValue(xule_context, XuleString(format_string, substitutions), 'string')


def evaluate_int_literal(literal, xule_context):
    """Evaluator for literal integer expressions
    
    :param literal: Rule expression
    :type literal: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue
    """
    return XuleValue(xule_context, int(literal['value']), 'int')


def evaluate_float_literal(literal, xule_context):
    """Evaluator for literal float expressions
    
    :param literal: Rule expression
    :type literal: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue
    """
    return XuleValue(xule_context, float(literal['value']), 'float')


def evaluate_void_literal(literal, xule_context):
    """Evaluator for literal void expressions
    
    :param literal: Rule expression
    :type literal: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue
    
    A void expression is either 'none' or 'skip'.
    """
    return XuleValue(xule_context, None, 'none' if literal['value'] == 'none' else 'unbound')


def evaluate_qname_literal(literal, xule_context):
    """Evaluator for literal qname expressions
    
    :param literal: Rule expression
    :type literal: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue
    """
    prefix = literal['prefix']
    return XuleValue(xule_context,
                     QName(prefix if prefix != '*' else None, literal['namespace_uri'], literal['localName']), 'qname')

def evaluate_namespace_group(ns_group, xule_context):
    """Evaluator for literal group qname expressions
    This is like a qname, but the prefix does not resolve to a specific namespace. Instead, the local name can come from a
    list of namespaces. This is used for refering to the different versions of a taxonomy with a single prefix. I.E. 
    usgaap could be a namespace group prefix for all the different versions of the USGAAP taxonomuy. The va
    
    :param literal: Rule expression
    :type literal: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue
    """

    ns_group_info = xule_context.find_namespace_group(ns_group['prefix'])
    if not ns_group_info['calculated']:
        ns_list = evaluate(ns_group_info['expr']['body'], xule_context, override_table_id=xule_context.iteration_table.current_table.table_id)
        if ns_list.type not in ('set', 'list', 'string'):
            raise XuleProcessingError(_("Value of a namespace group must be a string or a set or list of strings, found '%s'" % ns_list.type), xule_context)
        if ns_list.type in ('list', 'set'):
            for fragment in ns_list.value:
                if fragment.type != 'string':
                    raise XuleProcessingError(_("The values in a namespace group set or list must be strings, found '%s'"% fragment.type), xule_context)

        ns_group_info['calculated'] = True
        ns_group_info['value'] = ns_list

    return XuleValue(xule_context, (ns_group_info['value'], ns_group['localName']), 'groupQname')

def evaluate_severity(severity_expr, xule_context):
    """Evaluator for literal severity expressions
    
    :param literal: Rule expression
    :type literal: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue
    """
    return XuleValue(xule_context, severity_expr['value'], 'severity')


def evaluate_aspect_name(literal, xule_context):
    """Evaluator for literal aspect name expressions
    
    :param literal: Rule expression
    :type literal: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue
    
    An aspect name is literal is one of the build in aspects for a factset and is one of 'concept', 'unit', 'entity' or 'period'.
    """
    return XuleValue(xule_context, literal['value'], 'aspect_name')


def evaluate_string_keyword(expr, xule_context):
    """Evaluator for literal string based keywords expressions
    
    :param expr: Rule expression
    :type expr: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue
    
    Some keywords are evaluated as strings. This is used for balance types ('credit', 'debit') and period type ('instant', 'duration').
    """
    return XuleValue(xule_context, expr['value'], 'string')


def evaluate_tagged(tagged_expr, xule_context):
    """Evaluator for tagged expressions
    
    :param tagged_expr: Rule expression
    :type tagged_expr: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue
    
    A tagged expression is an expression followed by a # sign and a tag name. The evaluated value of the tagged expression
    is added to the rule processing context by its name. The tags can be used in creating messages.
    """
    try:
        tagged_value = evaluate(tagged_expr['expr'], xule_context)
    except XuleIterationStop as xis:
        xule_context.tags[tagged_expr['tagName']] = xis.stop_value.tag
        raise
    else:
        # xule_context.tags[tagged_expr.tagName] = tagged_value
        if tagged_value.tags is None:
            tagged_value.tags = {tagged_expr['tagName']: tagged_value}
        else:
            tagged_value.tags[tagged_expr['tagName']] = tagged_value

    return tagged_value


def tag_default_for_factset(aspect_filters, xule_context):
    """Get the value of the concept aspect for tagging the default None fact of a factset"""

    for aspect_info, aspect_value in aspect_filters.items():
        # aspect_inf is the aspect_info that is the key to the aspect_filters dictionary
        # aspect_info is a tuple. The 0 = type (builtin or explicit_dimension, 1 = aspect name, 2 = wildcard, 3 = operator, properties

        if aspect_info[0] == 'builtin' and aspect_info[1] == 'concept':
            if aspect_info[2] is None:  # there isn't a wildcard
                if aspect_info[3] == '=':  # the operator is '=' and it is not a wildcard
                    return str(aspect_value)
                elif aspect_info[3] == 'in':
                    # the aspect_value is a list or set of names.
                    concepts = []
                    for aspect_name in aspect_value.value:
                        if aspect_name == 'qname':
                            concepts.append(str(aspect_name.value))
                        elif aspect_name == 'concept':
                            concepts.append(str(aspect_name.value.qname))
                    if len(concepts) == 1:
                        return str(concepts[0])
                    else:
                        return 'one of (' + ', '.join(concepts) + ')'

    # If we get here, then the default tag is unknown
    return 'unknown'


def evaluate_block(block_expr, xule_context):
    """Evaluator for block expressions
    
    :param block_expr: Rule expression
    :type block_expr: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue
    
    A block expression is a series of variable declarations followed by an expression.
    """
    for var_assignment in block_expr['varDeclarations']:
        # for var_assignment in var_assignments:
        var_info = xule_context.add_var(var_assignment['varName'],
                                        var_assignment['node_id'],
                                        var_assignment['varName'],  # tagged - all variables are tagged
                                        var_assignment['body'])
        calc_var(var_info, None, xule_context)

    return evaluate(block_expr['expr'], xule_context)


def evaluate_var_ref(var_ref, xule_context):
    """Evaluator for block expressions
    
    :param var_ref: Rule expression
    :type var_ref: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue
    """
    # print(var_ref['node_id'], var_ref.varName)
    var_info = xule_context.find_var(var_ref['varName'], var_ref['var_declaration'])
    # xule_context.used_vars.append(var_ref.var_declaration)

    try:
        var_value = calc_var(var_info, var_ref, xule_context)
    except XuleIterationStop as xis:
        var_value = xis.stop_value
        raise

    return var_value


def calc_var(var_info, const_ref, xule_context):
    """Calculate the value of a variable
    
    :param var_info: A dictionary of meta data about the variable
    :type var_info: dict
    :param const_ref: The constant declaration if the variable reference is for a constant
    :type const_ref: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue
    
    This function evaluates the expression for the variable reference.
    """
    if var_info['type'] == xule_context._VAR_TYPE_ARG:
        var_value = var_info['value']
    elif var_info['type'] == xule_context._VAR_TYPE_VAR:
        if var_info['calculated'] == False:
            try:
                saved_aligned_result_only = xule_context.aligned_result_only
                saved_used_expressions = xule_context.used_expressions
                xule_context.aligned_result_only = False
                xule_context.used_expressions = set()
                try:
                    var_info['value'] = evaluate(var_info['expr'], xule_context)
                    var_info['value'].aligned_result_only = xule_context.aligned_result_only
                    var_info['value'].used_expressions = xule_context.used_expressions
                    var_info['calculated'] = True
                except XuleIterationStop as xis:
                    var_info['value'] = xis.stop_value
                    var_info['value'].aligned_result_only = xule_context.aligned_result_only
                    var_info['value'].used_expressions = xule_context.used_expressions
                    var_info['calculated'] = True
                    # raise
            finally:
                xule_context.aligned_result_only = xule_context.aligned_result_only or saved_aligned_result_only
                xule_context.used_expressions = saved_used_expressions | xule_context.used_expressions

        var_value = var_info['value']
    elif var_info['type'] == xule_context._VAR_TYPE_CONSTANT:
        # We should only end up here the first time the constant is referenced for the iteration.
        # The var_info is really the constant info from the global context
        var_value = evaluate(var_info['expr'], xule_context, override_table_id=const_ref['table_id'])
    else:
        raise XuleProcessingError(_("Internal error: unkown variable type '%s'" % var_info['type']), xule_context)

    var_value = var_value.clone()
    # if var_info['tagged']:
    if var_info['tagged'] is not None:
        xule_context.tags[var_info['tagged']] = var_value

    return var_value


def calc_constant(const_info, const_context):
    """Calculate the value of a constant
    
    :param const_info: Meta data about the constant
    :type const_info: dict
    :param const_context: Rule processing context
    :type const_context: XuleRuleContext
    :returns: The evaluated value or values
    :rtype: XuleValue or XuleValueSet
    
    Constants are evaluated in a separate table. This isolates the evaluation from the rule which is using the constant. If the
    constant produces a singleton value a single value is returned. If the constant produces multiple values, a value set is returned.
    """
    const_context.iteration_table.add_table(const_info['expr']['node_id'],
                                            const_context.get_processing_id(const_info['expr']['node_id']))

    const_values = XuleValueSet()

    while True:
        const_context.aligned_result_only = False
        const_context.used_expressions = set()
        try:
            const_value = evaluate(const_info['expr']['body'], const_context)
        except XuleIterationStop as xis:
            const_value = xis.stop_value  # XuleValue(const_context, None, 'unbound')

        const_value.facts = const_context.facts.copy()
        const_value.tags = const_context.tags.copy()
        const_value.aligned_result_only = const_context.aligned_result_only
        # const_value.used_expressions = const_context.used_expressions
        try:
            const_value.alignment = const_context.iteration_table.current_table.current_alignment
        except AttributeError:
            # This happens if there isn't a current table because it was never created.
            pass
        const_values.append(const_value)

        # const_context.iteration_table.del_current()
        if not const_context.iteration_table.is_empty:
            const_context.iteration_table.next(const_context.iteration_table.current_table.table_id)
        # if const_context.iteration_table.is_empty:
        if const_context.iteration_table.is_table_empty(const_info['expr']['node_id']):
            break
    #         else:
    #             const_context.reset_iteration()

    # reset the aligned only results.
    const_info['expr']['aligned_only_results'] = const_context.aligned_result_only

    const_info['value'] = const_values
    const_info['calculated'] = True

def override_constant_calc(const_info, xule_context):

    const_values = XuleValueSet(xule_context.constant_overrides[const_info['name']])
    const_info['value'] = const_values
    const_info['calculated'] = True

def evaluate_constant_assign(const_assign, xule_context):
    """Evaluator a constant declaration
    
    :param const_assign: Rule expression for the constant declaration
    :type const_assign: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue
    """
    const_info = xule_context.find_var(const_assign['constantName'], const_assign['node_id'], constant_only=True)
    if const_info is None:
        raise XuleProcessingError(_("Constant '%s' not found" % const_assign['constantName']), xule_context)

    if not const_info['calculated']:
        # Check if there is a xule-arg that overrides the constant
        if const_assign['constantName'] in xule_context.constant_overrides:
            override_constant_calc(const_info, xule_context)
        else: # calc the constant
            const_context = XuleRuleContext(xule_context.global_context, None,
                                            xule_context.cat_file_num)
            calc_constant(const_info, const_context)
            # Clean up
            del const_context
    if 'is_iterable' in const_assign:
        # return the entire value set
        return const_info['value'].clone()
    else:
        # retrieve the single value
        return const_info['value'].values[None][0].clone()


def process_precalc_constants(global_context):
    """Precalculate constants
    
    :param global_context: Global processing context
    :type global_context: XuleGlobalContext
    
    This function will calculate constants that do not depend directly on the instance.
    """
    global_context.message_queue.logging("Precalcing non-instance constants")
    for constant_name, cat_constant in global_context.rule_set.catalog['constants'].items():
        if ('unused' not in cat_constant and
            not cat_constant['dependencies']['instance']):

            const_context = XuleRuleContext(global_context, constant_name, cat_constant['file'])
            const_info = const_context.find_var(constant_name, cat_constant['node_id'])
            if not const_info['calculated']:
                calc_constant(const_info, const_context)
            # Clean up
            del const_context

def process_reloadable_constants(global_context):
    """Load saved reloadable constants

    :param global_context: Global processing context
    :type global_context: XuleGlobalContext

    This function will calculate constants that do not depend directly on the instance.
    """
    if getattr(global_context.model, "log", None) is not None:
        global_context.model.log("DEBUG",
                                 "xule.loadSavedConstants", 
                                 "Reloading saved non-instance constants")
    else:
        global_context.message_queue.logging("Reloading saved non-instance constants")
    from arelle import FileSource
    xule_args_file = getattr(global_context.options,'xule_args_file', None)
    file_source = FileSource.openFileSource(xule_args_file, global_context.cntlr)
    with file_source.file(xule_args_file, binary=True)[0] as fh:
        const_objs = json.load(fh)
        for constant_name, obj in const_objs.items():
            try:
                cat_constant = global_context.rule_set.catalog['constants'][constant_name]
                const_context = XuleRuleContext(global_context, constant_name, cat_constant['file'])
                const_info = const_context.find_var(constant_name, cat_constant['node_id'])
                if not const_info['calculated']:
                    const_info['value'] = XuleValueSet(const_context.reload_value(obj))
                    const_info['calculated'] = True
                # Clean up
                del const_context
            except KeyError:
                pass # There is no constant stub to reload

def precalc_constant(global_context, constant_name):
        try:
            cat_constant = global_context.rule_set.catalog['constants'][constant_name]
        except KeyError:
            return # There is no constant to calc
        
        const_context = XuleRuleContext(global_context, constant_name, cat_constant['file'])
        const_info = const_context.find_var(constant_name, cat_constant['node_id'])
        if not const_info['calculated']:
            calc_constant(const_info, const_context)
        # Clean up
        del const_context
        return const_info

def evaluate_if(if_expr, xule_context):
    """Evaluator for if expressions
    
    :param if_expr: Rule expression for the constant declaration
    :type if_expr: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue    
    """
    if_thens = []
    if_thens.append((if_expr['condition'], if_expr['thenExpr']))

    for else_if in if_expr.get('elseIfExprs', []):
        if_thens.append((else_if['condition'], else_if['thenExpr']))

    for if_then in if_thens:
        condition_value = evaluate(if_then[0], xule_context)
        if condition_value.type in ('unbound', 'none'):
            return XuleValue(xule_context, None, 'unbound')
        elif condition_value.type != 'bool':
            raise XuleProcessingError(_("If condition is not a boolean, found '%s'" % condition_value.type),
                                      xule_context)
        else:
            if condition_value.value:
                return evaluate(if_then[1], xule_context)

    # This is only hit if none of the if conditions passed
    return evaluate(if_expr['elseExpr'], xule_context)


def evaluate_for(for_expr, xule_context):
    """Evaluator for for expressions
    
    :param for_expr: Rule expression for the constant declaration
    :type for_expr: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValueSet    
    """
    for_values = XuleValueSet()

    saved_used_expressions = xule_context.used_expressions
    xule_context.used_expressions = set()
    try:
        for_loop_collection = evaluate(for_expr['forLoopExpr'], xule_context)
    finally:
        used_expressions = xule_context.used_expressions
        xule_context.used_expressions = saved_used_expressions | used_expressions

    if for_loop_collection.type not in ('list', 'set'):
        raise XuleProcessingError(_("For loop requires a set or list, found '{}'.".format(for_loop_collection.type)),
                                  xule_context)

    for for_loop_var in for_loop_collection.value:
        if for_loop_var.used_expressions is None:
            for_loop_var.used_expressions = used_expressions
        else:
            for_loop_var.used_expressions.update(used_expressions)
        xule_context.add_arg(for_expr['forVar'],
                             for_expr['forLoopExpr']['node_id'],
                             for_expr['forVar'],  # tagged - all variables are automatically tagged
                             for_loop_var,
                             'single',
                             True)

        try:
            body_values = evaluate_for_body_detail(for_expr['forBodyExpr'],
                                                   for_expr['node_id'],
                                                   for_loop_var,
                                                   for_expr['forVar'],  # tag name
                                                   xule_context)
        finally:
            xule_context.del_arg(for_expr['forVar'], for_expr['forLoopExpr']['node_id'])
            del xule_context.tags[for_expr['forVar']]
            
        if for_loop_var.alignment is None:
            # add all
            for body_value in body_values.values.values():
                for_values.append(body_value)
        else:
            if for_loop_var.alignment in body_values.values:
                # take the aligned values
                for body_value in body_values.values[for_loop_var.alignment]:
                    for_values.append(body_value)
            else:
                # take only none aligned values and add alignment
                for body_value in body_values.values[None]:
                    body_value.alignment = for_loop_var.alignment
                    for_values.append(body_value)

    return for_values


def evaluate_for_body_detail(body_expr, table_id, for_loop_var, for_loop_tag, xule_context):
    """Evaluates the for body
    
    :param body_expr: Rule expression for the for body
    :type body_expr: dict
    :param table_id: The table id for the sub table to evaluate the for body
    :type table_id: int
    :param for_loop_var: The xuel value of the for loop variable
    :type for_loop_var: XuleValue
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValueSet        
    """
    body_values = XuleValueSet()

    aligned_result_only = False
    save_aligned_result_only = xule_context.aligned_result_only
    save_used_expressions = xule_context.used_expressions
    # for_body_table = xule_context.iteration_table.add_table(xule_context.get_processing_id(body_expr['node_id']), is_aggregation=True)

    for_body_table = xule_context.iteration_table.add_table(table_id, xule_context.get_processing_id(table_id),
                                                            is_aggregation=True)
    for_body_table.dependent_alignment = for_loop_var.alignment

    # add the loop control to the table
    # xule_context.iteration_table.add_column(rule_part, for_body_table.table_id, processing_id, values, xule_context)

    try:
        while True:
            xule_context.aligned_result_only = False
            xule_context.used_expressions = set()
            xule_context.tags[for_loop_tag] = for_loop_var
            body_value = XuleValue(xule_context, None, 'unbound')
            try:
                body_value = evaluate(body_expr, xule_context)
            except XuleIterationStop:
                pass

            aligned_result_only = aligned_result_only or xule_context.aligned_result_only
            body_value.alignment = for_body_table.current_alignment  # xule_context.iteration_table.dependent_alignment or xule_context.iteration_table.current_table.current_alignment
            body_value.aligned_result_only = aligned_result_only
            body_value.facts = xule_context.iteration_table.facts.copy()
            body_value.tags = xule_context.iteration_table.tags.copy()
            # print("for", body_expr['exprName'], body_expr['node_id'], len(xule_context.used_expressions), len(body_value.used_expressions))
            body_value.used_expressions = xule_context.used_expressions
            body_values.append(body_value)

            xule_context.iteration_table.next(for_body_table.table_id)
            if for_body_table.is_empty:
                break
    finally:
        xule_context.aligned_result_only = save_aligned_result_only
        xule_context.used_expressions = save_used_expressions
        xule_context.iteration_table.del_table(for_body_table.table_id)
    return body_values


def evaluate_unary(unary_expr, xule_context):
    """Evaluator for unary expressions
    
    :param unary_expr: Rule expression for the constant declaration
    :type unary_expr: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue   
    
    A unary expression is a plus or minus that flips the sign of a number. 
    """
    initial_value = evaluate(unary_expr['expr'], xule_context)

    if initial_value.type in ('unbound', 'none'):
        return initial_value.clone()

    if initial_value.type not in ('int', 'float', 'decimal'):
        raise XuleProcessingError(_("Unary operator requires a numeric operand, found '%s'" % initial_value.type),
                                  xule_context)

    if unary_expr['op'] == '-':
        return XuleValue(xule_context, initial_value.value * -1, initial_value.type)
    else:
        return initial_value.clone()

def evaluate_in(in_expr, xule_context):
    """Evaluator for in expressions

    :param in_expr: Rule expression for the in expression
    :type in_expr: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue
    """
    left = evaluate(in_expr['leftExpr'], xule_context)
    for right_side in in_expr['rights']:
        right = evaluate(right_side['rightExpr'], xule_context)
        if right.type in ('unbound', 'none'):
            left = XuleValue(xule_context, None, 'unbound')
        else:
            left = XuleProperties.property_contains(xule_context, right, left)

    return left

def evaluate_mult(mult_expr, xule_context):
    """Evaluator for multiplication expressions
    
    :param mult_expr: Rule expression for the constant declaration
    :type mult_expr: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue   
    
    This include multiplication and division
    """
    left = evaluate(mult_expr['leftExpr'], xule_context)

    for right_side in mult_expr['rights']:
        operator = right_side['op']
        right = evaluate(right_side['rightExpr'], xule_context)

        if left.type in ('unbound', 'none') or right.type in ('unbound', 'none'):
            left = XuleValue(xule_context, None, 'unbound')
        else:
            #             if left.type == 'unit' and right.type == 'unit':
            #                 #units have special handling
            #                 if operator == '*':
            #                     left = XuleValue(xule_context, unit_multiply(left.value, right.value), 'unit')
            #                 else:
            #                     left = XuleValue(xule_context, unit_divide(left.value, right.value), 'unit')
            #             else:
            # at this point there should only be numerics.
            if left.type not in ('int', 'float', 'decimal'):
                raise XuleProcessingError(
                    _("The left operand of '%s' is not numeric, found '%s'" % (operator, left.type)), xule_context)
            if right.type not in ('int', 'float', 'decimal'):
                raise XuleProcessingError(
                    _("The right operand of '%s' is not numeric, found '%s'" % (operator, right.type)), xule_context)

            combined_type, left_compute_value, right_compute_value = combine_xule_types(left, right, xule_context)
            '''NEED TO HANDLE CHNAGES IN UNIT ALIGNMENT'''
            if operator == '*':
                left = XuleValue(xule_context, left_compute_value * right_compute_value, combined_type)
            else:
                # This is division
                if right_compute_value == 0:
                    raise XuleProcessingError(_("Divide by zero error."), xule_context)
                left = XuleValue(xule_context, left_compute_value / right_compute_value, 'float' if combined_type == 'int' else combined_type)
    return left


def evaluate_intersect(inter_expr, xule_context):
    """Evaluator for intersection expressions
    
    :param inter_expr: Rule expression for the constant declaration
    :type inter_expr: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue   
    
    This is intersection of 2 sets.
    """
    left = evaluate(inter_expr['leftExpr'], xule_context)
    for right_side in inter_expr['rights']:
        right = evaluate(right_side['rightExpr'], xule_context)
        if left.type in ('unbound', 'none') or right.type in ('unbound', 'none'):
            left = XuleValue(xule_context, None, 'unbound')
        if left.type != 'set':
            raise XuleProcessingError(
                _("Intersection can only operatate on sets. The left side is a '{}'.".format(left.type)), xule_context)
        if right.type != 'set':
            raise XuleProcessingError(
                _("Intersection can only operatate on sets. The right side is a '{}'.".format(right.type)),
                xule_context)

        left = XuleUtility.intersect_sets(xule_context, left, right)

    return left


def evaluate_symetric_difference(sym_diff_expr, xule_context):
    """Evaluator for symetric difference expressions
    
    :param inter_expr: Rule expression for the constant declaration
    :type inter_expr: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue   
    
    This is the symetric difference of 2 sets.
    """
    left = evaluate(sym_diff_expr['leftExpr'], xule_context)
    for right_side in sym_diff_expr['rights']:
        right = evaluate(right_side['rightExpr'], xule_context)
        if left.type in ('unbound', 'none') or right.type in ('unbound', 'none'):
            left = XuleValue(xule_context, None, 'unbound')
        if left.type != 'set':
            raise XuleProcessingError(
                _("Symetric difference can only operatate on sets. The left side is a '{}'.".format(left.type)),
                xule_context)
        if right.type != 'set':
            raise XuleProcessingError(
                _("Symetric difference can only operatate on sets. The right side is a '{}'.".format(right.type)),
                xule_context)

        left = XuleUtility.symetric_difference(xule_context, left, right)

    return left


def evaluate_add(add_expr, xule_context):
    """Evaluator for add expressions
    
    :param add_expr: Rule expression for the constant declaration
    :type add_expr: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue   
    
    This includes add and subtract. These operations can be perforom on numbers, strings, sets and list. For strings
    adding wil concatenation. For lists and strings, adding will union. 
    
    In other binary operations, if an operand does not exist the operation is not performed. With add and subtract, if an
    operand is missing, it will treated as if it were zero. For example:
    
        @Assets + @Liabilities
        
    If there isn't a matching liability for an asset, the operation will return the value of assets.
    """
    left_bar = add_expr['rights'][0]['op'][0] == '<'  # the first operator
    if left_bar:
        left = evaluate(add_expr['leftExpr'], xule_context)
    else:
        # unbound is allowed, so it needs to be captured.
        try:
            left = evaluate(add_expr['leftExpr'], xule_context)
        except XuleIterationStop as xis:
            left = xis.stop_value  # XuleValue(xule_context, None, 'unbound')

    for right in add_expr['rights']:

        operator = right['op']
        right_bar = operator[-1] == '>'
        left_bar = operator[0] == '<'
        right_expr = right['rightExpr']

        if left.type not in (
        'int', 'float', 'decimal', 'string', 'uri', 'instant', 'time-period', 'set', 'list', 'dictionary', 'unbound', 'none'):
            raise XuleProcessingError(_("Left side of a {} operation cannot be {}.".format(operator, left.type)),
                                      xule_context)

        if right_bar:
            right = evaluate(right_expr, xule_context)
        else:
            # unbound is allowed, so it needs to be captured.
            try:
                right = evaluate(right_expr, xule_context)
            except XuleIterationStop as xis:
                right = xis.stop_value

        if right.type not in (
        'int', 'float', 'decimal', 'string', 'uri', 'instant', 'time-period', 'set', 'list', 'dictionary', 'unbound', 'none'):
            raise XuleProcessingError(_("Right side of a {} operation cannot be {}.".format(operator, right.type)),
                                      xule_context)

        # A time-period can be on the left only if the right is also a time period.
        if left.type == 'time-period' and right.type != 'time-period':
            raise XuleProcessingError(_("Incompatabile operands {} {} {}.".format(left.type, operator, right.type)),
                                      xule_context)
        do_calc = True

        if left_bar and left.type in ('unbound', 'none'):
            raise XuleIterationStop(XuleValue(xule_context, None, 'unbound'))
        if right_bar and right.type in ('unbound', 'none'):
            raise XuleIterationStop(XuleValue(xule_context, None, 'unbound'))

        if left.type in ('unbound', 'none'):
            # This is a special case for numbers. The left is none/unbound and the right is number. The new value will
            # be the negative of the right.
            if right.type in ('int', 'float', 'decimal') and '-' in operator:
                right = XuleValue(xule_context, right.value * -1, right.type)
            left = right.clone()
            do_calc = False
        if right.type in ('unbound', 'none'):
            do_calc = False

        # this ensures that if there is no value in the entire expression, the final value will be skipped.
        if left.type == 'none':
            left.type = 'unbound'

        if do_calc:
            combined_type, left_compute_value, right_compute_value = combine_xule_types(left, right, xule_context)
            if combined_type == 'unbound':
                raise XuleProcessingError(_("Incompatabile operands {} {} {}.".format(left.type, operator, right.type)),
                                          xule_context)

            if '+' in operator:
                if left.type == 'set' and right.type == 'set':
                    # use union for sets
                    # left = XuleValue(xule_context, left_compute_value | right_compute_value, 'set')
                    left = XuleUtility.add_sets(xule_context, left, right)
                elif left.type == 'dictionary' and right.type == 'dictionary':
                    left = XuleUtility.add_dictionaries(xule_context, left, right)
                elif left.type == 'dictionary' and right.type in ('set', 'list'):
                    raise XuleProcessingError(_("Cannot add a dictionary and a %s" % right.type), xule_context)
                elif left.type == 'instant' and right.type == 'instant':
                    raise XuleProcessingError(_("Dates cannot be added"), xule_context)
                else:
                    left = XuleValue(xule_context, left_compute_value + right_compute_value, combined_type)
            elif '-' in operator:
                if left.type == 'set' and right.type == 'set':
                    left = XuleUtility.subtract_sets(xule_context, left, right)
                elif left.type == 'dictionary' and right.type == 'dictionary':
                    left = XuleUtility.subtract_dictionaries(xule_context, left, right)
                elif left.type == 'dictionary' and right.type in ('set', 'list'):
                    left = XuleUtility.subtract_keys_from_dictionary(xule_context, left, right)
                elif left.type == 'list' or right.type == 'list':
                    raise XuleProcessingError(_("Lists cannot be subtracted"), xule_context)
                else:
                    if left.type == 'instant' and right.type == 'instant':
                        combined_type = 'time-period'
                    left = XuleValue(xule_context, left_compute_value - right_compute_value, combined_type)
            else:
                raise XuleProcessingError(
                    _("Unknown operator '%s' found in addition/subtraction operation." % operator), xule_context)

    return left

def evaluate_comp(comp_expr, xule_context):
    """Evaluator for comparison expressions
    
    :param comp_expr: Rule expression for the constant declaration
    :type comp_expr: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue   
    
    Comparison includes ==, !=, >, < >=, <=, in, not in
    """
    left = evaluate(comp_expr['leftExpr'], xule_context)

    for right in comp_expr['rights']:
        operator = right['op']
        right_expr = right['rightExpr']

        right = evaluate(right_expr, xule_context)

        interim_value = None

        combined_type, left_compute_value, right_compute_value = combine_xule_types(left, right, xule_context)

        if left.type in ('instant', 'duration') and right.type in ('instant', 'duration'):
            left_compute_value = XulePeriodComp(left_compute_value)
            right_compute_value = XulePeriodComp(right_compute_value)

        if left.type in ('list', 'set', 'dictionary'):
            left_compute_value = left.shadow_collection
        if right.type in ('list', 'set', 'dictionary'):
            right_compute_value = right.shadow_collection

        if operator == '==':
            interim_value = XuleValue(xule_context, left_compute_value == right_compute_value, 'bool')
        elif operator == '!=':
            interim_value = XuleValue(xule_context, left_compute_value != right_compute_value, 'bool')
        elif operator == 'in':
            interim_value = XuleProperties.property_contains(xule_context, right, left)
            #interim_value = XuleValue(xule_context, left_compute_value in right_compute_value, 'bool')
        elif operator == 'not in':
            positive_value = XuleProperties.property_contains(xule_context, right, left)
            interim_value = positive_value
            interim_value.value = not interim_value.value
            #interim_value = XuleValue(xule_context, left_compute_value not in right_compute_value, 'bool')
        elif operator in ('<', '>'):
            if left.type == 'none' or right.type == 'none':
                interim_value = XuleValue(xule_context, None, 'none')
            elif operator == '<':
                interim_value = XuleValue(xule_context, left_compute_value < right_compute_value, 'bool')
            elif operator == '>':
                interim_value = XuleValue(xule_context, left_compute_value > right_compute_value, 'bool')
        elif operator in ('<=', '>='):
            if left.type == 'none' and right.type == 'none':
                interim_value = XuleValue(xule_context, True, 'bool')
            elif left.type == 'none' or right.type == 'none':
                interim_value = XuleValue(xule_context, None, 'none')
            elif operator == '<=':
                interim_value = XuleValue(xule_context, left_compute_value <= right_compute_value, 'bool')
            elif operator == '>=':
                interim_value = XuleValue(xule_context, left_compute_value >= right_compute_value, 'bool')
        else:
            raise XuleProcessingError(_("Unknown operator '%s'." % operator), xule_context)

        left = interim_value

    return left


def evaluate_not(not_expr, xule_context):
    """Evaluator for not expressions
    
    :param not_expr: Rule expression for the constant declaration
    :type not_expr: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue   
    """
    initial_value = evaluate(not_expr['expr'], xule_context)

    if initial_value.type in ('unbound', 'none'):
        return initial_value

    if initial_value.type != 'bool':
        raise XuleProcessingError(
            _("The operand of the 'not' expression must be boolean, found '%s'" % initial_value.type), xule_context)

    return XuleValue(xule_context, not initial_value.value, 'bool')


def evaluate_and(and_expr, xule_context):
    """Evaluator for boolean and expressions
    
    :param and_expr: Rule expression for the constant declaration
    :type and_expr: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue   
    """
    value_found = False
    has_unbound = False
    left = XuleValue(xule_context, None, 'unbound')
    # Create a single list of expressions combing the left and each of the rights.
    exprs = [and_expr['leftExpr'], ] + [x['rightExpr'] for x in and_expr['rights']]

    # This process will allow unbounds if at some point a false is found. In this case the and expression is false. Otherwise, unbounds will make the result unbound.
    for expr in exprs:
        if value_found:
            break
        else:
            try:
                right = evaluate(expr, xule_context)
            except XuleIterationStop as xis:
                right = xis.stop_value  # XuleValue(xule_context, None, 'unbound')
            if right.type in ('unbound', 'none'):
                has_unbound = True
            if left.type not in ('unbound', 'none', 'bool') or right.type not in ('unbound', 'none', 'bool'):
                raise XuleProcessingError(_(
                    "Operand of 'and' expression is not boolean. Left and right operand types are '%s' and '%s'." % (
                    left.type, right.type)), xule_context)

            if left.type == 'bool' and right.type == 'bool':
                left = XuleValue(xule_context, left.value and right.value, 'bool')
                if left.value == False:
                    value_found = True
            elif left.type in ('unbound', 'none') and right.type in ('unbound', 'none'):
                continue
            elif left.type in ('unbound', 'none') and right.type == 'bool':
                left = right.clone()
                if left.value == False:
                    value_found = True
            elif left.type == 'bool' and right.type in ('unbound', 'none'):
                if left.value == False:
                    value_found = True

    if (has_unbound and value_found) or not has_unbound:
        return left
    else:
        return XuleValue(xule_context, None, 'unbound')


def evaluate_or(or_expr, xule_context):
    """Evaluator for boolean or expressions
    
    :param or_expr: Rule expression for the constant declaration
    :type or_expr: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValue   
    """
    value_found = False
    has_unbound = False
    left = XuleValue(xule_context, None, 'unbound')

    # Create a single list of expressions combing the left and each of the rights.
    exprs = [or_expr['leftExpr'], ] + [x['rightExpr'] for x in or_expr['rights']]
    for expr in exprs:
        if value_found:
            break

        else:
            try:
                right = evaluate(expr, xule_context)
            except XuleIterationStop as xis:
                right = xis.stop_value  # XuleValue(xule_context, None, 'unbound')
            if right.type in ('unbound', 'none'):
                has_unbound = True
            if left.type not in ('unbound', 'none', 'bool') or right.type not in ('unbound', 'none', 'bool'):
                raise XuleProcessingError(_(
                    "Operand of 'or' expression is not boolean. Left and right operand types are '%s' and '%s'." % (
                    left.type, right.type)), xule_context)

            if left.type == 'bool' and right.type == 'bool':
                left = XuleValue(xule_context, left.value or right.value, 'bool')
                if left.value == True:
                    value_found = True
            elif left.type in ('unbound', 'none') and right.type in ('unbound', 'none'):
                continue
            elif left.type in ('unbound', 'none') and right.type == 'bool':
                left = right.clone()
                if left.value == True:
                    value_found = True
            elif left.type == 'bool' and right.type in ('unbound', 'none'):
                if left.value == True:
                    value_found = True
    if (has_unbound and value_found) or not has_unbound:
        return left
    else:
        return XuleValue(xule_context, None, 'unbound')


def evaluate_factset(factset, xule_context):
    """Evaluator for a factset

    :param factset: Rule expression for the constant declaration
    :type factset: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValueSet  
    
    There are two flavors of factsets.
    1. Starndard factset - finds facts in the instance (i.e. @Assets)
    2. Factset with inner expression - This is really an envelope that sets up aspect filters for other factsets that are inside the envelope.
    """
    if 'innerExpr' in factset:
        return evaluate_nesting_factset(factset, xule_context)
    else:
        return evaluate_factset_detail(factset, xule_context)


def evaluate_nesting_factset(factset, xule_context):
    """Evaluate a factset envolope
    
    :param factset: Rule expression for the constant declaration
    :type factset: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValueSet   
    """
    aspect_filters, _x, aspect_vars, _model, _y, _z = process_factset_aspects(factset, xule_context)
    # TODO - Probably, the model from the outer factset should be passed to the inner factsets

    # verify that there are not already filters in place
    current_filters, _x = xule_context.get_current_filters()
    current_aspects = {aspect_info[ASPECT] for aspect_info in current_filters}

    factset_aspects = {aspect_info[ASPECT] for aspect_info in aspect_filters}

    if current_aspects & factset_aspects:
        raise XuleProcessingError(_(
            "A nested factset clause cannot include aspects in an outer factset clause, found '%s'." % ", ".join(
                current_aspects & factset_aspects)), xule_context)

    # add the align aspects to the nested_filter in the context
    xule_context.filter_add('nested', aspect_filters)

    save_aligned_result_only = xule_context.aligned_result_only
    save_used_expressions = xule_context.used_expressions

    nested_table = xule_context.iteration_table.add_table(factset['node_id'],
                                                          xule_context.get_processing_id(factset['node_id']),
                                                          is_aggregation=True)
    nested_values = XuleValueSet()
    try:
        while True:
            xule_context.aligned_result_only = False
            xule_context.used_expressions = set()
            try:
                nested_value = evaluate(factset['innerExpr'], xule_context)
            except XuleIterationStop:
                nested_value = XuleValue(xule_context, None, 'unbound', tag=XuleValue(xule_context, None, 'none'))

            # if not(xule_context.iteration_table.current_table.current_alignment is None and xule_context.aligned_result_only):
            # remove the with portion of the alignment
            #if xule_context.iteration_table.current_table.current_alignment is not None:  # this should be the alignment on the with table
            remove_aspects = [(with_filter[0], with_filter[1]) for with_filter in aspect_filters]
            if  factset.get('covered', False):
                new_alignment = None
                remove_alignments(nested_value)
            elif xule_context.iteration_table.current_alignment is None:
                new_alignment = None
            else:
                new_alignment = remove_from_alignment(xule_context.iteration_table.current_alignment,
                                                  remove_aspects, xule_context)
            nested_value.alignment = new_alignment

            nested_value.facts = xule_context.facts.copy()
            nested_value.tags = xule_context.tags.copy()

            nested_values.append(nested_value)

            # xule_context.iteration_table.del_current()
            xule_context.iteration_table.next(nested_table.table_id)
            if nested_table.is_empty:
                break
    finally:
        xule_context.aligned_result_only = save_aligned_result_only
        xule_context.used_expressions = save_used_expressions
        # delete the with table (in case it is left behind from an exception)
        xule_context.iteration_table.del_table(nested_table.table_id)
        xule_context.filter_del()

    return nested_values

def remove_alignments(val):
    val.alignment = None
    val.aligned_result_only = False

    if val.type in ('list', 'set'):
        for child_val in val.value:
            remove_alignments(child_val)
    if val.type == 'dictionary':
        for key, child_val in val.value:
            remove_alignments(key)
            remove_alignments(child_val)

def evaluate_factset_detail(factset, xule_context):
    """Evaluate a factset
    
    :param factset: Rule expression for the constant declaration
    :type factset: dict
    :param xule_context: Rule processing context
    :type xule_context: XuleRuleContext
    :rtype: XuleValueSet   

    The factset is divided into two parts. The first part contains aspects that will be used to filter the fact and will NOT
    be used for alignment. For example: "Assets[]" or "[lineItem=Assets]". These factsets will find all the 'Assets' facts in the 
    instance, but when these facts are compared to facts in other fact sets, the 'concept' aspect will not be used to check alignment.
    
    Actual aspects of the fact that are not specified in the first part of the factset will be used for alignment.
    
    Nested alignment:
        This would be put in the context for nested factset expressions. It would cause downstream factset evaluations
        to include the filter as part of getting facts. If the filter is 'closed', it would act like a closed factset and not allow
        facts that have dimenions in fact's alignment that are not in the filter. 'open' filters wouldn't care.
        
        This provides an alternative mechanism for handling alignment. Instead of getting all results for each side of an operation (i.e. property) 
        and then aligning them, it would allow the expression to iterate over one operand result set and evaluate for each result of the other operand. 
        By pushing the filter, first, only the aligned results will come back. 
    """
    #The no alignment flag indicates that the results of the factset should all have none alignment. It is set by the 'values' expression.

    saved_used_expressions = xule_context.used_expressions
    xule_context.used_expressions = set()
    try:
        non_align_aspects, align_aspects, aspect_vars, new_models, dimensions_special_value, dimensions_special_covered = process_factset_aspects(factset, xule_context)
    finally:
        used_expressions = xule_context.used_expressions
        xule_context.used_expressions = saved_used_expressions | xule_context.used_expressions

    # check if any non_align_aspects overlap with nested_filters
    nested_factset_filters, other_filters = xule_context.get_current_filters()
    # This restriction is removed to suport rules like r1923
    #     if with_aspects & factset_aspects:
    #         raise XuleProcessingError(_("The factset cannont contain any aspects in any bounding 'with' clause, found '%s'." % ", ".join(with_aspects & factset_aspects)), xule_context)

    # combine all the filtering aspects.
    all_aspect_filters = list(nested_factset_filters.items()) + list(other_filters.items()) + list(
        non_align_aspects.items())

    # If the the factset is dependent, then we only need to find facts that also match the current alignment. Create filters based on the current alignment.
    dependent_filters = list()
    # the first part of this if statement is dtermining if the factset has alignment or not. 
    if (not factset.get('covered') or ('aspectFilters' in factset and any({True for x in factset['aspectFilters'] if x['coverType'] == 'uncovered'}))) and factset['is_dependent']:
        if xule_context.dependent_alignment is None:
            # The table may acquire the dependent alignment after evaluating the aspect filters
            xule_context.iteration_table.current_table.make_dependent()
        try:
            if xule_context.dependent_alignment is not None:
                unfrozen_alignment = {k: v for k, v in xule_context.dependent_alignment}
                dependent_filters = list(alignment_to_aspect_info(unfrozen_alignment, xule_context).items())
        except IndexError:
            pass
    all_aspect_filters += list(align_aspects.items()) + dependent_filters

    '''Match facts based on the aspects in the first part of the factset and any additional filters.
    This is done by intersecting the sets of the fact_index. The fact index is a dictionary of dictionaries.
    The outer dictionary is keyed by aspect and the inner by member. So fact_index[aspect][member] contains a 
    set of facts that have that aspect and member.'''
    pre_matched_facts = factset_pre_match(factset, all_aspect_filters, non_align_aspects, align_aspects, new_models, xule_context,
                                          dimensions_special_value, dimensions_special_covered)

    # For dependent factset, set flag to check if the iteration table becomes aligned.
    # Bassically, there is no alignment yet. During the evaluation of the where clause a first time evaluated variable can create alignment.
    # If this happens, the pre matched facts should only include those facts that have matching alignment. So a flag in the context is set to check if the table becomes aligned.
    # When this happens a XuleReEvaluate exception is raised (this happens in the main evaluator()).

    saved_look_for_alignment = xule_context.look_for_alignment
    saved_where_table_ids = xule_context.where_table_ids
    saved_where_dependent_iterables = xule_context.where_dependent_iterables

    if factset['is_dependent'] and xule_context.iteration_table.any_alignment is None:
        xule_context.look_for_alignment = True
        # xule_context.where_table_ids = list(xule_context.iteration_table._ordered_tables.keys())
        xule_context.where_table_ids = [table.table_id for table in
                                        xule_context.iteration_table._ordered_tables.values()]
        xule_context.where_dependent_iterables = [di['node_id'] for di in factset['dependent_iterables']]
    else:
        # set_look_for_alignment = False
        xule_context.look_for_alignment = False
    # print(factset['node_id'], fact_value, no_pre_where_alignment, xule_context.iteration_table.current_table.table_id)

    default_where_used_expressions = set()
    try:
        results, default_where_used_expressions = process_filtered_facts(factset,
                                                                        pre_matched_facts,
                                                                        non_align_aspects,
                                                                        align_aspects,
                                                                        nested_factset_filters,
                                                                        aspect_vars, used_expressions, 
                                                                        dimensions_special_value, dimensions_special_covered,
                                                                        xule_context)
    except XuleReEvaluate as xac:
        # turn off looking for changes during the where evaluation. At this point either there is no alignment, so the result will be empty or
        # there is alignment and the pre_matched_facts will be refiltered with the alignment. In this case, we don't need to continue looking for
        # evaluating an iterable that can produce alignment.

        xule_context.iteration_table.current_table.make_dependent()

        xule_context.look_for_alignment = False
        if xac.alignment is None:
            # there are no matching facts in the dependent iterable (that has alignments)
            results = XuleValueSet()
        else:
            # This occurs if the alignment is created while processing where clause. Re-filter the pre_matched facts with the alignment information and try the where clause again
            # Add the alignment to the all_aspect filters
            #             if xac.alignment is None:
            #                 #in this case the no facts can match because the dependency is unaligned but would normally have alignment. This is basicaly the default value of 'unbound'
            #                 #for a factset.
            #                 results = XuleValueSet()
            #             else:
            unfrozen_alignment = {k: v for k, v in xac.alignment}
            additional_aspect_filters = list(alignment_to_aspect_info(unfrozen_alignment, xule_context).items())
            pre_matched_facts = factset_pre_match(factset, additional_aspect_filters, non_align_aspects, align_aspects,
                                                  new_models, xule_context, dimensions_special_value, dimensions_special_covered,
                                                  starting_facts=pre_matched_facts)
            # try again
            try:
                results, default_where_used_expressions = process_filtered_facts(factset, pre_matched_facts,
                                                                                non_align_aspects,
                                                                                align_aspects,
                                                                                nested_factset_filters, aspect_vars,
                                                                                used_expressions, 
                                                                                dimensions_special_value, dimensions_special_covered,xule_context)
            except XuleReEvaluate as xac:
                # In the second pass, the alignment change should not happen.
                raise XuleProcessingError(
                    _("Encountered 2nd alignment change while processing the 'where' clause for a factset"),
                    xule_context)
    finally:
        # if set_look_for_alignment:
        xule_context.look_for_alignment = saved_look_for_alignment
        xule_context.where_table_ids = saved_where_table_ids
        xule_context.where_dependent_iterables = saved_where_dependent_iterables

    if None not in results.values:
        expr_aspect_filters = non_align_aspects.copy()
        expr_aspect_filters.update(align_aspects)
        # default_value = XuleValue(xule_context, None, 'unbound', tag=XuleValue(xule_context, tag_default_for_factset(expr_aspect_filters, xule_context), 'empty_fact'))
        default_value = XuleValue(xule_context, None, 'unbound', tag=XuleValue(xule_context, None, 'none'))
        '''The current list of facts and tags are not inlcuded on the default None fact in a factset. This was causing problems with a exists() and missing().
        The default None fact in the missing would have the tags and facts from the first evaluation, but then these would be applied on consequent
        iterations where the tags from the first iteration would overwrite the tags on the consequent iterations.'''
        default_value.used_expressions = used_expressions | default_where_used_expressions
        # Added the second part of this 'if' statement to catch covered factsets that also have @@ which really makes them not covered.
        # This determines if the factset has alignment or not. Maybe - this could just look at the 'has_alignment' on the factset ???
        if not factset.get('covered') or ('aspectFilters' in factset and any({True for x in factset['aspectFilters'] if x['coverType'] == 'uncovered'})):
            default_value.aligned_result_only = True
        results.append(default_value)

    return results

def factset_pre_match(factset, filters, non_aligned_filters, align_aspects, models, xule_context, dimensions_special_value, dimensions_special_covered, starting_facts=None):
    """Match facts based on the factset  
       
    Match facts based on the aspects in the first part of the factset and any additional filters.
    
    This is done by intersecting the sets of the fact_index. The fact index is a dictionary of dictionaries.
    The outer dictionary is keyed by aspect and the inner by member. So fact_index[aspect][member] contains a 
    set of facts that have that aspect and member.
    """

    pre_matched_facts = set()

    for model in models:
        if starting_facts is None:
            pre_matched_model_facts = None
            first = True
        else:
            pre_matched_model_facts = copy.copy(starting_facts)
            first = False
        # models is a list of models based on the instance document
        fact_index = getattr(model, 'xuleFactIndex', dict()) # The model may be from a @instance which is different from the model in the xule_context

        for aspect_info, filter_member in filters:
            # Handle case where the filter only contains boolean values. Treat the filter_member as true.
            # For example: @concept.is-numeric This should be treated as @concept.is-numeric=true
            if (filter_member is None and
                aspect_info[ASPECT_OPERATOR] is None and
                aspect_info[SPECIAL_VALUE] is None and
                aspect_info[ASPECT_PROPERTY] is not None):

                aspect_info = list(aspect_info)
                aspect_info[ASPECT_OPERATOR] = '='
                filter_member = XuleValue(xule_context, True, 'bool')

            if filter_member is not None:
                index_key = fact_index_key(aspect_info, fact_index, xule_context)
                facts_by_aspect = set()

                '''THIS MIGHT BE MORE EFFICIENTLY HANDLED BY IGNORING THE ASPECT IF THE MEMBER IS None OR ELIMINATING ALL FACTS'''
                # When the aspect key is not in the fact index, then the instance doesn't use this aspect (dimension). So create an entry for the 'None' key and put all the facts in it.
                if index_key not in fact_index:
                    fact_index[index_key][None] = fact_index['all']

                if aspect_info[SPECIAL_VALUE] is not None:
                    if aspect_info[SPECIAL_VALUE] == '*':
                        if aspect_info[ASPECT_OPERATOR] == '=':
                            if aspect_info[TYPE] == 'builtin' and aspect_info[ASPECT] in ('concept', 'period', 'entity') and aspect_info[ASPECT_PROPERTY] is None:
                                # this is all facts
                                continue
                            else:
                                # need to combine all the facts that have that aspect
                                facts_by_aspect = set(it.chain.from_iterable(
                                    v for k, v in fact_index[index_key].items() if k is not None))
                        else:  # the operator is != ('in' and 'not in' are not allowed with a special value)
                            if aspect_info[TYPE] == 'builtin' and aspect_info[ASPECT] in ('concept', 'period', 'entity'):
                                # No facts can match these aspects not equal to * (i.e. @concept != *)
                                pre_matched_model_facts = []
                                break
                            else:
                                facts_by_aspect = fact_index[index_key][None]
                else:
                    if aspect_info[ASPECT_OPERATOR] == 'in' and filter_member.type not in ('list', 'set'):
                        raise XuleProcessingError(_("The value for '%s' with 'in' must be a set or list, found '%s'" % (
                        index_key[ASPECT], filter_member.type)), xule_context)

                    # fix for aspects that take qname members (concept and explicit dimensions. The member can be a concept or a qname. The index is by qname.
                    if index_key in (('builtin', 'concept'), ('property', 'cube', 'name')):
                        if aspect_info[ASPECT_OPERATOR] in ('=', '!='):
                            member_values = convert_value_to_qname(filter_member,model, xule_context)
                        else: # in operator
                            member_values = set()
                            for filter_val in filter_member.value:
                                for x in convert_value_to_qname(filter_val, model, xule_context):
                                    member_values.add(x)
                    # elif index_key[TYPE] == 'explicit_dimension':
                    #     if aspect_info[ASPECT_OPERATOR] in ('=', '!='):
                    #         member_values = convert_value_to_qname(filter_member, model, xule_context)
                    #     else:
                    #         member_values = {convert_value_to_qname(x, model, xule_context) if x.type == 'concept' else x.value for x
                    #                         in filter_member.value}
                    # Also fix for period aspect
                    elif index_key[TYPE] == 'explicit_dimension':
                        # Get the dimension concept to see if it is explicit or typed.
                        # With namespace groups, there may be more than one dimension concept
                        if isinstance(aspect_info[ASPECT], tuple):
                            # this is a namespace group qname for the dimension. This is not supported
                            raise XuleProcessingError(_(f"The prefix {aspect_info[ASPECT][0].value} is a namespace-group prefix. Namespace-group prefixes cannot be used for dimension aspects of a factset. Full namespace-group name is {aspect_info[ASPECT][0].value}:{aspect_info[ASPECT][1]}."))
                        model_dimension = model.qnameConcepts.get(aspect_info[ASPECT])
                        if model_dimension is None:
                                raise XuleProcessingError(_(f"Cannot find dimension concept for {aspect_info[ASPECT].clarkNotation} while processing a factset"), xule_context)
                        if aspect_info[ASPECT_OPERATOR] in ('=', '!='):
                            if model_dimension.isExplicitDimension:
                                member_values = convert_value_to_qname(filter_member, model, xule_context)
                            else:
                                # This is a typed dimension
                                member_values = {filter_member.value,}
                        else: # in or not in operator
                            if model_dimension.isExplicitDimension:
                                member_values = set()
                                for filter_val in filter_member.value:
                                    for x in convert_value_to_qname(filter_val, model, xule_context):
                                        member_values.add(x)
                            else: # Typed dimension
                                member_values = {x.value for x in filter_member.value}
                            
                    elif index_key == ('builtin', 'period'):
                        if aspect_info[ASPECT_OPERATOR] in ('=', '!='):
                            member_values = {convert_value_to_model_period(filter_member, xule_context), }
                        else:
                            member_values = {convert_value_to_model_period(x, xule_context) for x in filter_member.value}
                    # Allow units to be a qname or a xule 'unit'
                    elif index_key == ('builtin', 'unit'):
                        conversion_function = lambda x: XuleUnit(x) if x.type == 'qname' else x.value
                        if aspect_info[ASPECT_OPERATOR] in ('=', '!='):
                            member_values = {conversion_function(filter_member), }
                        else:
                            member_values = {conversion_function(x) for x in filter_member.value}
                    # Allow @table.drs-role to take a short role name
                    elif index_key == ('property', 'cube', 'drs-role'):
                        if aspect_info[ASPECT_OPERATOR] in ('=', '!='):
                            member_values = set(convert_value_to_role(filter_member, xule_context))
                        else:
                            member_values = set().union(*[set(convert_value_to_role(x, xule_context)) for x in filter_member.value])
                    else:
                        if aspect_info[ASPECT_OPERATOR] in ('=', '!='):
                            member_values = {filter_member.value, }
                        else:
                            member_values = {x.value for x in filter_member.value}
                            '''THIS COULD USE THE SHADOW COLLECTION
                            member_values = set(filter_member.shadow_collection)
                            '''
                    if aspect_info[ASPECT_OPERATOR] in ('=', 'in'):
                        found_members = member_values & fact_index[index_key].keys()
                    else:  # aspect operator is '!=' or 'not in'
                        found_members = (fact_index[index_key].keys() - {None, }) - member_values

                    for member in found_members:
                        facts_by_aspect |= fact_index[index_key][member]

                # intersect the facts with previous facts by aspect
                if first:
                    first = False
                    pre_matched_model_facts = facts_by_aspect
                else:
                    pre_matched_model_facts &= facts_by_aspect

        if first:
            # there were no apsects to start the matching, so use the full set
            # pre_matched_model_facts = xule_context.model.factsInInstance
            try:
                pre_matched_model_facts = fact_index['all']
            except KeyError: # This only happens because there is no model.
                pre_matched_model_facts = set()
        
        pre_matched_facts |= pre_matched_model_facts

    if starting_facts is None:
        # Check the alignment of pre matched facts to the dependent alignment
        if xule_context.dependent_alignment is not None and factset.get('is_dependent', False):
            match_aligned_facts = set()
            for fact in pre_matched_facts:
                fact_alignment = calc_fact_alignment(factset, fact, non_aligned_filters, align_aspects, True, 
                                                     dimensions_special_covered, xule_context)

                if fact_alignment == xule_context.dependent_alignment:
                    match_aligned_facts.add(fact)
            pre_matched_facts = match_aligned_facts

        '''  
        #This code reduces the pre matched facts to those that match alignment of the dependent alignment by using the fact index of dimensions that are not
        #in the dependent alignment. The method of checking the alignment used above proved to be more efficient. However, it may be that if the pre match includes a large number
        #of facts this method may be better.  
        if xule_context.dependent_alignment is not None and factset.is_dependent:
            if not hasattr(factset, 'empty_dimension_list'):            
                aligned_dimensions = [(k[0], k[1]) for k, v in aligned_filters if k[0] == 'explicit_dimension']
                empty_dimensions = [k for k in xule_context.fact_index.keys() if k[0] == 'explicit_dimension' and k not in aligned_dimensions]
                factset.empty_dimension_list = empty_dimensions
                
            for empty_dim_key in factset.empty_dimension_list:
                pre_matched_facts &= xule_context.fact_index[empty_dim_key][None]
        '''
    return pre_matched_facts

def fact_index_key(aspect_info, fact_index, xule_context):
    if aspect_info[ASPECT_PROPERTY] is None:
        index_key = (aspect_info[TYPE], aspect_info[ASPECT])
    else:
        # aspect_info[ASPECT_PROPERTY][0] is the aspect property name
        # aspect_info[ASPECT_PROPERTY][1] is a tuple of the arguments
        index_key = ('property', aspect_info[ASPECT], aspect_info[ASPECT_PROPERTY][0]) + \
                    aspect_info[ASPECT_PROPERTY][1]
        if index_key not in fact_index and index_key not in xmi.FACT_INDEX_PROPERTIES and index_key not in xmi.TABLE_INDEX_PROPERTIES and aspect_info[ASPECT_PROPERTY][0] != 'attribute':
            raise XuleProcessingError(_(
                "Factset aspect property '{}' is not a valid property of aspect '{}'.".format(index_key[2],
                                                                                                index_key[1])),
                                        xule_context)    
    return index_key                                            

def calc_fact_alignment(factset, fact, non_aligned_filters, align_aspects_filters, frozen, dimensions_special_covered, xule_context):
    if fact not in xule_context.fact_alignments[factset['node_id']]:
        unfrozen_alignment = get_alignment(fact,
                                           non_aligned_filters,
                                           align_aspects_filters,
                                           xule_context,
                                           factset.get('coveredDims', False),
                                           factset.get('covered', False),
                                           dimensions_special_covered)

        if len(unfrozen_alignment) == 0 and factset.get('covered', False):
            unfrozen_alignment = None
            fact_alignment = None
        else:
            fact_alignment = frozenset(unfrozen_alignment.items())
        xule_context.fact_alignments[factset['node_id']][fact] = (fact_alignment, unfrozen_alignment)
        return fact_alignment if frozen else unfrozen_alignment

    return xule_context.fact_alignments[factset['node_id']][fact][0 if frozen else 1]


def process_filtered_facts(factset, pre_matched_facts, non_align_aspects, align_aspects,
                           nested_filters, aspect_vars, pre_matched_used_expressions_ids, 
                           dimensions_special_value, dimensios_special_covered, xule_context):
    """Apply the where portion of the factset"""
    results = XuleValueSet()
    default_used_expressions = set()

    for model_fact in pre_matched_facts:
        # assume the fact will matach the where clause.
        matched = True

        # check if nill
        exclude_nils = (factset.get('excludeNils', False) or
                        (getattr(xule_context.global_context.options, "xule_exclude_nils", False) and not factset.get(
                            'includeNils', False))
                        )
        if exclude_nils and model_fact.isNil:
            # if not xule_context.include_nils and model_fact.isNil:
            continue

        '''The alignment is all the aspects that were not specified in the first part of the factset (non_align_aspects).'''
        # set up potential fact result
        alignment = calc_fact_alignment(factset, model_fact, non_align_aspects, align_aspects, False, dimensios_special_covered, xule_context)
        '''If we are in a innner factset, the alignment needs to be adjusted. Each aspect in the outer factset should be in the alignment even if
           if it is in the factset aspects (which would normally take that aspect out of the alignment).'''
        for nested_aspect_info in nested_filters:
            alignment_info = (nested_aspect_info[TYPE], nested_aspect_info[ASPECT])
            if alignment is None or alignment_info not in alignment:
                if alignment_info == ('builtin', 'concept'):
                    alignment_value = model_fact.qname
                    # alignment_value = model_fact.elementQname
                elif alignment_info == ('builtin', 'unit'):
                    if model_fact.isNumeric:
                        alignment_value = model_to_xule_unit(model_fact.unit, xule_context)
                elif alignment_info == ('builtin', 'period'):
                    alignment_value = model_to_xule_period(model_fact.context, xule_context)
                elif alignment_info == ('builtin', 'entity'):
                    alignment_value = model_to_xule_entity(model_fact.context, xule_context)
                elif alignment_info[TYPE] == 'explicit_dimension':
                    model_dimension = model_fact.context.qnameDims.get(alignment_info[ASPECT])
                    if model_dimension is None:
                        alignment_value = None
                    else:
                        if model_dimension.isExplicit:
                            alignment_value = model_dimension.memberQname
                        else:
                            alignment_value = model_dimension.typedMember.xValue
                # NEED TO CHECK WHAT THE VALUE SHOULD BE
                else:
                    raise XuleProcessingError(_(
                        "Pushing nested factset filter alignment, found unknown alignment '%s : %s'" % alignment_info),
                                              xule_context)
                if alignment is None:
                    # There was no alignment, but now an aspect is being added to the alignment
                    alignment = {} #dict
                alignment[alignment_info] = alignment_value

        '''Check closed factset'''
        if factset['factsetType'] == 'closed' or dimensions_special_value == 'none':
            aspect_dimensions = {aspect_info[ASPECT] for aspect_info in (non_align_aspects | align_aspects)}
            if len(set(model_fact.context.qnameDims.keys()) - aspect_dimensions) != 0:
                continue

        # if the dimensions_special_value is '*', then there must be atlest 1 dimension for the fact
        if dimensions_special_value == '*' and len(model_fact.context.qnameDims) == 0:
            continue

        if alignment is not None:
            # if not current_no_alignment and xule_context.iteration_table.is_dependent:
            # if not current_no_alignment and factset['is_dependent']:
            if factset['is_dependent']:
                if xule_context.dependent_alignment is not None:
                    if frozenset(alignment.items()) != xule_context.dependent_alignment:
                        # If this is in a 'with' clause, the first factset to be added to the with/agg table may be empty, The current alignment will be
                        # from a higher table which will not include the with filter aspects.
                        if len(
                                nested_filters) > 0 and xule_context.iteration_table.current_table.current_alignment is None:
                            remove_aspects = [(nested_filter[0], nested_filter[1]) for nested_filter in nested_filters]
                            adjusted_alignment = remove_from_alignment(frozenset(alignment.items()), remove_aspects,
                                                                       xule_context)
                            if adjusted_alignment != xule_context.dependent_alignment:
                                continue  # try the next fact from the pre match
                        else:
                            continue  # try the next fact from the pre match

        if factset.get('nilDefault', False) and model_fact.isNil:
            # Handle nil value where the value should be defaulted.
            if (model_fact.concept.baseXbrliTypeQname == qname('http://www.xbrl.org/2003/instance',
                                                               'monetaryItemType') or
                model_fact.concept.baseXbrliTypeQname == qname('http://www.xbrl.org/2003/instance',
                                                               'decimalItemType') or
                model_fact.concept.baseXbrliTypeQname == qname('http://www.xbrl.org/2003/instance',
                                                               'sharesItemType') or
                model_fact.concept.baseXbrliTypeQname == qname('http://www.xbrl.org/2003/instance',
                                                               'pureItemType')):
                xule_type = 'decimal'
                system_value = decimal.Decimal(0)
            elif (model_fact.concept.baseXbrliTypeQname == qname('http://www.xbrl.org/2003/instance',
                                                                 'floatItemType') or
                  model_fact.concept.baseXbrliTypeQname == qname('http://www.xbrl.org/2003/instance',
                                                                 'doubleItemType')):
                xule_type = 'float'
                system_value = 0.0
            elif (model_fact.concept.baseXbrliTypeQname == qname('http://www.xbrl.org/2003/instance',
                                                                 'integerItemType') or
                  model_fact.concept.baseXbrliTypeQname == qname('http://www.xbrl.org/2003/instance',
                                                                 'nonPositiveIntegerItemType') or
                  model_fact.concept.baseXbrliTypeQname == qname('http://www.xbrl.org/2003/instance',
                                                                 'longItemType') or
                  model_fact.concept.baseXbrliTypeQname == qname('http://www.xbrl.org/2003/instance',
                                                                 'intItemType') or
                  model_fact.concept.baseXbrliTypeQname == qname('http://www.xbrl.org/2003/instance',
                                                                 'shortItemType') or
                  model_fact.concept.baseXbrliTypeQname == qname('http://www.xbrl.org/2003/instance',
                                                                 'nonNegativeIntegerItemType') or
                  model_fact.concept.baseXbrliTypeQname == qname('http://www.xbrl.org/2003/instance',
                                                                 'unsignedLongItemType') or

                  model_fact.concept.baseXbrliTypeQname == qname('http://www.xbrl.org/2003/instance',
                                                                 'unsignedIntItemType') or
                  model_fact.concept.baseXbrliTypeQname == qname('http://www.xbrl.org/2003/instance',
                                                                 'unsignedShortItemType')):
                xule_type = 'int'
                system_value = 0
            elif (model_fact.concept.baseXbrliTypeQname == qname('http://www.xbrl.org/2003/instance',
                                                               'stringItemType') or
                model_fact.concept.baseXbrliTypeQname == qname('http://www.xbrl.org/2003/instance',
                                                               'normalizedStringItemType')):
                xule_type = 'string'
                system_value = ''
            else:
                # There is no default value for the type of the fact
                system_value = model_fact
                xule_type = 'fact'
        else:
            system_value = model_fact
            xule_type = 'fact'

        fact_value = XuleValue(xule_context, system_value, xule_type,
                               alignment=None if alignment is None else frozenset(alignment.items()))
        #if not current_no_alignment:
        if alignment is not None:
            fact_value.aligned_result_only = True

        '''Where clause'''
        if 'whereExpr' in factset:

            # push the apsect variables
            ''' aspect_var_info is a tuple: 0 = aspect type, 1 = aspect name'''

            #                 aspect_vars_flat = list(aspect_vars.items())
            #                 for declaration_index, aspect_var_flat in enumerate(aspect_vars_flat, 1):
            for var_name, aspect_var_tuple in aspect_vars.items():
                aspect_type = aspect_var_tuple[0]
                aspect_name = aspect_var_tuple[1]
                # declaration_index = aspect_var_tuple[2]
                # declaration_id = (factset['node_id'], declaration_index)
                declaration_id = aspect_var_tuple[2]
                if aspect_type == 'builtin':
                    if aspect_name == 'concept':
                        xule_context.add_arg(var_name,
                                             declaration_id,
                                             None,
                                             XuleValue(xule_context, model_fact.concept, 'concept'),
                                             'single')
                    elif aspect_name == 'period':
                        if model_fact.context.isStartEndPeriod:
                            xule_context.add_arg(var_name,
                                                 declaration_id,
                                                 None,
                                                 XuleValue(xule_context, (model_fact.context.startDatetime,
                                                                          model_fact.context.endDatetime),
                                                           'duration', from_model=True),
                                                 'single')
                        elif model_fact.context.isInstantPeriod:
                            xule_context.add_arg(var_name,
                                                 declaration_id,
                                                 None,
                                                 XuleValue(xule_context, model_fact.context.instantDatetime,
                                                           'instant', from_model=True),
                                                 'single')
                        else:
                            xule_context.add_arg(var_name,
                                                 declaration_id,
                                                 None,
                                                 XuleValue(xule_context, (datetime.datetime.min, datetime.datetime.max),
                                                           'duration', from_model=True),
                                                 'single')
                    elif aspect_name == 'unit':
                        xule_context.add_arg(var_name,
                                             declaration_id,
                                             None,
                                             XuleValue(xule_context, model_to_xule_unit(model_fact.unit, xule_context),
                                                       'unit'),
                                             'single')
                    elif aspect_name == 'entity':
                        xule_context.add_arg(var_name,
                                             declaration_id,
                                             None,
                                             XuleValue(xule_context, model_fact.context.entityIdentifier,
                                                       'entity'),
                                             'single')
                    else:
                        raise XuleProcessingError(_("Unknown built in aspect '%s'" % aspect_name), xule_context)
                elif aspect_type == 'explicit_dimension':
                    model_dimension = model_fact.context.qnameDims.get(aspect_name)
                    if model_dimension is None:
                        member = XuleValue(xule_context, None, 'qname')
                    else:
                        if model_dimension.isExplicit:
                            member = XuleValue(xule_context, model_dimension.memberQname, 'qname')
                        else:
                            # This is a typed dimension
                            member = XuleValue(xule_context, model_dimension.typedMember.xValue,
                                               model_to_xule_type(xule_context, model_dimension.typedMember.xValue)[0])

                    xule_context.add_arg(var_name,
                                         declaration_id,
                                         None,  # tagged,
                                         member,
                                         'single')

            # add the $item variable for the fact
            xule_context.add_arg('fact',
                                 # (factset['node_id'], 0),
                                 factset['whereExpr']['node_id'],
                                 None,  # tagged
                                 fact_value,
                                 'single')

            '''The where clause is evaluated in a sub table.'''
            #                 switched_alignment = False
            #                 if xule_context.dependent_alignment is None and alignment is not None:
            #                     switched_alignment = True
            #                     #xule_context.iteration_table.current_table.dependent_alignment = frozenset(alignment.items())
            save_aligned_result_only = xule_context.aligned_result_only
            # save_used_vars = xule_context.used_vars
            save_used_expressions = xule_context.used_expressions

            # pre_aggregation_table_list_size = len(xule_context.iteration_table)
            where_table = xule_context.iteration_table.add_table(factset['whereExpr']['node_id'],
                                                                 xule_context.get_processing_id(
                                                                     factset['whereExpr']['node_id']),
                                                                 is_aggregation=True)
            #                 if switched_alignment:
            #                     where_table.dependent_alignment = frozenset(alignment.items())

            try:
                while True:
                    xule_context.aligned_result_only = False
                    # xule_context.used_vars = []
                    xule_context.used_expressions = set()
                    where_matched = True
                    try:
                        where_value = evaluate(factset['whereExpr'], xule_context)
                    except XuleIterationStop as xis:
                        where_value = xis.stop_value  # XuleValue(xule_context, None, 'unbound')
                    finally:
                        #                             if xule_context.iteration_table.current_table.current_alignment is None and xule_context.aligned_result_only:
                        #                                 where_value = XuleValue(xule_context, None, 'unbound')
                        xule_context.aligned_result_only = save_aligned_result_only

                    if alignment is not None:
                        # if not current_no_alignment and xule_context.iteration_table.is_dependent:
                        # if not current_no_alignment and factset['is_dependent']:
                        if factset['is_dependent']:
                            if xule_context.dependent_alignment is not None:
                                if frozenset(alignment.items()) != xule_context.dependent_alignment:
                                    # If this is in a 'with' clause, the first factset to be added to the with/agg table may be empty, The current alignment will be
                                    # from a higher table which will not inlucde the with filter aspects.
                                    if len(
                                            nested_filters) > 0 and xule_context.iteration_table.current_table.current_alignment is None:
                                        remove_aspects = [(nested_filter[0], nested_filter[1]) for nested_filter in
                                                          nested_filters]
                                        adjusted_alignment = remove_from_alignment(frozenset(alignment.items()),
                                                                                   remove_aspects, xule_context)
                                        if adjusted_alignment != xule_context.dependent_alignment:
                                            where_matched = False
                                    else:
                                        where_matched = False

                    if where_matched:
                        default_used_expressions.update(set(xule_context.used_expressions))
                        if where_value.type in ('unbound', 'none'):
                            pass
                        elif where_value.type == 'bool':
                            if where_value.value:
                                new_fact_value = copy.copy(fact_value)
                                new_fact_value.facts = xule_context.iteration_table.facts.copy()
                                new_fact_value.tags = xule_context.iteration_table.tags.copy()
                                # new_fact_value.used_vars = get_used_vars(xule_context, pre_matched_used_var_ids + xule_context.used_vars)
                                new_fact_value.used_expressions = pre_matched_used_expressions_ids | xule_context.used_expressions
                                results.append(new_fact_value)
                            '''It may be that the false value should also be included with an unbound value'''
                        else:
                            raise XuleProcessingError(_(
                                "Where clause in a factset did not evaluate to a boolean. Found '%s'." % where_value.type),
                                                      xule_context)

                    # xule_context.iteration_table.del_current()
                    # if len(xule_context.iteration_table) == pre_aggregation_table_list_size:
                    xule_context.iteration_table.next(where_table.table_id)
                    if where_table.is_empty:
                        break

            finally:
                # remove $item
                xule_context.del_arg('fact',  # (factset['node_id'], 0),
                                     factset['whereExpr']['node_id'], )
                # remove aspect variables
                for var_name, aspect_var_tuple in aspect_vars.items():
                    declaration_id = aspect_var_tuple[2]
                    xule_context.del_arg(var_name, declaration_id)
                # remove where table (if this is the result of an exception the where table may be left behind)
                xule_context.iteration_table.del_table(where_table.table_id)
                # restore aligned results, used_vars and used_expressions
                # xule_context.aligned_result_only = save_aligned_result_only
                # xule_context.used_vars = save_used_vars
                xule_context.used_expressions = save_used_expressions
        else:
            # fact_value.used_vars = get_used_vars(xule_context, pre_matched_used_var_ids)
            fact_value.used_expressions = pre_matched_used_expressions_ids
            results.append(fact_value)

    return results, default_used_expressions


# def evaluate_dict(dict_expr, xule_context):
#     result_dict = dict()
#     result_shadow = dict()
#
#     for pair in dict_expr['items']:
#         key = evaluate(pair['key'], xule_context)
#         value = evaluate(pair['value'], xule_context)
#
#         if key.type == 'dictionary':
#             raise XuleProcessingError(_("Key to a dictionary cannot be a dictionary."), xule_context)
#
#         if key.shadow_collection if key.type in ('set', 'list') else key.value not in result_shadow:
#             result_dict[key] = value
#             result_shadow[key.shadow_collection if key.type in (
#             'set', 'list') else key.value] = value.shadow_collection if value.type in (
#             'set', 'list', 'dictionary') else value.value
#
#     return XuleValue(xule_context, frozenset(result_dict.items()), 'dictionary',
#                      shadow_collection=frozenset(result_shadow.items()))
#
#
# def evaluate_list(list_expr, xule_context):
#     result = list()
#     result_shadow = list()
#     for item in list_expr['items']:
#         item_value = evaluate(item, xule_context)
#         result.append(item_value)
#         result_shadow.append(item_value.value)
#
#     return XuleValue(xule_context, tuple(result), 'list', shadow_collection=tuple(result_shadow))
#
#
# def evaluate_set(set_expr, xule_context):
#     result = list()
#     result_shadow = list()
#     for item in set_expr['items']:
#         item_value = evaluate(item, xule_context)
#         if item_value.shadow_collection if item_value.type in (
#         'set', 'list', 'dictionary') else item_value.value not in result_shadow:
#             result.append(item_value)
#             result_shadow.append(item_value.value)
#
#     return XuleValue(xule_context, frozenset(result), 'set', shadow_collection=frozenset(result_shadow))


def evaluate_filter(filter_expr, xule_context):
    collection_value = evaluate(filter_expr['expr'], xule_context)

    if collection_value.type not in ('set', 'list'):
        raise XuleProcessingError(
            _("Filter expresssion can only be used on a 'set' or 'list', found '{}'.".format(collection_value.type)),
            xule_context)

    # do nothing if there is no filtering
    if 'whereExpr' not in filter_expr and 'returnsExpr' not in filter_expr:
        return collection_value

    if collection_value.type == 'set':
        results = set()
        results_shadow = set()
    else:  # list
        results = list()
        results_shadow = list()

    for item_number, item_value in enumerate(collection_value.value):
        xule_context.add_arg('item',
                             filter_expr['expr']['node_id'],
                             None,
                             item_value,
                             'single')
        try:
            xule_context.column_prefix.append("{}-{}".format(filter_expr['node_id'], item_number))
            try:
                keep = True
                if 'whereExpr' in filter_expr:
                    keep = False
                    filter_where_result = evaluate(filter_expr['whereExpr'], xule_context)

                    if filter_where_result.type == 'bool':
                        keep = filter_where_result.value
                    elif filter_where_result.type not in ('unbound', 'none'):
                        raise XuleProcessingError(_(
                            "The where clause on a filter expression must evaluate to a boolean, found '{}'.".format(
                                filter_where_result.type)), xule_context)

                if keep:
                    if 'returnsExpr' in filter_expr:
                        keep_item = evaluate(filter_expr['returnsExpr'], xule_context)
                    else:
                        keep_item = item_value

                    if collection_value.type == 'set':
                        if (keep_item.shadow_collection if keep_item.type in (
                        'list', 'set', 'dictionary') else keep_item.value) not in results_shadow:
                            results.add(keep_item)
                            results_shadow.add(keep_item.shadow_collection if keep_item.type in (
                            'list', 'set', 'dictionary') else keep_item.value)
                        # otherwise, this a duplicate
                    else:  # list
                        results.append(keep_item)
                        results_shadow.append(keep_item.shadow_collection if keep_item.type in (
                        'list', 'set', 'dictionary') else keep_item.value)

            finally:
                # remove the args
                xule_context.del_arg('item',
                                    filter_expr['expr']['node_id'])
        finally:
            xule_context.column_prefix.pop()

    if collection_value.type == 'set':
        return XuleValue(xule_context, frozenset(results), 'set', shadow_collection=frozenset(results_shadow))
    else:  # list
        return XuleValue(xule_context, tuple(results), 'list', shadow_collection=tuple(results_shadow))


def evaluate_navigate(nav_expr, xule_context):
    # Get the taxonomy
    if 'taxonomy' in nav_expr:
        dts_value = evaluate(nav_expr['taxonomy'], xule_context)
        if dts_value.type != 'taxonomy':
            raise XuleProcessingError(
                _("Expecting a taxonomy for the 'taxonomy' clause of navigate. Found {}.".format(dts_value.type)),
                xule_context)
        dts = dts_value.value
    else:
        # Default to the taxonomy of the instance
        dts = xule_context.model

    # Set up the variables for the results of the traversal
    return_by_networks = nav_expr.get('return', dict()).get('byNetwork', False)
    if return_by_networks:
        results_by_networks = dict()
    result_items = list()

    # Determine if we are returning paths.
    paths = nav_expr.get('return', dict()).get('paths', False)

    return_names = get_return_component_names(nav_expr, xule_context)

    # Get the from and to concepts if they are supplied in the expression
    nav_from_concepts = nav_get_element(nav_expr, 'from', dts, xule_context)
    nav_to_concepts = nav_get_element(nav_expr, 'to', dts, xule_context)

    # This checks if there was 'from' concept in the expression, but no concepts were return. Same for the 'to' concept in the expression. Then there is not navigation to do
    if (nav_from_concepts is None or len(nav_from_concepts) > 0) and (
            nav_to_concepts is None or len(nav_to_concepts) > 0):
        arcroles = nav_get_role(nav_expr, 'arcrole', dts, xule_context)
        roles = nav_get_role(nav_expr, 'role', dts, xule_context)
        link_qname = evaluate(nav_expr['linkbase'], xule_context).value if 'linkbase' in nav_expr else None
        arc_qname = None  # This is always none.
        dimension_arcroles = None
        relationship_sets = list()
        # Find the relationships
        if not (('arcrole' in nav_expr and arcroles is None) or
                ('role' in nav_expr and roles is None) or
                ('linkbase' in nav_expr and link_qname is None)):
            # if the arcrole, or linkbase was provided, but it was None, then there are no relationshp sets to process and the relationships_sets will be left empty.

            # get the relationships
            if nav_expr.get('dimensional'):
                drs_roles = nav_get_role(nav_expr, 'drsRole', dts, xule_context)
                for drs_role in (drs_roles or (None,)):
                    table_concepts = nav_get_element(nav_expr, 'cube', dts, xule_context)
                    if arcroles is not None and len(arcroles) > 1:
                        raise XuleProcessingError(_(f"Dimensional navigation can only work with a single arcrole supplied. found {len(arcroles)}: {', '.join(arcroles)}"), xule_context)
                    arcrole = (arcroles or (None,))[0]
                    if arcrole is not None:
                        dimension_arcroles = xc.DIMENSION_PSEDDO_ARCROLES.get(arcrole, ('all', {arcrole, }))
                    relationship_sets += [XuleDimensionCube(dts, *x) for x in XuleDimensionCube.base_dimension_sets(dts) if
                                        ((drs_role is None or x[XuleDimensionCube.DIMENSION_SET_ROLE] == drs_role) and
                                        (table_concepts is None or x[
                                            XuleDimensionCube.DIMENSION_SET_HYPERCUBE] in table_concepts))]
            else:
                for arcrole in (arcroles or (None,)):
                    for role in (roles or (None,)):
                        relationship_set_infos = XuleProperties.get_base_set_info(dts, arcrole, role, link_qname, arc_qname)
                        relationship_sets += [
                            dts.relationshipSet(
                                relationship_set_info[XuleProperties.NETWORK_ARCROLE],
                                relationship_set_info[XuleProperties.NETWORK_ROLE],
                                relationship_set_info[XuleProperties.NETWORK_LINK],
                                relationship_set_info[XuleProperties.NETWORK_ARC]
                            ) for relationship_set_info in relationship_set_infos
                        ]

        direction = nav_expr['direction']
        include_start = nav_expr.get('includeStart', False)

        if nav_expr.get('acrossNetworks', False):
            # merge all the relationship sets into 1 pseudo relationship set
            relationship_sets = [XuleRelationshipSet(dts, relationship_sets),]

        for relationship_set in relationship_sets:
            if nav_from_concepts is None:
                # The from was not in tne navigate expression. Use the roots
                from_concepts = relationship_set.rootConcepts
            else:
                from_concepts = nav_from_concepts

            for from_concept in from_concepts:
                if direction == 'descendants':
                    for rel in nav_traverse(nav_expr, xule_context, 'down', relationship_set, from_concept,
                                            nav_to_concepts, int(nav_expr['depth']), return_names, dimension_arcroles):
                        result_items += nav_decorate(rel, 'down', return_names, include_start, paths, xule_context)
                if direction == 'children':
                    for rel in nav_traverse(nav_expr, xule_context, 'down', relationship_set, from_concept,
                                            nav_to_concepts, 1, return_names, dimension_arcroles):
                        result_items += nav_decorate(rel, 'down', return_names, include_start, paths, xule_context)
                if direction == 'ancestors':
                    for rel in nav_traverse(nav_expr, xule_context, 'up', relationship_set, from_concept,
                                            nav_to_concepts, int(nav_expr['depth']), return_names, dimension_arcroles):
                        result_items += nav_decorate(rel, 'up', return_names, include_start, paths, xule_context)
                if direction == 'parents':
                    for rel in nav_traverse(nav_expr, xule_context, 'up', relationship_set, from_concept,
                                            nav_to_concepts, 1, return_names, dimension_arcroles):
                        result_items += nav_decorate(rel, 'up', return_names, include_start, paths, xule_context)
                if direction == 'siblings':
                    for parent_rel in nav_traverse(nav_expr, xule_context, 'up', relationship_set, from_concept, None,
                                                   1, list(), dimension_arcroles, ignore_where=True):
                        for sibling_rel in nav_traverse(nav_expr, xule_context, 'down', relationship_set,
                                                        parent_rel['relationship'].fromModelObject, nav_to_concepts, 1,
                                                        return_names, dimension_arcroles):
                            if include_start or sibling_rel['relationship'] is not parent_rel['relationship']:
                                result_items += nav_decorate(sibling_rel, 'down', return_names, False, paths,
                                                             xule_context)
                if direction == 'previous-siblings':
                    for parent_rel in nav_traverse(nav_expr, xule_context, 'up', relationship_set, from_concept, None,
                                                   1, list(), dimension_arcroles, ignore_where=True):
                        for sibling_rel in nav_traverse(nav_expr, xule_context, 'down', relationship_set,
                                                        parent_rel['relationship'].fromModelObject, nav_to_concepts, 1,
                                                        return_names, dimension_arcroles):
                            if include_start or sibling_rel['relationship'] is not parent_rel['relationship']:
                                result_items += nav_decorate(sibling_rel, 'down', return_names, False, paths,
                                                             xule_context)
                            if sibling_rel['relationship'] is parent_rel['relationship']:
                                break  # We are done.
                if direction == 'following-siblings':
                    for parent_rel in nav_traverse(nav_expr, xule_context, 'up', relationship_set, from_concept, None,
                                                   1, list(), dimension_arcroles, ignore_where=True):
                        start_rel_found = False
                        for sibling_rel in nav_traverse(nav_expr, xule_context, 'down', relationship_set,
                                                        parent_rel['relationship'].fromModelObject, nav_to_concepts, 1,
                                                        return_names, dimension_arcroles):
                            if sibling_rel['relationship'] is parent_rel['relationship']:
                                start_rel_found = True
                            if start_rel_found:
                                if include_start or sibling_rel['relationship'] is not parent_rel['relationship']:
                                    result_items += nav_decorate(sibling_rel, 'down', return_names, False, paths,
                                                                 xule_context)

            if return_by_networks:
                if len(result_items) > 0:
                    results_by_networks[get_network_info(relationship_set, xule_context)] = result_items
                result_items = list()

    if return_by_networks:
        return nav_finish_results(nav_expr, results_by_networks, 'result-order' in return_names, xule_context)
    else:
        return nav_finish_results(nav_expr, result_items, 'result-order' in return_names, xule_context)


def nav_traverse(nav_expr, xule_context, direction, network, parent, end_concepts, remaining_depth, return_names,
                 dimension_arcroles=None, previous_concepts=None, nav_depth=1, result_order=0,
                 arc_attribute_names=None, ignore_where=False):
    """Traverse a network
    
    Arguments:
        direction (string): Either 'down' or 'up'
        network (ModelRelationshipSet): The network of relationships.
        parent (ModelConcept): The parent concept
        end_concepts (set of ModelConcepts): A set of concepts if encountered the traversal should stop
        depth (int): How many levels to traverse. -1 is infinite depth.
        previous_concepts (set of ModelConcepts): List concepts already traversed. Used to prevent loops in the traversal
        include_start (boolean): Indicates that the starting concept should be included in the results
    
    Return:
        list of tuples of (ModelRelationship, top)
    """
    # Initialize previous_concepts
    if previous_concepts is None:  # this only happens on the intial call.
        previous_concepts = {parent, }
        arc_attribute_names = {x for x in return_names if isinstance(x, QName)}
        first_time = True
    else:
        first_time = False

    if end_concepts is None:
        end_concepts = set()

    # initialize depth
    if remaining_depth == -1:
        depth = float('inf')

    if remaining_depth == 0:
        return list()

    paths = nav_expr.get('return', dict()).get('paths', False)

    # 'children' are parents if the direction is up.
    children = list()

    if parent in end_concepts:
        # This should only happen if in the inital call to nav_traverse the parent is in the list of 
        # end concepts. In this case there is no navigation.
        return children

    children_method = network.fromModelObject if direction == 'down' else network.toModelObject
    for rel_number, rel in enumerate(sorted(children_method(parent), key=lambda x: x.order or 1), 1):
        inner_children = list()

        child = rel.toModelObject if direction == 'down' else rel.fromModelObject
        if not isinstance(child, ModelConcept):
            # only navigate concept to concept relationships
            continue

        rel_info = {'relationship': rel}

        if first_time:
            rel_info['first'] = True
        if 'network' in return_names:
            rel_info['network'] = get_network_info(network, xule_context)
        if 'navigation-order' in return_names:
            rel_info['navigation-order'] = rel_number
        if 'navigation-depth' in return_names:
            rel_info['navigation-depth'] = nav_depth
        #         if 'result-order' in return_names:
        #             rel_info['result-order'] = result_order
        for arc_attribute_name in arc_attribute_names:
            rel_info[arc_attribute_name] = rel.arcElement.get(arc_attribute_name.clarkNotation)

        # Decide if the child will be in the results. If the child is not in the results, the navigation does not stop.
        if not ignore_where:
            if not (
                    nav_traverse_where(nav_expr, 'whereExpr', rel, xule_context) and
                    (
                            dimension_arcroles is None or
                            'dimensional' not in nav_expr or
                            ('dimensional' in nav_expr and rel.arcrole in dimension_arcroles[
                                xc.DIMENSION_PSEUD0_ARCROLE_PART] and
                            (
                                    dimension_arcroles[xc.DIMENSION_PSEUD0_SIDE] == 'all' or
                                    rel.side == dimension_arcroles[xc.DIMENSION_PSEUD0_SIDE]
                            )
                            )
                    )
                ):
                rel_info['relationship'] = None

        keep_rel = rel_info

        if child not in previous_concepts:

            if child in end_concepts:
                # This is the end of the traversal because the child is a 'to' concept.
                if paths:
                    inner_children.append([keep_rel, ])
                else:
                    inner_children += [keep_rel, ]
            else:
                if nav_traverse_where(nav_expr, 'stopExpr', rel, xule_context):
                    next_children = list()
                else:
                    next_children = nav_traverse(nav_expr,
                                                 xule_context,
                                                 direction,
                                                 network,
                                                 child,
                                                 end_concepts,
                                                 remaining_depth - 1,
                                                 return_names,
                                                 dimension_arcroles,
                                                 previous_concepts | {child, },
                                                 nav_depth + 1,
                                                 result_order,
                                                 arc_attribute_names)

                if len(next_children) == 0 and len(end_concepts) > 0:  # The to concept was never found
                    # Reset the inner_child list. This will throw away all reseults that lead to this moment.
                    inner_children = list()
                else:
                    if paths:
                        if len(next_children) == 0:
                            # On a leaf. This is where a path is initially created.
                            inner_children.append([keep_rel, ])
                        else:
                            # There are children of the current child (keep_rel), add the current child to each of the paths from the current childs children
                            for i in next_children:
                                inner_children.append([keep_rel, ] + i)
                    else:
                        inner_children += [keep_rel, ] + next_children
        else:
            if keep_rel['relationship'] is not None:
                # indicates a cycle
                keep_rel['cycle'] = True
            if paths:
                inner_children.append([keep_rel, ])
            else:
                inner_children.append(keep_rel)

        children += inner_children
        # This only allows the first child of the initial call to nav_traverse to be marked as first.The first is used to indicate when to use
        # the from side of the relationship for 'ionclude start'.
        if not paths:
            first_time = False
    return children


def nav_traverse_where(nav_expr, clause_name, relationship, xule_context):
    if clause_name not in nav_expr:
        if clause_name == 'whereExpr':
            return True
        else:  # 'stopExpr'
            return False
    else:
        xule_context.add_arg('relationship',
                             nav_expr[clause_name]['node_id'],
                             None,
                             XuleValue(xule_context, relationship, 'relationship'),
                             'single')

        try:
            nav_where_results = evaluate(nav_expr[clause_name], xule_context)
        finally:
            # remove the args
            xule_context.del_arg('relationship',
                                 nav_expr[clause_name]['node_id'])

        if nav_where_results.type == 'bool':
            return nav_where_results.value
        elif nav_where_results.type in ('unbound', 'none'):
            return False
        elif nav_where_results.type not in ('unbound', 'none'):
            raise XuleProcessingError(_(
                "The {} clause on a navigation expression must evaluate to a boolean, found '{}'.".format(
                    clause_name[:clause_name.find('Expr')], nav_where_results.type)), xule_context)


def nav_get_role(nav_expr, role_type, dts, xule_context):
    """Get the full role from the navigation expression.
    
    A role in the navigation expressions is either a string, uri or a non prefixed qname. If it is a string or uri, it is a full arcrole. If it is
    a non prefixed qname, than the local name of the qname is used to match an arcrole that ends in 'localName'. If more than one arcrole is found then
    and error is raise. This allows short form of an arcrole i.e parent-child.
    """
    if role_type in nav_expr:
        expr_role_value = evaluate(nav_expr[role_type], xule_context)
        if expr_role_value.type in ('set', 'list'):
            role_values = expr_role_value.value
        else:
            role_values = (expr_role_value,) # a tuple of 1

        # return value will be a list of the role/arcrole uris
        return_values = []
        errors = []
        for role_value in role_values:
            if role_value.type in ('string', 'uri'):
                return_values.append(role_value.value)
            elif role_value.type == 'qname':
                return_values.extend(XuleUtility.resolve_role(role_value, role_type, dts, xule_context))
            elif role_value.type == 'role':
                return_values.append(role_value.value.roleURI)
            else:
                errors.append("Navigation is expecting a role (role, string, uri, or short role name), found '{}'.".format(role_value.type))

        if len(errors) > 0:
            raise XuleProcessingError(_('\n'.join(errors)), xule_context)
        if len(return_values) == 0:
            return None
        else:
            return return_values
    else:
        return None  # There is no arcrole on the navigation expression


def nav_get_element(nav_expr, side, dts, xule_context):
    """Get the element or set of elements on the from or to side of a navigation expression'
    
    This determines the from/to elements of a navigation expression. If the navigation expression includes the from/to component, this will be evaluated.
    The result can be a qname, concept or a set/list of qname or concepts.
    
    Arguments:
        nav_expr (dictionary): The navigation expression AST node
        side (string): Either 'from' or 'to'.
        xule_context (XuleRuleContext): The processing context
        
    Returns:
        None - indicates that the side is not in the navigation expression
        set of concepts - the set of the concepts if the side evaluates to a set or list of concept/concepts
    """

    if side in nav_expr:
        side_value = evaluate(nav_expr[side], xule_context)
        if side_value.type == 'qname':
            concept = XuleProperties.get_concept(dts, side_value.value)
            if concept is None:
                return set()
            else:
                return {concept, }
        elif side_value.type == 'concept':
            return {side_value.value, }
        elif side_value.type in ('set', 'list'):
            concepts = set()
            for item in side_value.value:
                if item.type == 'qname':
                    concept = XuleProperties.get_concept(dts, item.value)
                    if concept is not None:
                        concepts.add(concept)
                elif item.type == 'concept':
                    concepts.add(item.value)
                else:
                    raise XuleProcessingError(_(
                        "In navigation, expecting a collection of concepts or concepts, but found {}.".format(
                            item.type)))
            return concepts
        else:
            raise XuleProcessingError(
                _("In navigation, expecting a concept or qname, but found {}.".format(side_value.type)))
    else:
        return None  # The side is not in the navigation expression


def nav_decorate(rel, direction, return_names, include_start, paths, xule_context):
    """Determine what will be outputted for a single navigation item.
    
    Arguments:
        rel (dict): A dition of the relationship and additional information from the traversal
        direction (string): 'up' or 'down'. Direction of the traversal
        return_names (list of strings): List of the return components to decoreate the relationship.
        include_start (boolean or 'start_only'): Indicates if the start concept should be returned.
        xule_context (XuleRuleContext): Processing context
        
    Returns:
        A tuple of tuples of the return components for the relationship. The tuple is composed of:
        0. value
        1. xule type
        2. return component name
    """
    result = list()
    if paths:
        # In this case, rel is a list of relatioships in the path.
        path = list()
        for i in rel:
            if include_start and i.get('first', False):
                path.append(nav_decorate_gather_components(i, direction, return_names, True, xule_context))
            path.append(nav_decorate_gather_components(i, direction, return_names, False, xule_context))
        result.append(path)
    else:
        if rel['relationship'] is not None:
            if include_start and rel.get('first', False):
                result.append(nav_decorate_gather_components(rel, direction, return_names, True, xule_context))
            result.append(nav_decorate_gather_components(rel, direction, return_names, False, xule_context))

    return result


def get_return_component_names(nav_expr, xule_context):
    component_names = list()

    if 'return' in nav_expr:
        if 'returnComponents' in nav_expr['return']:
            for component in nav_expr['return']['returnComponents']:
                if isinstance(component, dict):
                    # This is an expression which should evluate to a qname. This is an arc attribute name.
                    component_name_value = evaluate(component, xule_context)
                    if component_name_value.type != 'qname':
                        raise XuleProcessingError(_(
                            "Expression in return components of a navigate expression did not evaluate to a qname. Expecting a qname for the name of a arc attribute. Found {}").format(
                            component_name_value.type), xule_context)
                    component_names.append(component_name_value.value)
                else:
                    component_names.append(component)
        else:
            component_names.append('target')
    else:
        component_names.append('target')
    return component_names

    # nav_expr.get('return', {'returnComponents': ('target',)}).get('returnComponents', ('target',))


def nav_decorate_gather_components(rel, direction, component_names, is_start, xule_context):
    """Get the values for all the return components for a relationship.
    
    Arguments:
        rel (dict): A dition of the relationship and additional information from the traversal
        direction (string): 'up' or 'down'. Direction of the traversal
        component_names (list of strings): The list of return components to get for the relationship.
        is_start (boolean): Indicates if the relationship should be treated as a starting relationship. This is used if the 'include start' keyword is
                            is included in the navigation expression and the relationship is from the start of the navigation. In this case there is
                            an extra result for the relationships for the starting concept.
        xule_context (XuleRuleContext): The processing context.
    Returns:
        A tupple of return component values. Each value is a tuple of:
            1. component value
            2. xule type for the value
            3. component name (used for ease of debugging)
    """
    result = list()

    for component_name in component_names:
        if component_name == 'result-order':
            result.append('result-order')  # This will be handled in the finishing step
        else:
            result.append(nav_decorate_component_value(rel, direction, component_name, is_start, xule_context))
    return result


def nav_decorate_component_value(rel, direction, component_name, is_start, xule_context):
    """Get the return component value for a relationship and a single return component.
    
    Arguments:
        rel (dict): A dition of the relationship and additional information from the traversal
        direction (string): 'up' or 'down'. Direction of the traversal
        component_name (string): Component name
        is_start (boolean): Indicates if the relationship should be treated as a starting relationship. This is used if the 'include start' keyword is
                            is included in the navigation expression and the relationship is from the start of the navigation. In this case there is
                            an extra result for the relationships for the starting concept.
        xule_context (XuleRuleContext): The processing context.
    Returns:
        A tuple of:
            1. component value
            2. xule type for the value
            3. component name (used for ease of debugging)
    """

    if component_name in NAVIGATE_RETURN_COMPONENTS:
        if rel['relationship'] is None:
            return (None, 'none', component_name)

        if NAVIGATE_RETURN_COMPONENTS[component_name][NAVIGATE_ALLOWED_ON_START]:
            if rel['relationship'] is None:
                return (None, 'none', component_name)
            else:
                return NAVIGATE_RETURN_COMPONENTS[component_name][NAVIGATE_RETURN_FUCTION](rel, direction,
                                                                                           component_name, is_start,
                                                                                           xule_context)
        else:
            if is_start:
                return (None, 'none', component_name)
            else:
                if rel['relationship'] is None:
                    return (None, 'none', component_name)
                else:
                    return NAVIGATE_RETURN_COMPONENTS[component_name][NAVIGATE_RETURN_FUCTION](rel, direction,
                                                                                               component_name, is_start, 
                                                                                               xule_context)
    else:
        # Could be an arc attribute name (qname)
        if isinstance(component_name, QName):
            if rel['relationship'] is None:
                return (None, 'none', component_name)
            else:
                attribute_value = rel[component_name]
                if attribute_value is None:
                    return (None, 'none', component_name)
                else:
                    return (attribute_value, 'string', component_name)
        else:
            raise XuleProcessingError(_("Component {} is not currently supported.".format(component_name)),
                                      xule_context)


def nav_decorate_component_target(rel, direction, component_name, is_start, xule_context):
    if is_start:
        # If it is the start concept, then get the opposide side of the relationship. 
        if component_name == 'target':
            if direction == 'up':
                return (rel['relationship'].toModelObject, 'concept', component_name)
            else:
                return (rel['relationship'].fromModelObject, 'concept', component_name)
        elif component_name == 'target-name':
            if direction == 'up':
                return (rel['relationship'].toModelObject.qname, 'qname', component_name)
            else:
                return (rel['relationship'].fromModelObject.qname, 'qname', component_name)
    else:
        if component_name == 'target':
            if direction == 'up':
                return (rel['relationship'].fromModelObject, 'concept', component_name)
            else:
                return (rel['relationship'].toModelObject, 'concept', component_name)
        if component_name == 'target-name':
            if direction == 'up':
                return (rel['relationship'].fromModelObject.qname, 'qname', component_name)
            else:
                return (rel['relationship'].toModelObject.qname, 'qname', component_name)


def nav_decorate_component_source(rel, direction, component_name, is_start, xule_context):
    if component_name == 'source':
        if direction == 'up':
            return (rel['relationship'].toModelObject, 'concept', component_name)
        else:
            return (rel['relationship'].fromModelObject, 'concept', component_name)
    elif component_name == 'source-name':
        if direction == 'up':
            return (rel['relationship'].toModelObject.qname, 'qname', component_name)
        else:
            return (rel['relationship'].fromModelObject.qname, 'qname', component_name)


def nav_decorate_component_order(rel, direction, component_name, is_start, xule_context):
    return (rel['relationship'].orderDecimal, 'decimal', component_name)


def nav_decorate_component_weight(rel, direction, component_name, is_start, xule_context):
    if rel['relationship'].weightDecimal is None:
        return (None, 'none', component_name)
    else:
        return (rel['relationship'].weightDecimal, 'decimal', component_name)


def nav_decorate_component_preferred_label_role(rel, direction, component_name, is_start, xule_context):
    if rel['relationship'].preferredLabel is not None:
        #return (rel['relationship'].preferredLabel, 'uri', component_name)

        return (XuleUtility.role_uri_to_model_role(rel['relationship'].modelXbrl, rel['relationship'].preferredLabel), 'role', component_name)
    else:
        return (None, 'none', component_name)


def nav_decorate_component_preferred_label(rel, direction, component_name, is_start, xule_context):
    if rel['relationship'].preferredLabel is not None:
        label = XuleProperties.get_label(xule_context, rel['relationship'].toModelObject, rel['relationship'].preferredLabel, None)
        if label is None:
            return (None, 'none', component_name)
        else:
            return (label, 'label', component_name)
    else:
        return (None, 'none', component_name)


def nav_decorate_component_relationship(rel, direction, component_name, is_start, xule_context):
    return (rel['relationship'], 'relationship', component_name)


def nav_decorate_component_role(rel, direction, component_name, is_start, xule_context):
    role = get_role(rel['relationship'], xule_context)
    return (role, 'role', component_name)


def get_role(relationship, xule_context):
    role_uri = relationship.linkrole
    if role_uri in relationship.modelXbrl.roleTypes:
        return relationship.modelXbrl.roleTypes[role_uri][0]
    else:
        return XuleRole(role_uri)


def nav_decorate_component_role_uri(rel, direction, component_name, is_start, xule_context):
    role_uri = rel['relationship'].linkrole

    return (role_uri, 'uri', component_name)


def nav_decorate_component_role_description(rel, direction, component_name, is_start, xule_context):
    role = get_role(rel['relationship'], xule_context)

    return (role.definition, 'string', component_name)


def nav_decorate_component_arcrole(rel, direction, component_name, is_start, xule_context):
    arcrole = get_arcrole(rel['relationship'], xule_context)
    return (arcrole, 'role', component_name)


def get_arcrole(relationship, xule_context):
    arcrole_uri = relationship.arcrole
    if arcrole_uri in relationship.modelXbrl.arcroleTypes:
        return relationship.modelXbrl.arcroleTypes[arcrole_uri][0]
    else:
        return XuleArcrole(arcrole_uri)


def nav_decorate_component_arcrole_uri(rel, direction, component_name, is_start, xule_context):
    return (rel['relationship'].arcrole, 'uri', component_name)


def nav_decorate_component_arcrole_description(rel, direction, component_name, is_start, xule_context):
    arcrole = get_arcrole(rel['relationship'], xule_context)
    return (arcrole.definition, 'string', component_name)


def nav_decorate_component_cycles_allowed(rel, direction, component_name, is_start, xule_context):
    arcrole = get_arcrole(rel['relationship'], xule_context)
    return (arcrole.cyclesAllowed, 'string', component_name)


def nav_decorate_component_link_name(rel, direction, component_name, is_start, xule_context):
    return (rel['relationship'].linkQname, 'qname', component_name)


def nav_decorate_component_arc_name(rel, direction, component_name, is_start, xule_context):
    return (rel['relationship'].qname, 'qname', component_name)


def nav_decorate_component_network(rel, direction, component_name, is_start, xule_context):
    return (rel['network'], 'network', component_name)


def get_network_info(network, xule_context):
    network_info = (network.arcrole, network.linkrole, network.arcqname, network.linkqname, False)
    return (network_info, network)


def nav_decorate_component_cycle(rel, direction, component_name, is_start, xule_context):
    return (False if is_start else rel.get('cycle', False), 'bool', component_name)

def nav_decorate_component_navigation_order(rel, direction, component_name, is_start, xule_context):
    if is_start:
        return (0, 'int', component_name)
    else:
        return (rel['navigation-order'], 'int', component_name)


def nav_decorate_component_navigation_depth(rel, direction, component_name, is_start, xule_context):
    if is_start:
        return (0, 'int', component_name)
    else:
        return (rel['navigation-depth'], 'int', component_name)


def nav_decorate_component_result_order(rel, direction, component_name, is_start, xule_context):
    if is_start:
        return (0, 'int', component_name)
    else:
        return (rel['result-order'], 'int', component_name)


def nav_decorate_component_drs_role(rel, direction, component_name, is_start, xule_context):
    # if hasattr(rel['relationship'], 'dimension_set'):
    if isinstance(rel['relationship'], DimensionRelationship):
        return (rel['relationship'].dimension_set.drs_role, 'role', component_name)
    else:
        return (None, 'none', component_name)


def nav_decorate_component_dimension_type(rel, direction, component_name, is_start, xule_context):
    if isinstance(rel['relationship'], DimensionRelationship):
        if (is_start and direction == 'down') or (not is_start and direction == 'up'):
            dim_type = rel['relationship'].fromDimensionType
        else:
            dim_type = rel['relationship'].toDimensionType

        if dim_type is None:
            return (None, 'none', component_name)
        else:
            return (dim_type, 'string', component_name)
    else:
        return (None, 'none', component_name)


def nav_decorate_component_dimension_sub_type(rel, direction, component_name, is_start, xule_context):
    if isinstance(rel['relationship'], DimensionRelationship):
        if (is_start and direction == 'down') or (not is_start and direction == 'up'):
            dim_type = rel['relationship'].fromDimensionSubType
        else:
            dim_type = rel['relationship'].toDimensionSubType

        if dim_type is None:
            return (None, 'none', component_name)
        else:
            return (dim_type, 'string', component_name)
    else:
        return (None, 'none', component_name)


def nav_decorate_component_usable(rel, direction, component_name, is_start, xule_context):
    if isinstance(rel['relationship'], DimensionRelationship):
        if (is_start and direction == 'down') or (not is_start and direction == 'up'):
            concept = rel['relationship'].fromModelObject
        else:
            concept = rel['relationship'].toModelObject

        if rel['relationship'].dimension_set.isUsable(concept) is None:
            return (None, 'none', component_name)
        else:
            return (rel['relationship'].dimension_set.isUsable(concept), 'bool', component_name)
    else:
        return (None, 'none', component_name)


# Navigation return component tuple locations
NAVIGATE_RETURN_FUCTION = 0
NAVIGATE_ALLOWED_ON_START = 1

NAVIGATE_RETURN_COMPONENTS = {'source': (nav_decorate_component_source, False),
                              'source-name': (nav_decorate_component_source, False),
                              'target': (nav_decorate_component_target, True),
                              'target-name': (nav_decorate_component_target, True),
                              'order': (nav_decorate_component_order, False),
                              'weight': (nav_decorate_component_weight, False),
                              'preferred-label-role': (nav_decorate_component_preferred_label_role, False),
                              'preferred-label': (nav_decorate_component_preferred_label, False),
                              'relationship': (nav_decorate_component_relationship, False),
                              'role': (nav_decorate_component_role, True),
                              'role-uri': (nav_decorate_component_role_uri, True),
                              'role-description': (nav_decorate_component_role_description, True),
                              'arcrole': (nav_decorate_component_arcrole, True),
                              'arcrole-uri': (nav_decorate_component_arcrole_uri, True),
                              'arcrole-description': (nav_decorate_component_arcrole_description, True),
                              'arcrole-cycles-allowed': (nav_decorate_component_cycles_allowed, True),
                              'link-name': (nav_decorate_component_link_name, True),
                              'arc-name': (nav_decorate_component_arc_name, True),
                              'network': (nav_decorate_component_network, True),
                              'cycle': (nav_decorate_component_cycle, True),
                              'navigation-order': (nav_decorate_component_navigation_order, True),
                              'navigation-depth': (nav_decorate_component_navigation_depth, True),
                              'result-order': (nav_decorate_component_result_order, True),
                              'drs-role': (nav_decorate_component_drs_role, True),
                              'dimension-type': (nav_decorate_component_dimension_type, True),
                              'dimension-sub-type': (nav_decorate_component_dimension_sub_type, True),
                              'usable': (nav_decorate_component_usable, True),
                              }


def nav_finish_results(nav_expr, return_items, add_result_order, xule_context):
    """Format the results of navigation.
    
    This function processes the list of results and puts them in their final form.
    The options for the final forma are if the results are:
     1. a list of relationships
     2. a set of relations
     3. a dictionary organzed by network of relationships
     4. a nested list of paths of the traversal. The outer list is a path and the inner list contains the relationships that make up the path.

    Arguments:
        nav_expr (dictionary): The navigation expression AST node
        return_items (list): The list of the decorated return items. 
        add_result_order (bool): An indicator if the result order should be added to the results. This is calculated as the return_items are processed.
        xule_context (XuleRuleContext): The processing context.
        
    Returns a XuleValue which is the final result of the navigation.
    """

    if nav_expr.get('return', dict()).get('byNetwork', False):
        by_network = dict()
        by_network_shadow = dict()
        for network, network_return_items in return_items.items():
            processed_items = nav_finish_return_items(nav_expr, network_return_items, add_result_order, xule_context)
            by_network[XuleValue(xule_context, network, 'network')] = processed_items
            by_network_shadow[network] = processed_items.shadow_collection
        return XuleValue(xule_context, frozenset(by_network.items()), 'dictionary',
                         shadow_collection=frozenset(by_network_shadow.items()))
    else:
        return nav_finish_return_items(nav_expr, return_items, add_result_order, xule_context)


def nav_finish_return_items(nav_expr, return_items, add_result_order, xule_context):
    paths = nav_expr.get('return', dict()).get('paths', False)

    if 'return' in nav_expr:
        return_type = nav_expr['return'].get('returnType', 'list')
    else:
        return_type = 'list'

    # The return_component_type determins if the return concompents are a list or a dictionary keyed by the component name.
    return_component_type = nav_expr.get('return', dict()).get('returnComponentType', 'list')

    cur_order = 1

    final_results = list()
    final_shadow = list()

    def handle_return_item(return_item, cur_order):

        if add_result_order:
            # replace reuslt-order
            return_item = [x if x != 'result-order' else (cur_order, 'int', 'result-order') for x in return_item]
            # return_item.append((cur_order, 'int', 'result-order'))
            cur_order += 1

        if return_component_type == 'list':
            if len(return_item) == 1:
                # A list of single items is returned. 
                # The return item only has one return component
                return XuleValue(xule_context, return_item[0][0], return_item[0][1]), return_item[0][0], cur_order
            #
            #                 if return_type == 'list' or (return_type == 'set' and return_item[0][0] not in results_shadow):
            #                     if return_component_type == 'list':
            #                         results.append(XuleValue(xule_context, return_item[0][0], return_item[0][1]))
            #                         results_shadow.append(return_item[0][0])
            else:
                # A list of list of components is returned.
                # The return_item has multiple return componenets
                multi_result = list()
                multi_shadow = list()
                for return_component in return_item:
                    multi_result.append(XuleValue(xule_context, return_component[0], return_component[1]))
                    multi_shadow.append(return_component[0])
                multi_shadow_tuple = tuple(multi_shadow)

                return XuleValue(xule_context, tuple(multi_result), 'list',
                                 shadow_collection=tuple(multi_shadow)), tuple(multi_shadow_tuple), cur_order

        #                 if return_type == 'list' or (return_type == 'set' and  multi_shadow_tuple not in results_shadow):
        #                     results.append(XuleValue(xule_context, tuple(multi_result), 'list', shadow_collection=tuple(multi_shadow)))
        #                     results_shadow.append(multi_shadow_tuple)
        else:  # the return_component_type is a dictionary
            multi_result = dict()
            multi_shadow = dict()
            for return_component in return_item:
                multi_result[XuleValue(xule_context, return_component[2], 'string')] = XuleValue(xule_context,
                                                                                                 return_component[0],
                                                                                                 return_component[1])
                multi_shadow[return_component[2]] = return_component[0]

            return XuleValue(xule_context, frozenset(multi_result.items()), 'dictionary'), frozenset(
                multi_shadow.items()), cur_order

    #
    #             if return_type == 'list' or (return_type == 'set' and frozenset(multi_shadow.items()) not in results_shadow):
    #                 results.append(XuleValue(xule_context, frozenset(multi_result.items()), 'dictionary'))
    #                 results_shadow.append(frozenset(multi_shadow.items()))
    #
    #         return results, results_shadow

    for return_item in return_items:
        if paths:
            path_result = list()
            path_shadow = list()
            # The return item is a list of the items that make up a path. Process each item in the path.
            for path_item in return_item:
                path_item_result, path_item_shadow, cur_order = handle_return_item(path_item, cur_order)
                path_result.append(path_item_result)
                path_shadow.append(path_item_shadow)
            # Now the path is complete
            item_result = XuleValue(xule_context, tuple(path_result), 'list', shadow_collection=tuple(path_shadow))
            item_shadow = tuple(path_shadow)

        else:
            item_result, item_shadow, cur_order = handle_return_item(return_item, cur_order)

        if return_type == 'list':
            final_results.append(item_result)
            final_shadow.append(item_shadow)
        else:  # Set
            if item_shadow not in final_shadow:
                final_results.append(item_result)
                final_shadow.append(item_shadow)

    if return_type == 'list':
        return XuleValue(xule_context, tuple(final_results), 'list', shadow_collection=tuple(final_shadow))
    else:
        return XuleValue(xule_context, frozenset(final_results), 'set', shadow_collection=frozenset(final_shadow))


def evaluate_function_ref(function_ref, xule_context):
    if function_ref['functionName'] in BUILTIN_FUNCTIONS:
        function_info = BUILTIN_FUNCTIONS.get(function_ref['functionName'])
        if function_info[FUNCTION_TYPE] == 'aggregate':
            return evaluate_aggregate_function(function_ref, function_info, xule_context)
        elif function_info[FUNCTION_TYPE] == 'regular':
            return regular_function(xule_context, function_ref, function_info)
        else:
            raise XuleProcessingError(_("Unknown function type '{}'.".format(function_info[FUNCTION_TYPE])),
                                      xule_context)
    elif function_ref['functionName'] in XuleProperties.PROPERTIES:
        return property_as_function(xule_context, function_ref)
    else:
        # xule defined function
        return user_defined_function(xule_context, function_ref)


def property_as_function(xule_context, function_ref):
    """Evaluate a function that is a property.
    
    Some functions are just a function version of a property. In these cases, the first argument is the object of the property and the rest of the args are
    arguments to the property. 
    
    Example:
        -10.abs is the same as abs(-10)
    """
    # Get the property information
    property_info = XuleProperties.PROPERTIES[function_ref['functionName']]

    # Check that there is at least one argument. This is the property object
    if len(function_ref['functionArgs']) == 0:
        raise XuleProcessingError(
            _("The '{}' function must have at least one argument, found none.".format(function_ref['functionName'])),
            xule_context)

    # Check that the first argument is the right type
    property_object = evaluate(function_ref['functionArgs'][0], xule_context)

    if len(property_info[XuleProperties.PROP_OPERAND_TYPES]) > 0:
        # if not (property_object.type in property_info[XuleProperties.PROP_OPERAND_TYPES] or
        #         property_object.is_fact and 'fact' in property_info[XuleProperties.PROP_OPERAND_TYPES] or
        #         any([xule_castable(property_object, allowable_type, xule_context) for allowable_type in
        #              property_info[XuleProperties.PROP_OPERAND_TYPES]])):

        if not (property_object.type in property_info[XuleProperties.PROP_OPERAND_TYPES] or
                property_object.is_fact and 'fact' in property_info[XuleProperties.PROP_OPERAND_TYPES] or
                any([xule_castable(property_object, allowable_type, xule_context) for allowable_type in
                        property_info[XuleProperties.PROP_OPERAND_TYPES]]) or
                'noncollection' in property_info[XuleProperties.PROP_OPERAND_TYPES] or
                (property_object.type in ('none', 'unbound') and property_info[XuleProperties.PROP_UNBOUND_ALLOWED])):
            raise XuleProcessingError(
                _("The first argument of function '{}' must be {}, found '{}'.".format(function_ref['functionName'],
                                                                                       ', '.join(property_info[
                                                                                                     XuleProperties.PROP_OPERAND_TYPES]),
                                                                                       property_object.type)),
                xule_context)

    if property_info[XuleProperties.PROP_ARG_NUM] is not None:
        property_args = function_ref['functionArgs'][1:]
        if property_info[XuleProperties.PROP_ARG_NUM] >= 0 and len(property_args) != property_info[
            XuleProperties.PROP_ARG_NUM]:
            raise XuleProcessingError(
                _("Property '%s' must have %s arguments. Found %i." % (function_ref['functionName'],
                                                                       property_info[XuleProperties.PROP_ARG_NUM],
                                                                       len(property_args))),
                xule_context)
        elif len(property_args) > property_info[XuleProperties.PROP_ARG_NUM] * -1 and property_info[
            XuleProperties.PROP_ARG_NUM] < 0:
            raise XuleProcessingError(
                _("Property '%s' must have no more than %s arguments. Found %i." % (function_ref['functionName'],
                                                                                    property_info[
                                                                                        XuleProperties.PROP_ARG_NUM] * -1,
                                                                                    len(property_args))),
                xule_context)
    # prepare the arguments
    arg_values = []
    for arg_expr in property_args:
        arg_value = evaluate(arg_expr, xule_context)
        arg_values.append(arg_value)

    if len(property_info) > 4:
        return property_info[XuleProperties.PROP_FUNCTION](xule_context, property_object,
                                                           property_info[XuleProperties.PROP_DATA], *arg_values)
    else:
        return property_info[XuleProperties.PROP_FUNCTION](xule_context, property_object, *arg_values)


def regular_function(xule_context, function_ref, function_info):
    if function_info[FUNCTION_ARG_NUM] is not None:
        # if the number of argument is none, then the function can have any number of arguments
        if function_info[FUNCTION_ARG_NUM] >= 0:
            if function_info[FUNCTION_TYPE] == 'regular' and len(function_ref['functionArgs']) != function_info[
                FUNCTION_ARG_NUM]:
                raise XuleProcessingError(
                    _("The '%s' function must have only %i argument, found %i." % (function_ref['functionName'],
                                                                                   function_info[FUNCTION_ARG_NUM],
                                                                                   len(function_ref['functionArgs']))),
                    xule_context)
        else:
            # The function can have no more than the specified number of arguments.
            if function_info[FUNCTION_TYPE] == 'regular' and len(function_ref['functionArgs']) > (
                    function_info[FUNCTION_ARG_NUM] * -1):
                raise XuleProcessingError(
                    _("The '%s' function must have no more than %i arguments, found %i." % (function_ref['functionName'],
                                                                                            function_info[
                                                                                                FUNCTION_ARG_NUM] * -1,
                                                                                            len(function_ref[
                                                                                                    'functionArgs']))),
                    xule_context)

    function_args = []
    for function_arg in function_ref['functionArgs']:
        if len(function_info) > FUNCTION_POST_EVALUATE_ARGS and function_info[FUNCTION_POST_EVALUATE_ARGS]:
            # the function arguments will be evaluated by the function instead of pre-evaluated here
            function_args.append(function_arg)
        else:
            if function_info[FUNCTION_ALLOW_UNBOUND_ARGS]:
                try:
                    arg = evaluate(function_arg, xule_context)
                except XuleIterationStop as xis:
                    arg = xis.stop_value
            else:
                arg = evaluate(function_arg, xule_context)

            function_args.append(arg)

    if len(function_info) > FUNCTION_POST_EVALUATE_ARGS and function_info[FUNCTION_POST_EVALUATE_ARGS]:
        return function_info[FUNCTION_EVALUATOR](xule_context, *function_args, evaluate_function=evaluate, iteration_stop=XuleIterationStop)
    else:
        return function_info[FUNCTION_EVALUATOR](xule_context, *function_args)


def user_defined_function(xule_context, function_ref):
    # check fucntion cache - The function cache is very basic. It only caches on functions that have no args.
    if len(function_ref['functionArgs']) == 0 and function_ref[
        'functionName'] in xule_context.global_context.function_cache:
        return xule_context.global_context.function_cache[function_ref['functionName']]

    function_info = xule_context.find_function(function_ref['functionName'])
    if function_info is None:
        raise XuleProcessingError("Function '%s' not found" % function_ref['functionName'], xule_context)
    else:
#        # Get the list of variables and their values. This will put the current single value for the variable as an argument
#        for var_ref in sorted(function_ref['var_refs'], key=lambda x: x[1]):
#            '''NOT SURE THIS IS NEEDED. THE ARGUMENTS WILL BE EXVALUTED WHEN THE for arg in matched_args IS PROCESSED'''
#            # 0 = var declaration id, 1 = var name, 2 = var_ref, 3 = var type (1 = var/arg, 2 = constant)
#            var_value = evaluate(var_ref[2], xule_context)

        matched_args = match_function_arguments(function_ref, function_info['function_declaration'], xule_context)
        for arg in matched_args:
            try:
                arg_value = evaluate(arg['expr'], xule_context)
            except XuleIterationStop as xis:
                arg_value = xis.stop_value
            xule_context.add_arg(arg['name'],
                                 arg['node_id'],
                                 arg['tagged'],
                                 arg_value,
                                 'single')

            # add the node_id of the function reference to the prefix used for calculating the processing node id
        # This is done before adding the args so the id prefix is set with the function delcaration id before the args are added as varaiables.
        xule_context.id_prefix.append(function_ref['node_id'])

        body_expr = function_info['function_declaration']['body']
        save_aligned_result_only = xule_context.aligned_result_only

        def cleanup_function():
            # remove the args
            for arg in matched_args:
                xule_context.del_arg(arg['name'],
                                     arg['node_id'])

            # pop the function reference node id off the prefix
            xule_context.id_prefix.pop()
            # reset the aligned only results.
            xule_context.aligned_result_only = save_aligned_result_only
            # xule_context.used_expressions = save_used_expressions

        function_result_values = isolated_evaluation(xule_context,
                                                     function_info['function_declaration']['node_id'],
                                                     body_expr,
                                                     cleanup_function=cleanup_function  # ,
                                                     # iteration_reset_function=iteration_reset,
                                                     )
        if 'is_iterable' in function_ref:
            function_results = function_result_values
        else:
            if None in function_result_values.values:
                function_results = function_result_values.values[None][0]
            else:
                function_results = XuleValue(xule_context, None, 'unbound')

        # Cache fucntion results that don't have any arguments.
        if len(function_ref['functionArgs']) == 0:
            xule_context.global_context.function_cache[function_ref['functionName']] = function_results
        #             if function_ref.get('cacheable') == True:
        #                 xule_context.global_context.function_cache[cache_key] = function_results
        #
        return function_results


def isolated_evaluation(xule_context, node_id, expr, setup_function=None, cleanup_function=None,
                        iteration_reset_function=None):
    save_aligned_result_only = xule_context.aligned_result_only
    save_used_expressions = xule_context.used_expressions
    # pre_aggregation_table_list_size = len(xule_context.iteration_table)
    isolated_table = xule_context.iteration_table.add_table(node_id, xule_context.get_processing_id(node_id),
                                                            is_aggregation=True)
    try:
        if setup_function is not None:
            setup_function()

        return_values = XuleValueSet()

        while True:
            xule_context.aligned_result_only = False
            xule_context.used_expressions = set()
            try:
                return_value = evaluate(expr,
                                        xule_context)  # , glob_cache_key=body_expr['node_id'] if len(matched_args)==0 else None)
            except XuleIterationStop as xis:
                return_value = xis.stop_value  # XuleValue(xule_context, None, 'unbound')

            return_value.facts = xule_context.facts.copy()
            return_value.tags = xule_context.tags.copy()
            return_value.aligned_result_only = xule_context.aligned_result_only
            return_value.used_expressions = xule_context.used_expressions
            return_value.alignment = return_value.alignment or xule_context.iteration_table.current_table.current_alignment
            # return_value.alignment = isolated_table.current_alignment
            return_values.append(return_value)

            # xule_context.iteration_table.del_current()
            xule_context.iteration_table.next(isolated_table.table_id)
            if iteration_reset_function is not None:
                iteration_reset_function()
            # if len(xule_context.iteration_table) == pre_aggregation_table_list_size:
            if isolated_table.is_empty:
                break
    finally:
        # ensure that the isolated table is removed if there is an exception.
        xule_context.iteration_table.del_table(isolated_table.table_id)
        # reset the aligned only results.
        xule_context.aligned_result_only = save_aligned_result_only
        xule_context.used_expressions = save_used_expressions
        if cleanup_function is not None:
            cleanup_function()

    return return_values

def evaluate_aggregate_function(function_ref, function_info, xule_context):
    '''Aggregation functions

    Aggregation functions perform 2 types of aggregation. The first is to collapse iterations generated from evaluating
    the arguments of the aggregation fucntion. This is essentially the opposite of a for loop. The second is to combine
    the values generated from each argument.
    '''

    # Evaluate each argument
    values_by_argument = list()
    for function_arg in function_ref['functionArgs']:
        values_by_argument.append(isolated_evaluation(xule_context, function_ref['node_id'], function_arg))

    # Combine the value sets created from each argument
    # Get all alignments
    all_alignments = set()
    for arg_value_set in values_by_argument:
        for alignment in arg_value_set:
            all_alignments.add(alignment)

    # Go through each alignment and pull the values from each of the arguments
    values_by_alignment = collections.defaultdict(list)
    # The aligned_result_only and used_expressions need to be aggregated
    aligned_result_only_by_alignment = collections.defaultdict(lambda: False)
    used_expressions_by_alignment = collections.defaultdict(set)

    for alignment in all_alignments:
        values_by_alignment[alignment] = list()
        for arg_value_set in values_by_argument:
            if alignment in arg_value_set:
                arg_alignment = alignment
            else:
                # This will match none aligned values to aligned values (i.e. 1 and @Assets)
                arg_alignment = None

            for arg_value in arg_value_set.values[arg_alignment]:
                if arg_value.type != 'unbound':
                    values_by_alignment[alignment].append(arg_value)
                aligned_result_only_by_alignment[alignment] = aligned_result_only_by_alignment[alignment] or arg_value.aligned_result_only
                used_expressions_by_alignment[alignment].update((arg_value.used_expressions))

    # Go through each alignment and apply the aggregation function
    agg_values = XuleValueSet()

    # add default value if there are no None aligned results and the aggregation has a default value.
    # if None not in values_by_alignment and function_info[FUNCTION_DEFAULT_VALUE] is not None:
    if len(values_by_alignment) == 0:
        agg_values.append(XuleValue(xule_context, function_info[FUNCTION_DEFAULT_VALUE], function_info[FUNCTION_DEFAULT_TYPE]))

    for alignment in values_by_alignment:
        if len(values_by_alignment[alignment]) > 0:
            agg_value = function_info[FUNCTION_EVALUATOR](xule_context, values_by_alignment[alignment])
        else:
            # Add the default value if there is one
            if function_info[FUNCTION_DEFAULT_VALUE] is not None:
                agg_value = XuleValue(xule_context, function_info[FUNCTION_DEFAULT_VALUE],
                                      function_info[FUNCTION_DEFAULT_TYPE])
            else:
                agg_value = None

        if agg_value is not None:
            #if (aligned_result_only_by_alignment[alignment] and 
            #    alignment is None and
            #    xule_context.iteration_table.current_table.current_alignment is None):
            #    agg_value = XuleValue(xule_context, None, 'unbound')

            #if (aligned_result_only_by_alignment[alignment] and 
            #    alignment is None ):
            #    agg_value = XuleValue(xule_context, None, 'unbound')

            agg_value.alignment = alignment
            agg_value.aligned_result_only = aligned_result_only_by_alignment[alignment]
            # print("agg", function_ref['exprName'], function_ref['node_id'], len(xule_context.used_expressions), len(used_expressions))
            agg_value.used_expressions = used_expressions_by_alignment[alignment]
            agg_values.append(agg_value)
        
    return agg_values


def evaluate_property(property_expr, xule_context):
    # The object_value is the left side of the property expression
    object_value = evaluate(property_expr['expr'], xule_context)

    # The properties expression is an object and then a chain of properties (i.e. Assets[]::concept::name::local-part)
    for current_property_expr in property_expr['properties']:
        # Check that this is a valid property
        if current_property_expr['propertyName'] not in XuleProperties.PROPERTIES:
            raise XuleProcessingError(_("'%s' is not a valid property." % current_property_expr['propertyName']),
                                      xule_context)

        property_info = XuleProperties.PROPERTIES[current_property_expr['propertyName']]

        # Check if the property can operate on a set or list.
        if (object_value.type not in ('set', 'list') or # It's not a collection so the property should not operate on a collection
            (object_value.type in ('set', 'list') and (
            len({'set', 'list'} & set(property_info[XuleProperties.PROP_OPERAND_TYPES])) > 0) or # It's a collection but the property is for a collection
            (object_value.is_fact and 'fact' in property_info[XuleProperties.PROP_OPERAND_TYPES])) or # It's a fact and the property works for a fact
            (
            # property operates on all types, so should operate on the list or set
            len(property_info[XuleProperties.PROP_OPERAND_TYPES]) == 0)
            ):
            object_value = process_property(current_property_expr, object_value, property_info, xule_context)
        else:
            # This is a set or list. The property is not for a set or list, so try to create a new set or list after applying the property to the members.
            if object_value.type == 'set':
                new_list = set()
                new_shadow = set()
            else:
                new_list = list()
                new_shadow = list()
            for item in object_value.value:
                new_value = process_property(current_property_expr, item, property_info, xule_context)
                if object_value.type == 'set':
                    if (new_value.shadow_collection if new_value.type in (
                    'set', 'list', 'dictionary') else new_value.value) not in new_shadow:
                        new_list.add(new_value)
                        new_shadow.add(new_value.shadow_collection if new_value.type in (
                        'set', 'list', 'dictionary') else new_value.value)
                else:  # list
                    new_list.append(new_value)
                    new_shadow.append(new_value.shadow_collection if new_value.type in (
                    'set', 'list', 'dictionary') else new_value.value)

            if object_value.type == 'set':
                object_value = XuleValue(xule_context, frozenset(new_list), 'set',
                                         shadow_collection=frozenset(new_shadow))
            else:  # list
                object_value = XuleValue(xule_context, tuple(new_list), 'list', shadow_collection=frozenset(new_shadow))

    return object_value


def process_property(current_property_expr, object_value, property_info, xule_context):
    # Check that the left object is the right type
    # if the left object is unbound then return unbound
    if not property_info[XuleProperties.PROP_UNBOUND_ALLOWED] and object_value.type in ('unbound', 'none'):
        return object_value
    else:
        # check the left object is the right type
        if len(property_info[XuleProperties.PROP_OPERAND_TYPES]) > 0:
            if not (object_value.type in property_info[XuleProperties.PROP_OPERAND_TYPES] or
                    object_value.is_fact and 'fact' in property_info[XuleProperties.PROP_OPERAND_TYPES] or
                    any([xule_castable(object_value, allowable_type, xule_context) for allowable_type in
                         property_info[XuleProperties.PROP_OPERAND_TYPES]]) or
                    'noncollection' in property_info[XuleProperties.PROP_OPERAND_TYPES] or
                    (object_value.type in ('none', 'unbound') and property_info[XuleProperties.PROP_UNBOUND_ALLOWED])):
       
                raise XuleProcessingError(
                    _("Property '%s' is not a property of a '%s'.") % (current_property_expr['propertyName'],
                                                                       object_value.type),
                    xule_context)
            
    if property_info[XuleProperties.PROP_ARG_NUM] is not None:
        property_args = current_property_expr.get('propertyArgs', [])
        if property_info[XuleProperties.PROP_ARG_NUM] >= 0 and len(property_args) != property_info[
            XuleProperties.PROP_ARG_NUM]:
            raise XuleProcessingError(
                _("Property '%s' must have %s arguments. Found %i." % (current_property_expr['propertyName'],
                                                                       property_info[XuleProperties.PROP_ARG_NUM],
                                                                       len(property_args))),
                xule_context)
        elif len(property_args) > property_info[XuleProperties.PROP_ARG_NUM] * -1 and property_info[
            XuleProperties.PROP_ARG_NUM] < 0:
            raise XuleProcessingError(_(
                "Property '%s' must have no more than %s arguments. Found %i." % (current_property_expr['propertyName'],
                                                                                  property_info[
                                                                                      XuleProperties.PROP_ARG_NUM] * -1,
                                                                                  len(property_args))),
                                      xule_context)
    # prepare the arguments
    arg_values = []
    for arg_expr in property_args:
        arg_value = evaluate(arg_expr, xule_context)
        arg_values.append(arg_value)

    if len(property_info) > 4:  # There is property data
        object_value = property_info[XuleProperties.PROP_FUNCTION](xule_context, object_value,
                                                                   property_info[XuleProperties.PROP_DATA], *arg_values)
    else:
        object_value = property_info[XuleProperties.PROP_FUNCTION](xule_context, object_value, *arg_values)

    if 'tagName' in current_property_expr:
        xule_context.tags[current_property_expr['tagName']] = object_value

    return object_value

def evaluate_index(index_expr, xule_context):
    # evaluate the left side of the expression
    left_value = evaluate(index_expr['expr'], xule_context)

    for index in index_expr['indexes']:
        index_value = evaluate(index, xule_context)
        # An index expression is used for lists and dictionaries.
        left_value = XuleProperties.property_index(xule_context, left_value, index_value)

    return left_value

def evaluate_tag_ref(tag_ref, xule_context):
    if tag_ref['varName'] in xule_context.tags:
        # When tags are evaluated for message production, the tags associated with the tagged value
        # should overwrite the the current tags. Making a copy of the value and removing the tags (and likewise for facts)
        tag_value =  copy.copy(xule_context.tags[tag_ref['varName']])
        tag_value.tags = dict()
        tag_value.facts = collections.OrderedDict()
        return tag_value
    else:
        # The reference may be to a constant
        cat_const = xule_context.global_context.catalog['constants'].get(tag_ref['varName'])
        if cat_const is not None:
            ast_const = xule_context.global_context.rule_set.getItem(cat_const)
            # If the constant is iterable and it was never used in the rule body, it cannot be calculated for the message. 
            # There would be no way to determing which value to use.
            if ast_const['number'] == 'single':
                const_info = xule_context.find_var(tag_ref['varName'], ast_const['node_id'])
                if const_info['type'] == xule_context._VAR_TYPE_CONSTANT:
                    if not const_info.get('calculated'):
                        var_value = evaluate(ast_const, xule_context)

                    return const_info['value'].values[None][0]
    # If here the tag could not be found
    return XuleValue(xule_context, None, 'none')


# aspect info indexes
TYPE = 0
ASPECT = 1
SPECIAL_VALUE = 2
ASPECT_OPERATOR = 3
ASPECT_PROPERTY = 4

EVALUATOR = {
    # rules
    "assertion": evaluate_assertion,
    "outputRule": evaluate_output_rule,

    # literals
    "boolean": evaluate_bool_literal,
    "integer": evaluate_int_literal,
    "float": evaluate_float_literal,
    "string": evaluate_string_literal,
    "period": evaluate_period_literal,
    "qname": evaluate_qname_literal,
    "groupQname": evaluate_namespace_group,
    # "skip": evaluate_void_literal,
    "none": evaluate_void_literal,

    # atomic expressions
    "constantDeclaration": evaluate_constant_assign,
    "ifExpr": evaluate_if,
    "forExpr": evaluate_for,
    # "forControl": evaluate_for_control,
    # "forLoop": evaluate_for_loop,
    # "forBodyExpr": evaluate_for_body,
    # "withExpr": evaluate_with,

    "blockExpr": evaluate_block,
    "varRef": evaluate_var_ref,
    "tagRef": evaluate_tag_ref,
    "functionReference": evaluate_function_ref,
    "taggedExpr": evaluate_tagged,
    "propertyExpr": evaluate_property,
    "indexExpr": evaluate_index,

    "factset": evaluate_factset,
    "navigation": evaluate_navigate,
    "filter": evaluate_filter,
    #"dictExpr": evaluate_dict,
    #"listExpr": evaluate_list,
    #"setExpr": evaluate_set,

    # expressions with order of operations
    "unaryExpr": evaluate_unary,
    "inExpr": evaluate_in,
    "multExpr": evaluate_mult,
    "addExpr": evaluate_add,
    "intersectExpr": evaluate_intersect,
    "symetricDifferenceExpr": evaluate_symetric_difference,
    "compExpr": evaluate_comp,
    "notExpr": evaluate_not,
    "andExpr": evaluate_and,
    "orExpr": evaluate_or,

    # severity
    'severity': evaluate_severity,

    # aspect name literal
    'aspectName': evaluate_aspect_name,

    # balance
    'balance': evaluate_string_keyword,
    'periodType': evaluate_string_keyword,

}

# the position of the function information
FUNCTION_TYPE = 0
FUNCTION_EVALUATOR = 1
FUNCTION_ARG_NUM = 2
# aggregate only
FUNCTION_DEFAULT_VALUE = 3
FUNCTION_DEFAULT_TYPE = 4
# non aggregate only
FUNCTION_ALLOW_UNBOUND_ARGS = 3
FUNCTION_RESULT_NUMBER = 4
FUNCTION_POST_EVALUATE_ARGS = 5


def built_in_functions():
    return XuleFunctions.BUILTIN_FUNCTIONS


BUILTIN_FUNCTIONS = XuleFunctions.BUILTIN_FUNCTIONS


def process_factset_aspects(factset, xule_context):
    """Build list of aspects as descried in the factset
    
    This function returns 3 dictionaries. 
    1. A dictionary of aspect filters that will not be included in the alignment. These are indicated with @ (single) in the factset expression.
    2. A dictionary of aspect filters that will be included in the alginment. These are indicated with @@ (double) in the factset expression.
    3. A dictionary of variables based on aspect aliases.
    
    The aspect filters dictionaries (the first 2 dictionaries) are keyed by a tuple of 4 items:    
    1. aspect type - This is either 'builtin' or 'explicit_dimension'
    2. aspect name - For built in aspects it is one of 'concept', 'entity', 'period', 'unit' or 'cube'
                     For dimensional aspects it is the qname of the dimension
    3. special value - If the aspect uses a wildcard, this will contain the wildcard character '*'. Otherwise it is None.
    4. aspect operator - This is the operator used for the filter. It is one of '=', '!=', 'in' or 'not in'. If there is not aspect operator
                         the value is None.
    5. aspect properties - This is a tuple. The first time is the property name and second item is a list of evaluated tuple of arguments.
    
    The value of the key is the evaluated expression on the right of the aspect operator. If there is no operator, then the value is None.                     
    
    The dictionary of aspect variables is key by a the alias name. The value is a tuple of 3 items:
    1. aspect type - This is either 'builtin' or 'explicit_dimension'
    2. aspect name - For built in aspects it is one of 'concept', 'entity', 'period', 'unit' or 'cube'
                     For dimensional aspects it is the qname of the dimension
    3. aspect index - This is the node_id of the aspect filter.

    """
    non_align_aspects = {}
    align_aspects = {}
    aspect_vars = {}
    alternate_notation = False

    # Set the model to the default model
    models = (xule_context.model,)
    # evaluate all the aspect names. If there is an @instance aspect then need to change the model
    aspects = dict()
    # These flags identify when the @dimensions=none or @dimensions=*
    dimensions_special_value = None
    dimensions_special_covered = None
    for aspect_filter in factset.get('aspectFilters', list()):
        aspect_name = evaluate(aspect_filter['aspectName'], xule_context)
        # Check if @instance. If so, change the model
        # Also, preprocess the @dimesnions
        # The @dimensions is a dictionary of key = diemsnion name and value = the member. Prepopulate the aspect_dictionary. If there
        # is another aspect filter for something in the @dimensions dictionary, then it will overrite that value in the aspect_dictionary
        if aspect_name.type == 'aspect_name' and aspect_name.value == 'instance':
            instance = evaluate(aspect_filter['aspectExpr'], xule_context)
            if aspect_filter['aspectOperator'] == '=':
                if instance.type == 'instance':
                    models = (instance.value,)
                else:
                    raise XuleProcessingError(_("Value of the @instance aspect of a factset must be an instance, but found {}".format(instance.type)), xule_context)
            elif aspect_filter['aspectOperator'] == 'in':
                if instance.type not in ('set', 'list'):
                    raise XuleProcessingError(_("The value of the @instance aspect with 'in' must be a set or a list, found {}".format(instance.type)), xule_context)
                models = list()
                for sub_instance in instance.value:
                    if sub_instance.type == 'instance':
                        models.append(sub_instance.value)
                    else:
                        raise XuleProcessingError(_("Value of the @instance aspect of a factset must be an instance, found {}".format(sub_instance.type)), xule_context)
        elif aspect_name.type == 'qname' and aspect_name.value.localName == 'dimensions' and aspect_name.value.prefix is None:
            # We are going to prepopulate the aspect_dictionary (either aligned or non_aligned.
            # Need to determine is this is aligned or non aligned
            aspect_dictionary = non_align_aspects if aspect_filter['coverType'] == 'covered' else align_aspects

            if aspect_filter.get('wildcard') == '*':
                # This is @dimensions=*
                # There is really nothing to do here. If it is covered (has @@) then set dimensions_are_all_covered flag
                dimensions_special_value = '*'
                dimensions_special_covered = aspect_filter.get('coverType')
            else:

                dimensions_value = evaluate(aspect_filter['aspectExpr'], xule_context)
                if dimensions_value.type == 'dictionary':
                    for dimension_value, member_value in dimensions_value.value:
                        x = 1
                        # dimension_value and member_value may be concepts or qnames. The dimension_value must be qname. I will make
                        # the member_value a qname as well. This will better support filtering where the dimesnion dictionary comes from
                        # one taxonomy but the factset is for another.
                        if dimension_value.type in ('qname', 'groupQname'):
                            dimension_name = dimension_value.value
                        elif dimension_value.type == 'concept':
                            dimension_name = dimension_value.value.qname
                        else:
                            raise XuleProcessingError(_(f"The key of the @dimensions dictionary filter must be a qname or a concept, found {dimension_value.type}"), xule_context)
                        aspect_info = ('explicit_dimension', dimension_name,  None, '=', None)

                        if member_value.type in ('qname', 'groupQname'):
                            member_qname_value = member_value
                        elif member_value.type == 'concept':
                            member_qname_value = XuleValue(xule_context, member_value.value.qname, 'qname')
                        else:
                            raise XuleProcessingError(_(f"The value of the @dimensions dictionary filter must be a qname or a concept, found {member_value.type}"), xule_context)
                        aspect_dictionary[aspect_info] = member_qname_value
                elif dimensions_value.type == 'none':
                    dimensions_special_value = 'none'
                    dimensions_special_covered = aspect_filter.get('coverType')
                else:
                    raise XuleProcessingError(_(f"The value of the @dimensions filter must be a dictionary, found {aspect_dictionary.type}"), xule_context)
        else:
            aspects[aspect_name] = aspect_filter
    '''COULD CHECK FOR DUPLICATE ASPECTS IN THE FACTSET'''
    for aspect_name, aspect_filter in aspects.items():
        # Set the dictionary to use based on if the aspect is covered (single @ - non aligned) vs. uncoverted (double @@ - aligned)
        aspect_dictionary = non_align_aspects if aspect_filter['coverType'] == 'covered' else align_aspects
        aspect_var_name = aspect_filter.get('alias')

        # evaluate the aspectName
        #aspect_name = evaluate(aspect_filter['aspectName'], xule_context)

        if aspect_name.type == 'aspect_name':
            # This is a built in aspect - one of concept, period, entity, unit or table
            add_aspect_var(aspect_vars, 'builtin', aspect_name.value, aspect_var_name,
                           aspect_filter['node_id'], xule_context)
            if aspect_name.value == 'concept' and alternate_notation:
                raise XuleProcessingError(_(
                    "The factset specifies the concept aspect as both @{0} and @concept={0}. Only one method should be used".format(
                        aspect_name.value)), xule_context)
            aspect_info, aspect_value = process_aspect_expr(aspect_filter, 'builtin', aspect_name.value, models, xule_context)
            if aspect_info is not None:
                aspect_dictionary[aspect_info] = aspect_value
        elif aspect_name.type in ('qname', 'groupQname'):
            # This is either a dimension aspect or the default concept aspect. The aspect name is determined by evaluating the aspectDimensionName
            # Get the model concept to determine if the aspect is a dimension.
            # Models is a list or tuple of models. Find the the concept in the first model where it exists
            aspect_filter_model_concept = None
            for model in models:
                if aspect_name.type == 'qname':
                    aspect_filter_model_concept = model.qnameConcepts.get(aspect_name.value)
                    if aspect_filter_model_concept is not None:
                        break
                else: # This is a groupQname
                    # the whole purpose of getting the model concept for the aspect_name is to see if it is a dimension. If it is 
                    # a dimension, then this is a dimension aspect. If it is not a dimension, then this is a shorthand notation for 
                    # the concept aspect (i.e @usgaap:Assets instead of @concept=usgaap:Assets). If there are multiple
                    # concepts that match, then they all have to be dimensions or all not dimensions, that is it can't 
                    # be a mix of some dimension concepts and some not. If there is a mix, then is will be an error.
                    # The groupQname xule value is a tuple of a list of namespace uri fragments and the local name.
                    for model_concept in model.nameConcepts.get(aspect_name.value[1], ()): # get all the concepts by local name
                        if match_namespace_group(aspect_name.value[0], model_concept.qname):
                            if aspect_filter_model_concept is None:
                                aspect_filter_model_concept = model_concept
                            else:
                                # verify that the first found model_concept matches the isDimension
                                if model_concept.isDimensionItem != aspect_filter_model_concept.isDimensionItem:
                                    raise XuleProcessingError(_("Found facset aspect that uses a namespace group prefix in the aspect name, but there are multiple concepts that are not all dimensions or all non dimensions. This is not allowed"), xule_context)

            if aspect_filter_model_concept is None:
                if aspect_name.type == 'groupQname':
                    missing_concept_name = f"{aspect_name.value[0].value}:{aspect_name.value[1]}"
                else:
                    missing_concept_name = aspect_name.value.clarkNotation
                raise XuleProcessingError(
                    _("Error while processing factset aspect. Concept %s not found." % missing_concept_name),
                    xule_context)
            if aspect_filter_model_concept.isDimensionItem:
                # This is a dimension aspect
                add_aspect_var(aspect_vars, 'explicit_dimension', aspect_name.value, aspect_var_name,
                               aspect_filter['node_id'], xule_context)
                aspect_info, aspect_value = process_aspect_expr(aspect_filter, 'explicit_dimension', aspect_name.value,
                                                               models, xule_context)
                # Need to clear the filter if it is already in the aspect_dictionary. This can happen if the @dimensions is used and an @xbrl-dimension=member for the same dimension.
                remove_existing_aspects(aspect_info[1], non_align_aspects, align_aspects)
                if aspect_info is not None:
                    aspect_dictionary[aspect_info] = aspect_value
            else:
                # This is a concept aspect and the filter name is really the aspect value (i.e. @Assets)
                if aspect_in_filters('builtin', 'concept', aspect_dictionary):
                    raise XuleProcessingError(_(
                        "The factset specifies the concept aspect as both @{0} and @concept={0}. Only one method should be used. Is dimension={1}. Sub Group={2}".format(
                            aspect_name.value, aspect_filter_model_concept.isDimensionItem, aspect_filter_model_concept.substitutionGroup)), xule_context)
                alternate_notation = True  # Indicate that the concept aspect is provided
                aspect_dictionary[('builtin', 'concept', None, '=', None)] = aspect_name
                add_aspect_var(aspect_vars, 'builtin', 'concept', aspect_var_name, aspect_filter['node_id'],
                               xule_context)
        else:
            raise XuleProcessingError(_(
                "An aspect name must be one of 'concept', 'unit', 'period', 'entity' or a dimension qname, found '{}'.".format(
                    aspect_name.type)), xule_context)

    return (non_align_aspects, align_aspects, aspect_vars, models, dimensions_special_value, dimensions_special_covered)

def remove_existing_aspects(dim_name, non_align_aspects, align_aspect):
    '''This is used to remove an aspect if it is in the aspect filters more than once.
       This can happen if the @dimensions is used and there is an aspect filter for one of the dimension in the @dimensions.
       In this case, the specified dimension (i.e. @dimension-name=member) is used in place of the the dimension in the @dimensions'''
    existing_keys = {x for x in non_align_aspects.keys() if x[1] == dim_name}
    for existing_key in existing_keys:
        del non_align_aspects[existing_key]
    existing_keys = {x for x in align_aspect.keys() if x[1] == dim_name}
    for existing_key in existing_keys:
        del align_aspect[existing_key]

def aspect_in_filters(aspect_type, aspect_name, filters):
    """Checks if an aspect is in the existing set of filters
    
    Arguments:
    aspect_type (string): Either 'builtin' or 'dimension'
    aspect_name (string or qname): if 'builtin' this will be a string with values of 'concept', 'unit', 'period', 'entity' or 'cube', otherwise it is a qname
                                   of the dimensional aspect name.
    filters (dictionary): Dictionary of aspect filters.
    """
    for filter_type, filter_name, _x, _y, _z in filters:
        if filter_type == aspect_type and filter_name == aspect_name:
            return True
    # This will only hit if the aspect was not found in the filters.
    return False

def process_aspect_expr(aspect_filter, aspect_type, aspect_name, models, xule_context):
    """Process the expression on the right side of an aspect filter.
    
    This looks at the aspectExpr for the aspect filter (if there is on). This will return a 2 item tuple of the aspect_info key and 
    the evaluated aspect expression value.
    
    Arguments:
        aspect_filter (dict): AST node for the aspect filter
        aspect_type (string): Either 'builtin' or 'explicit_dimension'
        xule_context (XuleContext): The processing context
    """
    aspect_info = None
    aspect_value = None
    prop = None

    # Properties 
    if 'propertyName' in aspect_filter:
        prop = (aspect_filter['propertyName'],
                tuple(evaluate(arg, xule_context).value for arg in aspect_filter.get('propertyArgs', tuple())))

    if 'wildcard' in aspect_filter:
        if aspect_filter['aspectOperator'] not in ('=', '!='):
            raise XuleProcessingError(_("In a factset a '*' can only be used with '=' or '!=', found '{}'".format(
                aspect_filter['aspectOperator'] + ' *')), xule_context)
        if aspect_filter['aspectOperator'] == '=':
            aspect_info = (aspect_type, aspect_name, aspect_filter['wildcard'], aspect_filter['aspectOperator'], prop)
            aspect_value = XuleValue(xule_context, None, 'none')
        else:
            # This is aspect_name != *. Really this is the fact does not have this aspect (this aspect is in the default)
            aspect_info = (aspect_type, aspect_name, None, '=', prop)
            aspect_value = XuleValue(xule_context, None, 'none')
    else:
        # Not a wildcard
        aspect_info = (aspect_type, aspect_name, None, aspect_filter.get('aspectOperator'), prop)
        if 'aspectExpr' in aspect_filter:
            aspect_value = evaluate(aspect_filter['aspectExpr'], xule_context)
            if aspect_type == 'explicit_dimension':
                aspect_value = fix_for_default_member(aspect_name, aspect_value, models, xule_context)
        else:
            aspect_value = None # There is nothing to filter, but the aspect info will be used for handling alignment

    return (aspect_info, aspect_value)

def fix_for_default_member(dim, aspect_value, models, xule_context):
    ''' If the member for an explicit dimension is the default member, change the value to none'''
    # There may be multiple models if @instance is used with a list of instances. So Check each model.
    # If there are are different defaults, this is a semantic problem and raise an error.

    default_names = set()
    for model in models:
        next_name = XuleDimensionCube.dimension_defaults_by_name(model).get(dim)
        if next_name is not None:
            default_names.add(next_name)
    if len(default_names) == 1:
        default_name = default_names.pop() # this just get the item out of the set
    elif len(default_names) ==0:
        default_name = None # there is no default for the dimension
    else: # there was more than 1 default for the dimension. This is a problem
        raise XuleProcessingError(_(f"Found different dimension defaults for {dim.clarkNotation}. This happens when there is a factset "
                                    f"with an @instance with multiple instances and the defaults are different in the different instances. "
                                    f"The default memebers are: {', '.join(default_names)}"))

    if default_name is None:
        return aspect_value
    new_values = list()
    for mem in aspect_value.value if aspect_value.type in ('list', 'set') else (aspect_value,):
        if mem.type == 'concept':
            mem_qname = mem.value.qname
        elif mem.type == 'qname':
            mem_qname = mem.value
        else:
            new_values.append(mem)
            continue

        if mem_qname == default_name:
            new_values.append(XuleValue(xule_context, None, 'none'))
        else:
            new_values.append(mem)

    if aspect_value.type == 'set':
        return XuleValue(xule_context, frozenset(new_values), 'set')
    elif aspect_value.type == 'list':
        return XuleValue(xule_context, tuple(new_values), 'list')
    else:
        return new_values[0]

def add_aspect_var(aspect_vars, aspect_type, aspect_name, var_name, aspect_index, xule_context):
    if var_name:
        if var_name in aspect_vars:
            raise XuleProcessingError(
                _("Found multiple aspects with same variable name '%s' in a factset." % (var_name)), xule_context)
        else:
            aspect_vars[var_name] = (aspect_type, aspect_name, aspect_index)

def match_namespace_group(fragment_list, namespace_uri):
    if fragment_list.type == 'string':
        fragments = (fragment_list,) 
    else:
        fragments = fragment_list.value
    for fragment in fragments:
        if fragment.value in namespace_uri.namespaceURI:
            return True

    # if here, there were no matching fragments
    return False

def convert_value_to_qname(value, model, xule_context):
    if value.type == 'concept':
        return {value.value.qname,}
    elif value.type == 'qname':
        return {value.value,}
    elif value.type == 'groupQname':
        return_values = set()
        for model_concept in model.nameConcepts.get(value.value[1], tuple()):
            if match_namespace_group(value.value[0], model_concept.qname):
                return_values.add(model_concept.qname)
        return return_values
    elif value.type in ('unbound', 'none'):
        return {None}
    # HF addition for testing rule 0118, not sure why set of one qname is showing up here
    elif value.type in ('list','set') and len(value.value) == 1 and next(iter(value.value)).type == "qname":
        return {next(iter(value.value)).value,}
    elif value.type in ('list','set') and len(value.value) == 0:
        return {None}
    else:
        raise XuleProcessingError(
            _("The value for a line item or dimension must be a qname or concept, found '%s'." % value.type),
            xule_context)

def convert_value_to_role(value, xule_context):
    if value.type in ('string', 'uri'):
        return [value.value,]
    elif value.type == 'qname':
        return XuleUtility.resolve_role(value, 'role', xule_context.model, xule_context)
    elif value.type == 'role':
        return [value.value.roleURI,]
    elif value.type in ('unbound', 'none'):
        return [] # empty list
    else:
        raise XuleProcessingError(
            _("The value for a role or arc role must be a string, uri or short role name, found '%s'." % value.type),
            xule_context)

def convert_value_to_model_period(value, xule_context):
    if value.type in ('unbound', 'none'):
        return None

    if value.from_model:
        return value.value
    else:
        # need to adjust instant and end_date. The model has instant and end dates of the next day because python treats midnight as the begining of the day.
        if value.type == 'instant':
            return value.value + datetime.timedelta(days=1)
        elif value.type == 'duration':
            if value.value[0] == datetime.datetime.min and value.value[1] == datetime.datetime.max:
                # this is forever, don't do anything
                return value.value
            else:
                return (value.value[0], value.value[1] + datetime.timedelta(days=1))
        else:
            raise XuleProcessingError(
                _("Converting result to a period, expected 'instant' or 'duration' but found '%s'" % value.type),
                xule_context)


def convert_value_to_model_unit(value, xule_context):
    if value.type == 'unit':
        return value.value
    elif value.type == 'qname':
        # A xule 'unit' is a tuple 
        return ((value.value,), tuple())


# def get_used_vars(xule_context, var_ids):    
#     return {var_id: xule_context.vars.get(var_id) for var_id in var_ids}

def match_function_arguments(reference, declaration, xule_context):
    ''' This function matches the arguments on the function reference (call) to the function declaration.
        It returns a list of matched arguments as a dictionary.
    '''
    if len(reference['functionArgs']) != len(declaration['functionArgs']):
        raise XuleProcessingError(_("Function call for '%s' has mismatched arguments." % reference['functionName']),
                                  xule_context)
    else:
        matched = []
        for i in range(len(reference['functionArgs'])):
            arg_name = declaration['functionArgs'][i]['argName']

            matched.append({"name": arg_name,
                            "node_id": declaration['functionArgs'][i]['node_id'],
                            "expr": reference['functionArgs'][i],
                            "tagged": declaration['functionArgs'][i].get('tagged', False)})

    return matched


def remove_from_alignment(alignment, remove_aspects, xule_context):
    unfrozen_alignment = {k: v for k, v in alignment}
    for aspect_key in remove_aspects:
        if aspect_key in unfrozen_alignment:
            del unfrozen_alignment[aspect_key]
    new_alignment = frozenset(unfrozen_alignment.items())
    return new_alignment


def alignment_to_aspect_info(alignment, xule_context):
    aspect_dict = {}
    for align_key, align_value in alignment.items():
        aspect_info = (align_key[0], align_key[1], None, '=', None)

        if align_key[0] == 'builtin':
            if align_key[1] == 'concept':
                aspect_value = XuleValue(xule_context, align_value, 'qname')
            elif align_key[1] == 'unit':
                aspect_value = XuleValue(xule_context, align_value, 'unit')

            elif align_key[1] == 'period':
                period_value = align_value
                if isinstance(period_value, tuple):
                    if period_value[1] == datetime.datetime.max:
                        # this is forever
                        aspect_value = XuleValue(xule_context, period_value, 'duration', from_model=True)
                    else:
                        # need to adjust the end date
                        #                         aspect_value = XuleResultSet(XuleResult((period_value[0],
                        #                                                                  period_value[1] - datetime.timedelta(days=1))
                        #                                                                 ,'duration'))
                        aspect_value = XuleValue(xule_context, (period_value[0],
                                                                period_value[1])
                                                 , 'duration',
                                                 from_model=True)
                else:
                    # need to adjust the date. This is from the model which handles midnight (end of day) as beginning of next day.
                    #                     aspect_value = XuleResultSet(XuleResult(period_value - datetime.timedelta(days=1), 'instant'))
                    aspect_value = XuleValue(xule_context, period_value, 'instant', from_model=True)

            elif align_key[1] == 'entity':
                aspect_value = XuleValue(xule_context, align_value, 'entity')

            else:
                raise XuleProcessingError(_("Unknown built in aspect '%s'" % align_key[1]), xule_context)
        elif align_key[0] == 'explicit_dimension':
            aspect_value = XuleValue(xule_context, align_value, model_to_xule_type(xule_context, align_value)[0])

        else:
            raise XuleProcessingError(_("Unknown aspect type '%s'" % align_key[0]), xule_context)

        aspect_dict[aspect_info] = aspect_value

    return aspect_dict


def sugar_trace(value, rule_part, xule_context):
    try: # HF: exception when rule part key isn't there
        part_name = rule_part['exprName']
        if part_name == 'forExpr':
            return (rule_part['forLoopExpr']['forVar'],)
        elif part_name == 'varRef':
            return (rule_part['varName'],)
        elif part_name == 'functionReference':
            function_info = xule_context.find_function(rule_part['functionName'])
            if function_info is None:
                return (rule_part['functionName'], tuple())
            else:
                args = tuple([x.argName for x in function_info['function_declaration']['functionArgs']])
                return (rule_part['functionName'], args)
        elif part_name == 'factset':
            return (value,)
        else:
            return tuple()
    except (KeyError, AttributeError):
        return tuple()

def format_trace(xule_context):
    trace_string = ""
    for step in xule_context.trace:
        trace_string += "  " * step[0] + step[3].format_value() + format_trace_info(step[1], step[2], {},
                                                                                    xule_context) + "\n"
    return trace_string


def format_trace_info(expr_name, sugar, common_aspects, xule_context):
    trace_info = ""

    try:
        if expr_name == 'forExpr':
            trace_info += 'for ($%s)' % sugar[0]
        elif expr_name == 'ifExpr':
            trace_info += 'if'
        elif expr_name == 'varRef':
            trace_info += 'var ($%s)' % sugar[0]
        elif expr_name == 'functionReference':
            if len(sugar[1]) == 0:
                args = "..."
            else:
                args = ",".join(sugar[1])
            trace_info += '%s(%s)' % (sugar[0], args)
        elif expr_name == 'addExpr':
            trace_info += 'add/subtract'
        elif expr_name == 'multExpr':
            trace_info += 'multiply/divide'
        elif expr_name == 'compExpr':
            trace_info += 'comparison'
        elif expr_name == 'andExpr':
            trace_info += 'and'
        elif expr_name == 'orExpr':
            trace_info += 'or'
        elif expr_name == 'property':
            trace_info += "::%s" % sugar[0]
        elif expr_name == 'factset':
            if sugar[0].fact is not None:
                fact_context = get_uncommon_aspects(sugar[0].fact, common_aspects, xule_context)
                trace_info = 'factset '
                if ('builtin', 'concept') not in fact_context:
                    trace_info += str(sugar[0].qname) + " "
                trace_info += format_alignment(fact_context, xule_context)
        else:
            trace_info += expr_name
    except IndexError:
        trace_info += "missing rule construct"

    if len(trace_info) > 0:
        trace_info = " - " + trace_info

    return trace_info

def result_file(rule_ast, xule_value, xule_context):

    file_content = None
    file_location = None
    file_append = None

    for result_ast in rule_ast.get('results', tuple()):
        if result_ast.get('resultName') in ('file-content', 'file-location', 'file-append'):
            try:
                message_context = xule_context.create_message_copy(rule_ast['node_id'], xule_context.get_processing_id(rule_ast['node_id']))
                message_context.tags['rule-value'] = xule_value
                message_context.tags['alignment'] = func_alignment(xule_context)
                message_context.result_alignment = func_alignment(xule_context)
                saved_no_cache = getattr(message_context.global_context.options, 'xule_no_cache', False)
                xule_context.global_context.options.xule_no_cache = True
                result = evaluate(result_ast['resultExpr'], message_context)
                if result_ast.get('resultName') == 'file-content':
                    file_content = result
                elif result_ast.get('resultName') == 'file-location':
                    file_location = result
                else: # This is file-append
                    file_append = result
            finally:
                xule_context.global_context.options.xule_no_cache = saved_no_cache

    if file_content is None or file_location is None:
        return # there is nothing to do. 

    if file_location.type != 'string':
        raise XuleProcessingError(_(f"The file-location needs to be a string, but it is a {file_location.type}"), xule_context)

    if file_content.type not in ('none', 'string', 'spreadsheet'):
        raise XuleProcessingError(_(f"Cannot write contents of type {file_content.type}"), xule_context)

    if file_append is not None and file_append.type != 'bool':
        raise XuleProcessingError(_(f"file-append must be a boolean, found {file_append.type}"))

    # Write the file
    if file_location.value in xule_context.global_context.output_files or (file_append is not None and file_append.value):
        open_mode = 'a'
    else:
        open_mode = 'w'
        xule_context.global_context.output_files.add(file_location.value)

    if file_content.type == 'spreadsheet':
        write_excel(file_location, file_content, open_mode, xule_context)
    else:
        try:
            with open(file_location.value, open_mode, encoding='utf8') as ofile:
                ofile.write(file_content.value if file_content.type == 'string' else '') # if none just write a blank string
        except FileNotFoundError:
            raise XuleProcessingError(_(f"Cannot open output file {file_location.value}"), xule_context)
    
def write_excel(file_location, file_content, open_mode, xule_context):
    if open_mode == 'a' and os.path.exists(file_location.value):
        wb = load_workbook(file_location.value)
    else:
        wb = Workbook()
        # the workbook will have a default sheet named 'Sheet'. We ware removing it so the workbook is empty.
        wb.remove(wb.get_sheet_by_name('Sheet'))

    data = file_content.value
    for sheet_name in sorted(data.keys()):
        content = data[sheet_name]
        if sheet_name in wb.get_sheet_names():
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
        
        for row in content:
            ws.append(row)

    # Save the workbook
    if len(wb.get_sheet_names()) == 0:
        # There has to be at least 1 sheet in order to save the workbook
        wb.create_sheet('Sheet')
    wb.save(file_location.value)


def result_message(rule_ast, result_ast, xule_value, xule_context):
    # validate_result_name(result_ast, xule_context)
    message_context = xule_context.create_message_copy(rule_ast['node_id'], xule_context.get_processing_id(rule_ast['node_id']))
    message_context.tags['rule-value'] = xule_value
    message_context.tags['alignment'] = func_alignment(xule_context)
    message_context.result_alignment = func_alignment(xule_context)

    try:
        # Caching does not work for expressions with tagRefs. The The results portion of a rule will have a tagRef for each varRef. This conversion is
        # done during the post parse step. So it is neccessary to turn local caching off when evaluating the result expression. There is a command line option
        # for doing this. This code will turn this command line option on.
        saved_no_cache = getattr(message_context.global_context.options, 'xule_no_cache', False)
        #         if hasattr(message_context.global_context.options, 'xule_no_cache'):
        #             xule_context.global_context.options.xule_no_cache = True
        xule_context.global_context.options.xule_no_cache = True

        message_value = evaluate(result_ast['resultExpr'], message_context)
    except XuleIterationStop as xis:
        raise XuleProcessingError(_("Cannot produce message. An expression in the message has a skip value."),
                                  xule_context)
    finally:
        #         if hasattr(message_context.global_context.options, 'xule_no_cache'):
        #
        xule_context.global_context.options.xule_no_cache = saved_no_cache

    if result_ast['resultName'] == 'rule-focus':
        # This is a special case. rule-focus requires some kind of a ModelObject. This will be passed to the logger as the modelObject argument.
        # Xule will allow a ModelFact or a ModelConcept
        if message_value.type == 'concept':
            message = message_value.value
        elif message_value.is_fact:
            message = message_value.fact
        elif message_value.type in ('list','set'):
            # The rule focus is a list/set of concepts or facts. The list/set cannot be nested
            message = []
            for rule_focus_item in message_value.value:
                if rule_focus_item.type == 'concept':
                    message.append(rule_focus_item.value)
                elif rule_focus_item.is_fact:
                    message.append(rule_focus_item.fact)
                elif rule_focus_item.type in ('unbound', 'none'):
                    message.append(None)
                else:
                    raise XuleProcessingError(
                        _("The rule-focus of a rule must be a concept or a fact, found {}".format(rule_focus_item.type)),
                        xule_context)
        elif message_value.type in ('unbound', 'none'):
            message = None
        else:
            raise XuleProcessingError(
                _("The rule-focus of a rule must be a concept or a fact, found {}".format(message_value.type)),
                xule_context)
    elif result_ast['resultName'] == 'message':
        if message_value.type == 'unbound':
            message = XuleValue(xule_context, '', 'string')
        else:
            message = message_value
    else:
        if message_value.type == 'unbound':
            message = ''
        else:
            message = str(message_value.value)

    return message


def validate_result_name(result, xule_context):
    if result['resultName'] not in ('message', 'severity', 'rule-suffix', 'rule-focus'):
        if not xule_context.rule_set.hasOutputAttribute(result['resultName']):
            raise XuleProcessingError(_(
                "Rule '{}' uses result name '{}' which does not have an output-attribute declaration.".format(
                    xule_context.rule_name, result['resultName'])))


def get_all_aspects(model_fact, xule_context):
    '''This function gets all the apsects of a fact'''
    return get_alignment(model_fact, {}, {}, xule_context)


def get_alignment(model_fact, non_align_aspects, align_aspects, xule_context, covered_dims=False, covered=False, dimensions_special_covered=False):
    '''The alignment contains the aspect/member pairs that are in the fact but not in the non_align_aspects.
       The alignment is done in two steps. First check each of the builtin aspects. Then check the dimesnions.'''

    '''non_align_aspect - dictionary
            Key is a tuple with the following parts:
                0 = TYPE -'builtin' or 'explicit_dimension', 
                1 = ASPECT - aspect, 
                2 = SPECIALVALUE - 'all' or 'allWithDefault', 
            value = MEMBER (None if there is a SPECIAL_VALUE)'''

    alignment = {}
    # builtin alignment
    if covered:
        # Don't need the non_align_builtins, so don't bother creating them.
        non_align_builtins = None
    else:
        non_align_builtins = {aspect_info[ASPECT] for aspect_info in non_align_aspects if aspect_info[TYPE] == 'builtin'}

    align_builtins = {aspect_info[ASPECT] for aspect_info in align_aspects if aspect_info[TYPE] == 'builtin'}

    # Only need to go through the builtins if they are not covered or they are covered and there are
    # aligned builtins
    if not covered or len(align_builtins) > 0:

        # lineItem
        if (not covered and 'concept' not in non_align_builtins) or 'concept' in align_builtins:
            alignment[('builtin', 'concept')] = model_fact.qname
            # alignment[('builtin', 'concept')] = model_fact.elementQname

        # unit
        if model_fact.isNumeric:
            if (not covered and 'unit' not in non_align_builtins) or 'unit' in align_builtins:
                alignment[('builtin', 'unit')] = model_to_xule_unit(model_fact.unit, xule_context)

        # period
        if (not covered and 'period' not in non_align_builtins) or 'period' in align_builtins:
            alignment[('builtin', 'period')] = model_to_xule_period(model_fact.context, xule_context)

        # entity
        if (not covered and 'entity' not in non_align_builtins) or 'entity' in align_builtins:
            alignment[('builtin', 'entity')] = model_to_xule_entity(model_fact.context, xule_context)

    # dimensional apsects
    if covered_dims or covered:
        # Non algined dimensions don't matter.
        non_align_dimesnions = set()
    else:
        non_align_dimensions = {aspect_info[ASPECT] for aspect_info in non_align_aspects if
                                aspect_info[TYPE] == 'explicit_dimension'}

    align_dimensions = {aspect_info[ASPECT] for aspect_info in align_aspects if
                                aspect_info[TYPE] == 'explicit_dimension'}

    # Only need to run through the dimensions if they are included or if they are not included there are
    # aligned dimensions
    if (not (covered_dims or dimensions_special_covered == 'covered') and not covered) or len(align_dimensions) > 0:
        for fact_dimension_qname, dimension_value in model_fact.context.qnameDims.items():
            if (not (covered_dims or dimensions_special_covered == 'covered') and not covered and fact_dimension_qname not in non_align_dimensions) or fact_dimension_qname in align_dimensions:
                alignment[('explicit_dimension', # This will included typed dimensions as well as explicit.
                           fact_dimension_qname)] = dimension_value.memberQname if dimension_value.isExplicit else dimension_value.typedMember.xValue

    return alignment

def get_uncommon_aspects(model_fact, common_aspects, xule_context):
    uncommon_aspects = {}

    fact_aspects = get_all_aspects(model_fact, xule_context)
    for aspect_info, aspect_value in fact_aspects.items():
        if aspect_info == ('builtin', 'concept'):
            uncommon_aspects[aspect_info] = aspect_value
        elif aspect_info not in common_aspects:
            uncommon_aspects[aspect_info] = aspect_value

    return uncommon_aspects


def format_alignment(aspects, xule_context):
    if len(aspects) == 0:
        return ''

    aspect_strings = []
    line_item_string = ""

    # built in aspects
    if ('builtin', 'concept') in aspects:
        line_item_string = format_qname(aspects[('builtin', 'concept')], xule_context)

    if ('builtin', 'period') in aspects:
        period_info = aspects[('builtin', 'period')]
        if isinstance(period_info, tuple):
            if period_info[0] == datetime.datetime.min and period_info[1] == datetime.datetime.max:
                aspect_strings.append("period=forever")
            else:
                aspect_strings.append("period=duration('%s', '%s')" % (period_info[0].strftime("%Y-%m-%d"), (
                            period_info[1] - datetime.timedelta(days=1)).strftime("%Y-%m-%d")))
        else:
            aspect_strings.append(
                "period=instant('%s')" % (period_info - datetime.timedelta(days=1)).strftime("%Y-%m-%d"))

    if ('builtin', 'unit') in aspects:
        model_unit = aspects[('builtin', 'unit')]
        aspect_strings.append(str(model_unit))
    #         if len(model_unit[1]) == 0:
    #             #no denominator
    #             aspect_strings.append("unit=%s" % " * ".join([x.localName for x in model_unit[0]]))
    #         else:
    #             aspect_strings.append("unit=%s/%s" % (" * ".join([x.localName for x in model_unit[0]]),
    #                                                   " * ".join([x.localName for x in model_unit[1]])))

    if ('builtin', 'entity') in aspects:
        entity_info = aspects[('builtin', 'entity')]
        aspect_strings.append("entity=(%s) %s" % (entity_info[0], entity_info[1]))

        # dimensions
    dimension_aspects = [(aspect_info[ASPECT], aspect_info, aspect_member) for aspect_info, aspect_member in
                         aspects.items() if aspect_info[TYPE] == 'explicit_dimension']
    # sort by the dimension qname
    dimension_aspects.sort(key=lambda tup: tup[0])

    # for aspect_info, aspect_member in aspects.items():
    for dimension_aspect in dimension_aspects:
        aspect_info = dimension_aspect[1]
        aspect_member = dimension_aspect[2]
        if aspect_info[TYPE] == 'explicit_dimension':
            aspect_member = aspects[aspect_info]
            '''THE formatted_member SHOULD HANDLE FORMATTING OF NON QNAME VALUES'''
            formatted_member = format_qname(aspect_member, xule_context) if type(
                aspect_member) == QName else aspect_member
            aspect_strings.append("%s=%s" % (format_qname(aspect_info[ASPECT], xule_context), formatted_member))

    if len(aspect_strings) > 0:
        aspect_string = "[" + ",\n".join(aspect_strings) + "]"
    else:
        aspect_string = ""

    return line_item_string + aspect_string


def format_fact_line_item(xule_context, xule_fact):
    return format_qname(xule_fact.fact.concept.qname, xule_context)


def format_fact_period(xule_context, xule_fact):
    if xule_fact.fact.context.isStartEndPeriod:
        period_string = xule_fact.fact.context.startDatetime.strftime("%m/%d/%Y") + " - " + (
                    xule_fact.fact.context.endDatetime - datetime.timedelta(days=1)).strftime("%m/%d/%Y")
    elif xule_fact.fact.context.isInstantPeriod:
        period_string = (xule_fact.fact.context.endDatetime - datetime.timedelta(days=1)).strftime("%m/%d/%Y")
    else:
        period_string = "Forever"

    return period_string


def format_fact_unit(xule_context, xule_fact):
    if xule_fact.fact.isNumeric:
        numerator = tuple(sorted(xule_fact.fact.unit.measures[0]))
        denominator = tuple(sorted(xule_fact.fact.unit.measures[1]))

        if len(denominator) == 0:
            # no denominator
            return " * ".join([x.localName for x in numerator])
        else:
            return "%s/%s" % (" * ".join([x.localName for x in numerator]),
                              " * ".join([x.localName for x in denominator]))
    else:
        return None


def format_fact_all_aspects(xule_context, xule_fact):
    aspect_strings = ["Line Item: " + format_fact_line_item(xule_context, xule_fact), ]
    aspect_strings.append("Period: " + format_fact_period(xule_context, xule_fact))
    unit = format_fact_unit(xule_context, xule_fact)
    if unit is not None:
        aspect_strings.append("Unit: " + unit)
    dimensions = format_fact_dimensions(xule_context, xule_fact)
    if dimensions is not None:
        aspect_strings.append("Dimensions:\n" + dimensions)
    return "\n".join(aspect_strings)


def format_fact_dimensions(xule_context, xule_fact):
    if len(xule_fact.fact.context.qnameDims) > 0:
        dim_pairs = []
        for axis_qname in sorted(xule_fact.fact.context.qnameDims):
            dim_pairs.append(format_qname(axis_qname, xule_context) + " = " + format_qname(
                xule_fact.fact.context.qnameDims[axis_qname].memberQname, xule_context))
        return "\n".join(dim_pairs)
    else:
        return None


def format_fact_label(xule_context, fact):
    label = XuleProperties.property_label(xule_context, fact)
    if label.type in ('unbound', 'none'):
        return "missing"
    else:
        return label.value.textValue


def format_qname(qname, xule_context):
    cat_namespace = xule_context.rule_set.getNamespaceInfoByUri(qname.namespaceURI)
    if cat_namespace:
        if cat_namespace['prefix'] == '*':
            return qname.localName
        else:
            return cat_namespace['prefix'] + ":" + qname.localName
    else:
        return str(qname)


MESSAGE_TAG_SUB_PARTS = (('context', format_alignment),
                         ('label', format_fact_label),
                         ('concept', format_fact_line_item),
                         ('period', format_fact_period),
                         ('unit', format_fact_unit),
                         ('aspects', format_fact_all_aspects),
                         ('dimensions', format_fact_dimensions)
                         )


def write_trace_count_string(trace_count_file, rule_name, traces, rule_part, total_iterations, total_time):
    display_string = display_trace_count(traces, rule_part, total_iterations, total_time)

    with open(trace_count_file + ".txt", 'a', newline='') as o:
        o.write(display_string)


def display_trace_count(traces, rule_part, total_iterations, total_time, level=0, display_string=""):
    if isinstance(rule_part, ParseResults):
        additional = ''
        if rule_part['exprName'] == 'varRef':
            if rule_part.is_constant:
                additional = " ($" + rule_part.varName + " declaration: " + str(
                    rule_part.var_declaration) + ") - constant)"
            else:
                additional = " ($" + rule_part.varName + " declaration: " + str(rule_part.var_declaration) + ")"
        elif rule_part['exprName'] == 'varAssign':
            additional = " (" + rule_part.varName + ")" + (" NOT USED" if rule_part.get('not_used') == True else "")
        elif rule_part['exprName'] == 'constantAssign':
            additional = " (" + rule_part.constantName + ")"
        elif rule_part['exprName'] == 'functionArg':
            additional = " (" + rule_part.argName + ")"
        elif rule_part['exprName'] == 'forExpr':
            additional = " (" + rule_part.forControl.forVar + ")"
        elif rule_part['exprName'] == 'reportDeclaration':
            additional = " (" + rule_part.reportName + ")"
        elif rule_part['exprName'] == 'raiseDeclaration':
            additional = " (" + rule_part.raiseName + ")"
        elif rule_part['exprName'] == 'formulaDeclaration':
            additional = " (" + rule_part.formulaName + ")"
        #         elif rule_part['exprName'] in ('functionDeclaration', 'functionReference'):
        #             additional = " (" + rule_part.functionName + ")"  + (" CACHEABLE" if rule_part.get('cacheable') == True else "")

        if rule_part.number != '':
            additional += (" [" +
                           ('i, ' if rule_part.get('instance') else '') +
                           ('r, ' if rule_part.get('rules-taxonomy') else '') +
                           ('1' if rule_part.number == 'single' else "*") +
                           (", Align" if rule_part.has_alignment else ", NoAlign") +
                           ((", " + ("D" if rule_part.is_dependent else "I")) if 'is_dependent' in rule_part else "") +
                           # ((", " + str(parse_res.var_refs)) if len(parse_res.var_refs) > 0 else "") +
                           #                            ((", v" + str({(x[0], x[1]) for x in rule_part.dependent_vars})) if len(rule_part.dependent_vars) > 0 else "") +
                           #                            ((", V" + str({(x[0], x[1]) for x in rule_part.var_refs})) if len(rule_part.var_refs) > 0 else "") +
                           #                            ((", VIds" + str(rule_part.var_ref_ids)) if 'var_ref_ids' in rule_part else "")  +
                           ((", i" + str({dep['node_id'] for dep in rule_part.dependent_iterables})) if len(
                               rule_part.dependent_iterables) > 0 else "") +
                           #                            ((", di" + str({dep['node_id'] for dep in rule_part.downstream_iterables})) if len(rule_part.downstream_iterables) > 0 else "") +
                           (", Values" if rule_part.get('values_expression') == True else "") +
                           (", Table %i" % rule_part.table_id if rule_part.get('table_id') is not None else "") +
                           "]")

        if 'is_iterable' in rule_part:
            additional += " iterable"
        if 'in_loop' in rule_part:
            additional += " LOOP"
        display_string += ("  " * level) + str(rule_part['node_id']) + ":" + rule_part['exprName'] + additional + "\n"
        if rule_part['node_id'] in traces:
            trace = traces[rule_part['node_id']]
            total_count = 0
            # display_string += ", ".join(trace.keys()) + "\n"
            for key in ('iterations', 'U', 'E', 'c', 'T', 'e', 'R', 'r', 'isE', 'ise', 'isu', 'ex'):
                if trace[key] > 0:
                    if key == 'iterations':
                        display_string += "{}{} {} {}\n".format("  " * (level + 1),
                                                                key,
                                                                trace[key],
                                                                (trace[
                                                                     key] / total_iterations) if total_iterations > 0 else 0)
                        # add step values
                        children_time, child_nodes = trace_count_next_time(rule_part, traces)
                        step_time = trace['iterations-t'] - children_time
                        display_string += "{}{} {} {} {}\n".format("  " * (level + 1),
                                                                   "Step",
                                                                   step_time.total_seconds(),
                                                                   (
                                                                               step_time / total_time) if total_time.total_seconds() > 0 else 0,
                                                                   str(child_nodes)[1:-1])
                    else:
                        try:
                            display_string += "{}{} {} - Avg: {}  Tot: {} - Avg: {:%}  Tot: {:%}\n".format(
                                "  " * (level + 2),
                                key,
                                trace[key],
                                trace[key + '-t'].total_seconds() / trace[key] if trace[key] > 0 else 0,
                                trace[key + '-t'].total_seconds(),
                                ((trace[key + '-t'].total_seconds() / trace[key] if trace[
                                                                                        key] > 0 else 0) / total_iterations) if total_iterations > 0 else 0,
                                (trace[key + '-t'] / total_time) if total_time.total_seconds() > 0 else 0)
                        except:
                            # TBD change to xule_context.global_context.message_queue.print
                            print("key", key, "key time", trace[key + '-t'])
                            raise
                    if key != 'iterations':
                        total_count += trace[key]
            if total_count != trace['iterations']:
                display_string += "%sCalc Total %i\n" % ("  " * (level + 1), total_count)
            display_string += "%sTime %f Average %f\n\n" % ("  " * (level + 1), trace['iterations-t'].total_seconds(), (
                        trace['iterations-t'].total_seconds() / total_count) if total_count > 0 else 0)
        for next_part in rule_part:
            display_string = display_trace_count(traces, next_part, total_iterations, total_time, level + 1,
                                                 display_string)
    return display_string


def write_trace_count_csv(trace_count_file, rule_name, traces, rule_part, total_iterations, total_time):
    import csv

    trace_count = calc_trace_count(rule_name, traces, rule_part, total_iterations, total_time)
    with open(trace_count_file + ".csv", 'a', newline='') as o:
        csv_writer = csv.writer(o)
        csv_writer.writerows(trace_count)


def calc_trace_count(rule_name, traces, rule_part, total_iterations, total_time, level=0, rows=None):
    if rows is None:
        rows = [['', '', '', '', '', '', '', '', '', '', ''
                    , 'Total', '', '', '', '', ''  # total iteraions
                    , 'Step', '', '', '', ''  # Step time
                    , 'Evaluations', '', '', '', '', ''  # Evaluations
                    , 'Table', '', '', '', '', ''  # Table
                    , 'Cache', '', '', '', '', ''  # cache
                    , 'Recalc', '', '', '', '', ''  # Recalc
                    , 'Stop', '', '', '', '', ''  # Iteration Stop
                 ],
                ['Rule', 'Id', 'Name', 'Notes', 'Instance', 'Rule Taxonomy', 'Number', 'Aligned', 'Dependent',
                 'Dependent Iterables', 'Iterable'
                    , 'it', 'it %', 'secs', 'secs %', 'avg', 'avg %'  # total iteraions
                    , 'secs', 'secs %', 'avg', 'avg %', 'nodes'  # step times
                    , 'it', 'it %', 'secs', 'secs %', 'avg', 'avg %'  # Evaluations
                    , 'it', 'it %', 'secs', 'secs %', 'avg', 'avg %'  # Table
                    , 'it', 'it %', 'secs', 'secs %', 'avg', 'avg %'  # cache
                    , 'it', 'it %', 'secs', 'secs %', 'avg', 'avg %'  # Recalc
                    , 'it', 'it %', 'secs', 'secs %', 'avg', 'avg %'  # Iteration Stop
                 ]]

    # Rows: name, notes, inst, rule tax, number, aligned, dependency, dependent iterables, iterable,
    #      For each count includes: iterations, percent, time, percent, avg time, percent
    #          total, E, T, c, R, is

    if isinstance(rule_part, ParseResults):
        additional = ''
        if rule_part['exprName'] == 'varRef':
            if rule_part.is_constant:
                additional = "$" + rule_part.varName + " declaration: " + str(
                    rule_part.var_declaration) + " - constant)"
            else:
                additional = "$" + rule_part.varName + " declaration: " + str(rule_part.var_declaration)
        elif rule_part['exprName'] == 'varAssign':
            additional = rule_part.varName + (" NOT USED" if rule_part.get('not_used') == True else "")
        elif rule_part['exprName'] == 'constantAssign':
            additional = rule_part.constantName
        elif rule_part['exprName'] == 'functionArg':
            additional = rule_part.argName
        elif rule_part['exprName'] == 'forExpr':
            additional = rule_part.forControl.forVar
        elif rule_part['exprName'] == 'reportDeclaration':
            additional = rule_part.reportName
        elif rule_part['exprName'] == 'raiseDeclaration':
            additional = rule_part.raiseName
        elif rule_part['exprName'] == 'formulaDeclaration':
            additional = rule_part.formulaName
        elif rule_part['exprName'] in ('functionDeclaration', 'functionReference'):
            additional = rule_part.functionName + (" CACHEABLE" if rule_part.get('cacheable') == True else "")

        row = [rule_name
            , rule_part['node_id']
            , ('  ' * level) + rule_part['exprName']  # name
            , additional  # notes
            , True if rule_part.get('instance') else False  # instance
            , True if rule_part.get('rules-taxonomy') else False  # rule taxonomy
            , rule_part.number == 'single'  # nubmer
            , True if rule_part.has_alignment else False  # aligned
            , ("D" if rule_part.is_dependent else "I") if 'is_dependent' in rule_part else ""  # dependency
            , str({dep['node_id'] for dep in rule_part.dependent_iterables}) if len(
                rule_part.dependent_iterables) > 0 else ""  # dependent iterables
            , 'is_iterable' in rule_part  # iterable
               ]

        if rule_part['node_id'] in traces:
            trace = traces[rule_part['node_id']]

            # add total values to the row
            row += trace_count_by_type('iterations', trace, total_iterations, total_time)

            # add step values
            children_time, child_nodes = trace_count_next_time(rule_part, traces)
            step_time = trace['iterations-t'] - children_time
            row += [step_time.total_seconds()
                , ((
                               step_time.total_seconds() / total_time.total_seconds()) if total_time.total_seconds() > 0 else 0) * 100
                , (step_time.total_seconds() / trace['iterations']) if trace['iterations'] > 0 else 0
                , ((((step_time.total_seconds() / trace['iterations']) if trace[
                                                                              'iterations'] > 0 else 0) / total_iterations) if total_iterations > 0 else 0) * 100
                , str(child_nodes)[1:-1]]

            # row += ['','','','','']

            # add values by evaluation type
            calc_total = 0
            for count_codes in [['E', 'e'], ['T'], ['c'], ['R', 'r'], ['isE', 'ise']]:
                if len(count_codes) == 1:
                    count_code = count_codes[0]
                else:
                    count_code = count_codes[0] if trace[count_codes[0]] != 0 else count_codes[1]
                calc_total += trace[count_code]
                row += trace_count_by_type(count_code, trace, total_iterations, total_time)

            if calc_total != trace['iterations']:
                row.append("ITERATION COUNT DOES NOT TOTAL. Calc total is", calc_total)

        rows.append(row)

        for next_part in rule_part:
            rows = calc_trace_count(rule_name, traces, next_part, total_iterations, total_time, level + 1, rows)
    return rows


def trace_count_by_type(count_code, trace, total_iterations, total_time):
    time_code = count_code + '-t'
    return [trace[count_code]
        , ''
        , trace[time_code].total_seconds()
        , (trace[time_code].total_seconds() / total_time.total_seconds() * 100) if total_time.total_seconds() > 0 else 0
        , (trace[time_code].total_seconds() / trace[count_code]) if trace[count_code] > 0 else 0
        , (((trace[time_code].total_seconds() / trace[count_code] if trace[
                                                                         count_code] > 0 else 0) / total_iterations) if total_iterations > 0 else 0) * 100
            ]


def trace_count_next_time(rule_part, traces):
    total_child_times = datetime.timedelta()
    total_child_nodes = []

    if isinstance(rule_part, ParseResults):
        for child in rule_part:

            if isinstance(child, ParseResults):
                if child['exprName'] != 'varAssign':
                    if child['node_id'] in traces:
                        total_child_times += traces[child['node_id']]['iterations-t']
                        total_child_nodes.append(child['node_id'])
                        # return (traces[child['node_id']]['iterations-t'], child['node_id'])
                    else:
                        child_info = trace_count_next_time(child, traces)
                        total_child_times += child_info[0]
                        total_child_nodes += child_info[1]
    return (total_child_times, total_child_nodes)

def fact_is_complete(model_fact):
    if model_fact.xValid < VALID:
        return False
    if model_fact.context is None or not context_has_period(model_fact.context):
        return False
    if model_fact.isNumeric and model_fact.unit is None:
        return False
    return True


def context_has_period(model_context):
    return model_context.isStartEndPeriod or model_context.isInstantPeriod or model_context.isForeverPeriod